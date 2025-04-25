"""
Combined QA Automation Script

This script combines the execution of multiple QA scripts into a single process:
1. test_api.py - Generates the main QA report with 3 tabs
2. qa_flight_v3.py - Validates flight dates 
3. name_assign.py - Checks naming conventions
4. targeting.py - Validates targeting settings
5. creative.py - Validates creative settings

All outputs are combined into a single Excel file with multiple tabs while
preserving the original formatting and styling from each individual script.

Environment variables (from beeswax_input_qa.env):
- ENV_PATH: Path to the environment file (default: ./input_folder/beeswax_input_qa.env)
- OUTPUT_DIR: Directory for output files (default: ./output_folder)
- OUTPUT_RAW_DIR: Directory for intermediate output files (default: ./output_raw)
- COMBINED_OUTPUT_PATH: Path for the combined output file (default: ./output_folder/combined_qa_report_{timestamp}.xlsx)
- BRIEF_PATH: Path to the campaign brief file
- QA_REPORT_PATH: Path to the QA report file (if already exists)
- QA_FLIGHT_OUTPUT_PATH: Path for qa_flight_v3.py output
- NAME_ASSIGN_OUTPUT_PATH: Path for name_assign.py output
- TARGETING_OUTPUT_PATH: Path for targeting.py output
- CREATIVE_OUTPUT_PATH: Path for creative.py output
"""

import os
import sys
import glob
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import importlib.util
from dotenv import load_dotenv
import subprocess

# Set working directory
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

def load_module_from_file(module_name, file_path):
    """
    Dynamically load a module from a file path
    """
    print(f"Loading module {module_name} from {file_path}")
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module

def find_latest_file(pattern):
    """
    Find the latest file matching the given pattern
    """
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def run_script_and_get_output(script_name, module_name, script_func=None):
    """
    Run a script either by importing its main function or by executing it as a subprocess
    Returns the path to the output file
    """
    print(f"\n{'='*80}\nRunning {script_name}...\n{'='*80}")
    
    # If a specific function is provided, use that
    if script_func:
        output_file = script_func()
        print(f"{script_name} completed, output file: {output_file}")
        return output_file
    
    # Try to import and run the main function
    try:
        module = load_module_from_file(module_name, script_name)
        if hasattr(module, 'main'):
            output_file = module.main()
            print(f"{script_name} completed, output file: {output_file}")
            return output_file
    except Exception as e:
        print(f"Could not import {script_name} as a module: {e}")
    
    # Fallback: Run as subprocess
    try:
        subprocess.run([sys.executable, script_name], check=True)
        print(f"{script_name} completed via subprocess")
        return None  # We'll need to find the output file separately
    except subprocess.CalledProcessError as e:
        print(f"Error running {script_name}: {e}")
        return None

def run_test_api():
    """
    Run the test_api.py script and return the output file path
    """
    # Check if a specific QA report is already provided in env vars
    qa_report_path = os.environ.get("QA_REPORT_PATH")
    if qa_report_path and os.path.exists(qa_report_path):
        print(f"Using existing QA report from environment variables: {qa_report_path}")
        return qa_report_path
    
    try:
        # Import test_api.py and run
        test_api_module = load_module_from_file("test_api", "test_api.py")
        
        # Get parameters from environment variables
        brief_path = os.environ.get("BRIEF_PATH")
        env_path = os.environ.get("ENV_PATH")
        output_dir = os.environ.get("OUTPUT_DIR")
        
        # Create an instance of BeeswaxQA with environment parameters
        if brief_path and env_path and output_dir:
            qa = test_api_module.BeeswaxQA(brief_path, env_path, output_dir)
        elif brief_path and output_dir:
            qa = test_api_module.BeeswaxQA(brief_path, output_dir=output_dir)
        elif brief_path:
            qa = test_api_module.BeeswaxQA(brief_path)
        else:
            qa = test_api_module.BeeswaxQA()
        
        # Generate the QA report
        output_file = qa.generate_qa_report()
        
        print(f"test_api.py completed, output file: {output_file}")
        return output_file
    
    except Exception as e:
        print(f"Error running test_api.py: {e}")
        
        # Fallback: Run as subprocess
        subprocess.run([sys.executable, "test_api.py"], check=True)
        
        # Find the latest QA report in the output directory
        output_dir = os.environ.get("OUTPUT_DIR", "./output_folder")
        latest_file = find_latest_file(os.path.join(output_dir, "qa_report_*.xlsx"))
        
        print(f"Found latest QA report: {latest_file}")
        return latest_file

def copy_cell_format(source_cell, target_cell):
    """
    Safely copy cell formatting properties to avoid StyleProxy issues
    """
    # Copy font properties individually
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    
    # Copy fill
    if source_cell.fill:
        if source_cell.fill.fill_type:
            # Only copy if there's an actual fill
            fill_color = source_cell.fill.start_color.rgb if source_cell.fill.start_color else None
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=fill_color,
                end_color=source_cell.fill.end_color.rgb if source_cell.fill.end_color else None
            )
    
    # Copy alignment
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )
    
    # Copy border
    if source_cell.border:
        sides = {}
        for side in ['left', 'right', 'top', 'bottom']:
            side_obj = getattr(source_cell.border, side)
            if side_obj and side_obj.style:
                sides[side] = Side(
                    style=side_obj.style,
                    color=side_obj.color.rgb if side_obj.color else None
                )
            else:
                sides[side] = Side(style=None)
        
        target_cell.border = Border(**sides)
    
    # Copy number format
    target_cell.number_format = source_cell.number_format

def create_combined_report(qa_report_path, other_outputs):
    """
    Create a combined report with all the tabs from individual scripts
    Preserves original formatting
    
    Args:
        qa_report_path: Path to the QA report from test_api.py
        other_outputs: Dictionary mapping script names to their output file paths
    
    Returns:
        Path to the combined report
    """
    print("\nCreating combined QA report...")
    
    # Create timestamp for the output file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Get output path from environment or use default
    combined_output_path = os.environ.get("COMBINED_OUTPUT_PATH")
    if combined_output_path:
        # Replace timestamp placeholder if present
        if "{timestamp}" in combined_output_path:
            combined_output_path = combined_output_path.replace("{timestamp}", timestamp)
    else:
        # Define the output directory and file path
        output_dir = os.environ.get("OUTPUT_DIR", "./output_folder")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        combined_output_path = os.path.join(output_dir, f"combined_qa_report_{timestamp}.xlsx")
    
    # Ensure the directory exists
    os.makedirs(os.path.dirname(os.path.abspath(combined_output_path)), exist_ok=True)
    
    # Start with the QA report workbook
    print(f"Loading QA report from {qa_report_path}")
    qa_workbook = openpyxl.load_workbook(qa_report_path)
    
    # Create a new workbook for the combined report
    combined_wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = combined_wb.active
    combined_wb.remove(default_sheet)
    
    # First, copy all sheets from the QA report
    print("Copying sheets from QA report...")
    for sheet_name in qa_workbook.sheetnames:
        # Get the source sheet
        source_sheet = qa_workbook[sheet_name]
        
        # Create a new sheet in the combined workbook
        new_sheet = combined_wb.create_sheet(title=sheet_name)
        
        # Copy the data and formatting
        for row in source_sheet.rows:
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                # Copy cell formatting safely
                if cell.has_style:
                    copy_cell_format(cell, new_cell)
        
        # Copy column dimensions
        for col_letter, dimension in source_sheet.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = dimension.width
        
        # Copy row dimensions
        for row_number, dimension in source_sheet.row_dimensions.items():
            new_sheet.row_dimensions[row_number].height = dimension.height
            
        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
    
    # Define standardized tab names for each script
    script_tab_names = {
        "qa_flight_v3.py": "QA_Dates",
        "name_assign.py": "QA_Name_Assign",
        "targeting.py": "QA_Targeting",
        "creative.py": "QA_Creative"
    }
    
    # Now copy sheets from other output files
    for script_name, output_file in other_outputs.items():
        if not output_file or not os.path.exists(output_file):
            print(f"Warning: Output file for {script_name} not found at {output_file}")
            continue
        
        print(f"Copying sheets from {script_name} output: {output_file}")
        
        # Load the workbook
        try:
            source_wb = openpyxl.load_workbook(output_file)
            
            # Use standardized prefix if available, otherwise use the script name
            script_prefix = script_tab_names.get(script_name, script_name.replace(".py", "").replace("_", " ").title())
            
            # Copy each sheet with a standardized name
            for idx, sheet_name in enumerate(source_wb.sheetnames):
                source_sheet = source_wb[sheet_name]
                
                # Create a unique name for the new sheet
                # If there's only one sheet, use just the prefix; otherwise add a suffix
                if len(source_wb.sheetnames) == 1:
                    new_sheet_name = script_prefix
                else:
                    # Extract a short suffix from the original sheet name or use an index
                    if idx == 0:
                        new_sheet_name = script_prefix  # First sheet gets no suffix
                    else:
                        # Create a suffix based on position or extract a word from the sheet name
                        sheet_words = sheet_name.replace("Check", "").replace("Result", "").strip()
                        if len(sheet_words) > 0:
                            new_sheet_name = f"{script_prefix}_{sheet_words}"
                        else:
                            new_sheet_name = f"{script_prefix}_{idx+1}"
                
                # Ensure name is valid and not too long
                new_sheet_name = new_sheet_name[:31]  # Excel sheet names limited to 31 chars
                
                # If the sheet name already exists, append a number to make it unique
                base_name = new_sheet_name
                suffix_num = 1
                while new_sheet_name in combined_wb.sheetnames:
                    new_sheet_name = f"{base_name}_{suffix_num}"
                    suffix_num += 1
                    if len(new_sheet_name) > 31:
                        # Truncate if too long after adding suffix
                        new_sheet_name = f"{base_name[:27]}_{suffix_num}"
                
                # Create a new sheet in the combined workbook
                print(f"  Creating sheet '{new_sheet_name}'")
                new_sheet = combined_wb.create_sheet(title=new_sheet_name)
                
                # Copy the data and formatting
                for row in source_sheet.rows:
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                        new_cell.value = cell.value
                        
                        # Copy cell formatting safely
                        if cell.has_style:
                            copy_cell_format(cell, new_cell)
                
                # Copy column dimensions
                for col_letter, dimension in source_sheet.column_dimensions.items():
                    new_sheet.column_dimensions[col_letter].width = dimension.width
                
                # Copy row dimensions
                for row_number, dimension in source_sheet.row_dimensions.items():
                    new_sheet.row_dimensions[row_number].height = dimension.height
                    
                # Copy merged cells
                for merged_range in source_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_range))
                
                # Copy conditional formatting if possible
                try:
                    if hasattr(source_sheet, 'conditional_formatting'):
                        for cf_range, cf_rules in source_sheet.conditional_formatting.items():
                            for rule in cf_rules:
                                try:
                                    new_sheet.conditional_formatting.add(cf_range, rule)
                                except Exception as e:
                                    print(f"Warning: Could not copy conditional formatting: {e}")
                except Exception as e:
                    print(f"Warning: Issue with conditional formatting: {e}")
                
        except Exception as e:
            print(f"Error copying sheets from {output_file}: {e}")
    
    # Save the combined workbook
    print(f"Saving combined workbook to {combined_output_path}")
    combined_wb.save(combined_output_path)
    print(f"Combined QA report saved to: {combined_output_path}")
    
    return combined_output_path

def ensure_output_paths():
    """
    Make sure all output directories exist based on environment variables
    """
    # Get environment paths
    env_path = os.environ.get("ENV_PATH", "./input_folder/beeswax_input_qa.env")
    # Load environment variables
    if os.path.exists(env_path):
        load_dotenv(env_path)
    
    # Get output directories from environment variables
    output_dirs = [
        os.environ.get("OUTPUT_DIR", "./output_folder"),
        os.environ.get("OUTPUT_RAW_DIR", "./output_raw"),
    ]
    
    # Add directories from file paths
    file_paths = [
        os.environ.get("QA_FLIGHT_OUTPUT_PATH", "./qa_flight_v3_output.xlsx"),
        os.environ.get("NAME_ASSIGN_OUTPUT_PATH", "./name_assign_output.xlsx"),
        os.environ.get("TARGETING_OUTPUT_PATH", "./targeting_check_output.xlsx"),
        os.environ.get("CREATIVE_OUTPUT_PATH", "./creative_qa_output.xlsx"),
        os.environ.get("COMBINED_OUTPUT_PATH", "./output_folder/combined_qa_report.xlsx")
    ]
    
    # Extract directories from file paths
    for file_path in file_paths:
        if file_path:
            dir_path = os.path.dirname(file_path)
            if dir_path:
                output_dirs.append(dir_path)
    
    # Create directories if they don't exist
    for directory in output_dirs:
        if directory and not os.path.exists(directory):
            print(f"Creating output directory: {directory}")
            os.makedirs(directory)

def main():
    """
    Main function to run all QA scripts and combine their outputs
    """
    # Start time
    start_time = time.time()
    
    # Load environment variables
    env_path = os.environ.get("ENV_PATH", "./input_folder/beeswax_input_qa.env")
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        load_dotenv(env_path)
    else:
        print(f"Warning: Environment file {env_path} not found. Using default environment.")
        load_dotenv()  # Try default locations
    
    # Make sure output directories exist
    ensure_output_paths()
    
    # 1. Run test_api.py first
    print("\nStep 1: Running test_api.py...")
    qa_report_path = run_test_api()
    
    if not qa_report_path or not os.path.exists(qa_report_path):
        print("Error: Failed to generate QA report from test_api.py")
        return
    
    print(f"QA report generated: {qa_report_path}")
    
    # Set the QA report path in environment for other scripts
    os.environ["QA_REPORT_PATH"] = qa_report_path
    
    # Dictionary to store output paths for each script
    output_files = {}
    
    # 2. Run qa_flight_v3.py
    print("\nStep 2: Running qa_flight_v3.py...")
    qa_flight_output = run_script_and_get_output("qa_flight_v3.py", "qa_flight_v3")
    if not qa_flight_output:
        # Try to find the output file
        qa_flight_output = os.environ.get("QA_FLIGHT_OUTPUT_PATH", "./qa_flight_v3_output.xlsx")
        if not os.path.exists(qa_flight_output):
            output_raw_dir = os.environ.get("OUTPUT_RAW_DIR", "./output_raw")
            qa_flight_output = find_latest_file(os.path.join(output_raw_dir, "qa_flight_v3_output*.xlsx"))
    
    output_files["qa_flight_v3.py"] = qa_flight_output
    print(f"QA Flight output: {qa_flight_output}")
    
    # 3. Run name_assign.py
    print("\nStep 3: Running name_assign.py...")
    name_assign_output = run_script_and_get_output("name_assign.py", "name_assign")
    if not name_assign_output:
        # Try to find the output file
        name_assign_output = os.environ.get("NAME_ASSIGN_OUTPUT_PATH", "./name_assign_output.xlsx")
        if not os.path.exists(name_assign_output):
            output_raw_dir = os.environ.get("OUTPUT_RAW_DIR", "./output_raw")
            name_assign_output = find_latest_file(os.path.join(output_raw_dir, "name_assign_output*.xlsx"))
    
    output_files["name_assign.py"] = name_assign_output
    print(f"Name Assign output: {name_assign_output}")
    
    # 4. Run targeting.py
    print("\nStep 4: Running targeting.py...")
    targeting_output = run_script_and_get_output("targeting.py", "targeting")
    if not targeting_output:
        # Try to find the output file
        targeting_output = os.environ.get("TARGETING_OUTPUT_PATH", "./targeting_check_output.xlsx")
        if not os.path.exists(targeting_output):
            output_raw_dir = os.environ.get("OUTPUT_RAW_DIR", "./output_raw")
            targeting_output = find_latest_file(os.path.join(output_raw_dir, "targeting_check_output*.xlsx"))
    
    output_files["targeting.py"] = targeting_output
    print(f"Targeting output: {targeting_output}")
    
    # 5. Run creative.py
    print("\nStep 5: Running creative.py...")
    creative_output = run_script_and_get_output("creative.py", "creative")
    if not creative_output:
        # Try to find the output file
        creative_output = os.environ.get("CREATIVE_OUTPUT_PATH", "./creative_qa_output.xlsx")
        if not os.path.exists(creative_output):
            output_raw_dir = os.environ.get("OUTPUT_RAW_DIR", "./output_raw")
            creative_output = find_latest_file(os.path.join(output_raw_dir, "creative_qa_output*.xlsx"))
    
    output_files["creative.py"] = creative_output
    print(f"Creative output: {creative_output}")
    
    # 6. Combine all outputs into a single Excel file
    print("\nStep 6: Combining all outputs...")
    combined_report_path = create_combined_report(qa_report_path, output_files)
    
    # Calculate and print execution time
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"\nExecution completed in {execution_time:.2f} seconds")
    print(f"Combined QA report saved to: {combined_report_path}")
    
    return combined_report_path

if __name__ == "__main__":
    main() 