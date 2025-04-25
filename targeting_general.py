"""
Targeting Data General Check Script

This script performs general checks on targeting data from a QA report and compares it against
a Campaign Brief to validate configuration. It serves as the foundation for more specific 
targeting validation scripts.

Usage:
1. Run this script
2. Results will be used by targeting.py to generate a comprehensive report
"""

import pandas as pd
import numpy as np
import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import dotenv
from pathlib import Path
import json
import math
import traceback
import glob

# Import the function from brief_extractor
from brief_extractor import extract_structured_brief_data

# Find the latest QA report file
def find_latest_qa_report(output_dir=None):
    """Find the latest QA report file in the output directory or current directory"""
    if output_dir and os.path.exists(output_dir):
        qa_report_files = glob.glob(os.path.join(output_dir, "qa_report_*.xlsx"))
    else:
        qa_report_files = glob.glob("qa_report_*.xlsx")
    
    if not qa_report_files:
        return None
    
    # Sort by modification time, newest first
    latest_file = max(qa_report_files, key=os.path.getmtime)
    return latest_file

# Load environment variables
def load_env():
    """Load environment variables from beeswax_input_qa.env file"""
    # Try to load from the standard location first
    env_path = "./input_folder/beeswax_input_qa.env"
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists("./beeswax_input_qa.env"):
        env_path = "./beeswax_input_qa.env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists("config.env"):
        env_path = "config.env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists(".env"):
        env_path = ".env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    else:
        print("No environment file found, using defaults")
    
    # If environment variables weren't loaded or the paths are not set, use defaults
    base_dir = os.getcwd()
    brief_path = os.getenv("BRIEF_PATH", "./Brief/Campaign_Brief.xlsx")
    
    # Determine QA report path - first check env variable, then find latest
    qa_report_path = os.getenv("QA_REPORT_PATH", "./qa_report.xlsx")
    output_dir = os.getenv("OUTPUT_DIR", "./output_folder")
    
    # Look for the latest QA report if the default doesn't exist
    if not os.path.exists(qa_report_path):
        print(f"QA report not found at {qa_report_path}, looking for the latest report...")
        latest_report = find_latest_qa_report(output_dir)
        if latest_report:
            qa_report_path = latest_report
            print(f"Using latest QA report: {qa_report_path}")
        else:
            print(f"No QA report found, will use the default path: {qa_report_path}")
    
    # Ensure paths use proper separators for the current OS
    brief_path = brief_path.replace('/', os.sep)
    qa_report_path = qa_report_path.replace('/', os.sep)
    
    # Construct the full paths (handle both relative and absolute paths)
    if not os.path.isabs(brief_path):
        full_brief_path = os.path.join(base_dir, brief_path)
    else:
        full_brief_path = brief_path
        
    if not os.path.isabs(qa_report_path):
        full_qa_report_path = os.path.join(base_dir, qa_report_path)
    else:
        full_qa_report_path = qa_report_path
    
    print(f"Using brief path: {full_brief_path}")
    print(f"Using QA report path: {full_qa_report_path}")
    
    return full_brief_path, full_qa_report_path

# --- Define QA Check Functions ---
def check_exact_match(value, expected):
    """Check if value exactly matches expected (case-insensitive)"""
    return str(value).strip().lower() == str(expected).lower()

def check_is_empty(value, expected):
    """Check if value is empty or contains only whitespace/special characters"""
    if value is None:
        return True
    if pd.isna(value):
        return True
    value_str = str(value).strip()
    if value_str == '':
        return True
    if value_str.lower() == 'nan':
        return True
    if value_str == '-':
        return True
    if value_str == 'None':
        return True
    return False

def check_is_number_one(value, expected):
    """Check if value is 1 or 1.0"""
    val_str = str(value).strip()
    return val_str == '1' or val_str == '1.0'

def check_is_false(value, expected):
    """Check if value is false"""
    return str(value).strip().lower() == 'false'

def check_is_true(value, expected):
    """Check if value is true"""
    return str(value).strip().lower() == 'true'

def check_frequency_duration(value, expected):
    """Check if value matches any of the expected values"""
    return str(value).strip().lower() in [v.lower() for v in expected]

def check_campaign_budget_format(value, expected):
    """Check if campaign budget follows the format {'lifetime': number, 'daily': None}"""
    try:
        # Remove spaces and single quotes
        value_str = str(value).replace(" ", "").replace("'", '"')
        # Check if it matches the pattern
        return bool(re.match(r'^\{"lifetime":\d+,"daily":None\}$', value_str))
    except:
        return False

# Define QA checks with their corresponding functions and expected values
qa_checks = {
    # Standard checks that are actually used in QA
    'Programmatic Guaranteed': (check_is_false, 'false'),
    'Budget Type': (check_exact_match, 'impressions'),
    'Bidding Strategy': (check_exact_match, 'CPM_PACED'),
    'Pacing': (check_exact_match, 'lifetime'),
    'Pacing Behavior': (check_exact_match, 'even'),
    'Multiplier': (check_is_number_one, '1'),
    'Catchup Behavior': (check_exact_match, 'even'),
    'Bid Shading': (check_is_empty, None),
    'Use Custom Bidding Agent': (check_is_false, 'false'),
    'Frequency Cap ID Type': (check_exact_match, 'STANDARD'),
    'Frequency Duration': (check_frequency_duration, ['(1;1;week)', '(2;1;week)', '(3;1;week)']),
    'Use Fallback': (check_is_true, 'true'),
    'campaign_impressions_budget': (check_campaign_budget_format, None)  # New check for campaign budget format
}

# Map check names to actual column names in the data
check_to_column_map = {
    'Programmatic Guaranteed': 'Programmatic Guaranteed',
    'Budget Type': 'Budget Type',
    'Bidding Strategy': 'Bidding Strategy',
    'Pacing': 'Pacing',
    'Pacing Behavior': 'Pacing Behavior',
    'Multiplier': 'Multiplier',
    'Catchup Behavior': 'Catchup Behavior',
    'Bid Shading': 'Bid Shading',
    'Use Custom Bidding Agent': 'Use Custom Bidding Agent',
    'Frequency Cap ID Type': 'Frequency Cap ID Type',
    'Frequency Duration': 'Frequency Duration',
    'Use Fallback': 'Use Fallback',
    'campaign_impressions_budget': 'campaign_impressions_budget'  # Added mapping for campaign budget
}

def safe_date_convert(date_val):
    """Safely convert various date formats to pandas datetime"""
    if pd.isna(date_val) or date_val == '':
        return None
    
    if isinstance(date_val, (datetime, pd.Timestamp)):
        return pd.to_datetime(date_val)
    
    try:
        return pd.to_datetime(date_val)
    except Exception as e1:
        try:
            if isinstance(date_val, (int, float)):
                if 30000 < date_val < 70000: 
                    return pd.to_datetime('1899-12-30') + pd.Timedelta(days=float(date_val))
        except Exception as e2:
            print(f"Excel date conversion error for {date_val}: {e2}")
            pass
            
        date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y', '%m-%d-%Y',
            '%Y/%m/%d', '%b %d, %Y', '%d %b %Y',
            '%m/%d/%y', '%d/%m/%y', '%y-%m-%d',
            '%m.%d.%Y', '%d.%m.%Y', '%Y.%m.%d'
        ]
        
        for fmt in date_formats:
            try:
                return pd.to_datetime(str(date_val).strip(), format=fmt)
            except (ValueError, TypeError):
                continue
                
    print(f"Warning: Could not convert '{date_val}' to datetime after trying multiple formats. Error: {e1}")
    return None

def get_field_value(df, field_pattern, default=None):
    """
    Extract value for a specific field pattern from a DataFrame with 'Field'/'Value' columns.
    Case-insensitive search. Returns the first match or default.
    """
    if df is None or df.empty or 'Field' not in df.columns or 'Value' not in df.columns:
        return default
        
    try:
        df['Field'] = df['Field'].astype(str) 
        matching_rows = df[df['Field'].str.contains(field_pattern, case=False, na=False, regex=False)]
        
        if not matching_rows.empty:
            value = matching_rows.iloc[0]['Value']
            if isinstance(value, np.generic):
                 value = value.item()
            return value if pd.notna(value) else default
        else:
            # Reduced verbosity
            # print(f"Field containing pattern '{field_pattern}' not found. Available fields: {list(df['Field'])}")
            return default
            
    except KeyError:
        print(f"Error: 'Field' or 'Value' column not found in DataFrame while searching for '{field_pattern}'.")
        return default
    except Exception as e:
        print(f"Error getting field value for '{field_pattern}': {e}")
        return default

def find_col_name(columns, possible_names):
    """Find the first matching column name from a list of possibilities (case-insensitive)."""
    col_map = {str(col).lower().strip(): str(col) for col in columns}
    for name in possible_names:
        if name.lower().strip() in col_map:
            return col_map[name.lower().strip()]
    return None

def get_base_cpm(platform_media, geo_required, lda_compliant):
    """
    Determine base CPM based on platform/media type and requirements.
    
    Args:
        platform_media (str): Platform and media type (e.g., 'Mobile/Banner', 'Desktop/Video')
        geo_required (str): Whether geo-targeting is required ('Yes' or 'No')
        lda_compliant (str): Whether LDA compliant is required ('Yes' or 'No')
    
    Returns:
        float: Base CPM value
    """
    # Convert inputs to lowercase for comparison
    platform_media = str(platform_media).lower().strip()
    geo_required = str(geo_required).lower().strip()
    lda_compliant = str(lda_compliant).lower().strip()
    
    # Mobile CPMs
    if platform_media.startswith('mobile'):
        if 'banner' in platform_media:
            return 2.34 if geo_required == 'yes' else 2.00
        elif 'rich media' in platform_media:
            return 3.15  # regardless of geo
        elif 'video' in platform_media:
            return 6.30  # regardless of geo
            
    # Desktop CPMs
    elif platform_media.startswith('desktop'):
        if 'banner' in platform_media:
            return 2.89 if geo_required == 'yes' else 2.36
        elif 'rich media' in platform_media:
            return 2.89  # regardless of geo
        elif 'video' in platform_media:
            return 7.35  # regardless of geo
            
    # CTV CPMs
    elif platform_media.startswith('ctv'):
        return 19.00 if lda_compliant == 'yes' else 17.00
        
    return None  # Return None if no matching rule found

def normalize_viewability_goal(viewability_goal):
    """
    Normalize viewability goal to a percentage between 0-100.
    Handles both decimal (0.7) and percentage (70) formats.
    """
    try:
        value = float(str(viewability_goal).replace('%', ''))
        # If value is less than 1, assume it's in decimal format (0.7)
        if value < 1:
            value = value * 100
        return value
    except (ValueError, TypeError):
        return None

def get_viewability_addon_cpm(platform_media, viewability_goal):
    """
    Calculate viewability add-on CPM based on platform/media and viewability goal.
    
    Args:
        platform_media (str): Platform and media type
        viewability_goal (str): Viewability goal percentage or decimal
        
    Returns:
        tuple: (float, str) - (Add-on CPM value, normalized viewability percentage string)
    """
    try:
        platform_media = str(platform_media).lower().strip()
        
        # Normalize viewability goal to percentage
        viewability = normalize_viewability_goal(viewability_goal)
        if viewability is None:
            return 0.0, "N/A"
            
        # Format normalized viewability for display
        normalized_viewability = f"{viewability:.1f}%"
        
        # Define viewability ranges and their corresponding CPMs
        viewability_cpms = {
            'mobile/banner': {
                (60, 74): 0.15,
                (75, 84): 0.40,
                (85, 94): 1.43,
                (95, 100): 2.47
            },
            'mobile/geo-targeting': {
                (60, 74): 0.15,
                (75, 84): 0.47,
                (85, 94): 1.67,
                (95, 100): 2.90
            },
            'mobile/rich media': {
                (60, 74): 0.15,
                (75, 84): 0.63,
                (85, 94): 2.25,
                (95, 100): 3.90
            },
            'mobile/video': {
                (60, 74): 2.45,
                (75, 84): 3.90,
                (85, 94): 4.50,
                (95, 100): 7.80
            },
            'desktop/banner': {
                (60, 74): 0.45,
                (75, 84): 0.83,
                (85, 94): 2.95,
                (95, 100): 5.06
            },
            'desktop/geo-targeting': {
                (60, 74): 0.45,
                (75, 84): 1.02,
                (85, 94): 3.60,
                (95, 100): 6.19
            },
            'desktop/rich media': {
                (60, 74): 0.45,
                (75, 84): 1.02,
                (85, 94): 3.60,
                (95, 100): 6.19
            },
            'desktop/video': {
                (60, 74): 2.45,
                (75, 84): 4.55,
                (85, 94): 9.17,
                (95, 100): 15.75
            }
        }
        
        # Find matching platform CPM rules
        platform_rules = None
        for key in viewability_cpms.keys():
            if platform_media.startswith(key):
                platform_rules = viewability_cpms[key]
                break
        
        if platform_rules:
            # Find matching viewability range
            for (min_view, max_view), addon_cpm in platform_rules.items():
                if min_view <= viewability <= max_view:
                    return addon_cpm, normalized_viewability
                    
    except Exception as e:
        print(f"Error calculating viewability add-on: {e}")
        
    return 0.0, "N/A"

def extract_cpm_bid(value):
    """Extract CPM bid value from JSON-like string."""
    try:
        # Remove curly braces and split by colon
        value = str(value).strip('{}').replace('"', '')
        key_value = value.split(':')
        if len(key_value) == 2 and 'cpm_bid' in key_value[0]:
            return float(key_value[1].strip())
    except Exception as e:
        print(f"Error extracting CPM bid: {e}")
    return None

def validate_bidding_value(row):
    """Validate the bidding value (CPM) for a row based on platform, geo, LDA, and viewability requirements."""
    try:
        # Get required values
        platform_media = str(row.get('brief_platform_media', '')).strip()
        geo_required = str(row.get('brief_geo_required', '')).strip()
        lda_compliant = str(row.get('brief_lda_compliant', '')).strip()
        viewability_goal = str(row.get('brief_viewability_goal', '')).strip()
        
        # Extract CPM from JSON-like string
        bidding_value = str(row.get('Bidding Values', ''))
        actual_cpm = extract_cpm_bid(bidding_value)
        
        if actual_cpm is None:
            return False, 0, 0, "N/A", f"Could not extract CPM from bidding value: {bidding_value}"
        
        # Calculate expected CPM
        base_cpm = get_base_cpm(platform_media, geo_required, lda_compliant)
        if base_cpm is None:
            return False, 0, actual_cpm, "N/A", f"No matching base CPM rule for platform/media: {platform_media}"
            
        addon_cpm, normalized_viewability = get_viewability_addon_cpm(platform_media, viewability_goal)
        expected_cpm = base_cpm + addon_cpm
        
        # Exact match comparison (no tolerance)
        is_valid = expected_cpm == actual_cpm
        
        explanation = (
            f"Platform: {platform_media}, "
            f"Geo Required: {geo_required}, "
            f"LDA: {lda_compliant}, "
            f"Viewability: {normalized_viewability}, "
            f"Expected: ${expected_cpm:.2f} (Base: ${base_cpm:.2f} + Addon: ${addon_cpm:.2f}), "
            f"Actual: ${actual_cpm:.2f}"
        )
        
        return is_valid, expected_cpm, actual_cpm, normalized_viewability, explanation
        
    except Exception as e:
        return False, 0, 0, "N/A", f"Error validating CPM: {str(e)}"

def extract_campaign_budget(budget_str):
    """Extract lifetime budget value from campaign_impressions_budget JSON string."""
    try:
        # Remove curly braces and single quotes
        budget_str = budget_str.strip('{}').replace("'", '"')
        # Split by comma and find lifetime value
        parts = budget_str.split(',')
        for part in parts:
            if 'lifetime' in part:
                value = part.split(':')[1].strip()
                return int(value) if value.lower() != 'none' else None
    except Exception as e:
        print(f"Error extracting campaign budget: {e}")
    return None

def calculate_campaign_metrics(df):
    """Calculate campaign-level metrics including total impressions and buffer percentage."""
    try:
        # Group by Campaign ID and calculate total brief impressions
        campaign_totals = df.groupby('Campaign ID')['brief_impressions'].sum().reset_index()
        campaign_totals.rename(columns={'brief_impressions': 'Campaign_Imps_Required'}, inplace=True)
        
        # Merge back with original dataframe
        df = pd.merge(df, campaign_totals, on='Campaign ID', how='left')
        
        # Calculate Campaign Buffer %
        df['Campaign Buffer %'] = None
        mask = (df['campaign_impressions_budget'].notna()) & (df['Campaign_Imps_Required'].notna())
        df.loc[mask, 'Campaign Buffer %'] = ((df.loc[mask, 'campaign_impressions_budget'].apply(extract_campaign_budget) - 
                                            df.loc[mask, 'Campaign_Imps_Required']) / 
                                           df.loc[mask, 'Campaign_Imps_Required'] * 100)
        
        # Format Campaign_Imps_Required as whole numbers with commas
        df['Campaign_Imps_Required'] = df['Campaign_Imps_Required'].fillna(0).astype(int).apply(lambda x: f"{x:,}")
        
        # Format Campaign Buffer % with 2 decimal places and % symbol
        df.loc[df['Campaign Buffer %'].notna(), 'Campaign Buffer %'] = df.loc[df['Campaign Buffer %'].notna(), 'Campaign Buffer %'].apply(lambda x: f"{x:.2f}%")
        
        return df
    except Exception as e:
        print(f"Error calculating campaign metrics: {e}")
        return df

def apply_formatting(wb, qa_df):
    """Apply formatting to the Excel output using openpyxl."""
    ws = wb.active
    
    # Define formats
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray color
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical='top')
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Apply header format
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Get column indices
    col_indices = {name: idx for idx, name in enumerate(qa_df.columns, 1)}
    
    # Prepare for budget split validation
    budget_split_col_idx = col_indices.get('Budget Split %')
    alt_id_col_idx = None
    alt_id_col_name = find_col_name(qa_df.columns, ['Alt ID', 'line_item_alternative_id'])
    if alt_id_col_name:
        alt_id_col_idx = col_indices.get(alt_id_col_name)
    
    # Group budget splits by Alt ID to check if they sum to 100%
    alt_id_to_rows = {}
    if budget_split_col_idx and alt_id_col_idx:
        for row in range(2, ws.max_row + 1):
            alt_id_cell = ws.cell(row=row, column=alt_id_col_idx)
            alt_id = str(alt_id_cell.value).strip()
            
            if alt_id not in alt_id_to_rows:
                alt_id_to_rows[alt_id] = []
            alt_id_to_rows[alt_id].append(row)
    
    # Apply QA check highlighting
    for row in range(2, ws.max_row + 1):
        row_data = qa_df.iloc[row-2]
        
        # Apply standard checks from qa_checks dictionary
        for col_name, (check_func, expected_value) in qa_checks.items():
            if col_name in col_indices:
                cell = ws.cell(row=row, column=col_indices[col_name])
                cell_value = row_data.get(col_name)
                
                # Skip special handling cases
                if check_func is None:
                    continue
            
                # Apply the check
                if check_func(cell_value, expected_value):
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
                    print(f"Check failed: {col_name} - Expected: {expected_value}, Got: {cell_value}")
        
        # Special handling for Bidding Values
        if 'Bidding Values' in col_indices:
            cell = ws.cell(row=row, column=col_indices['Bidding Values'])
            is_valid, _, _, _, _ = validate_bidding_value(row_data)
            if is_valid:
                cell.fill = green_fill
            else:
                cell.fill = red_fill
            
        # Campaign Budget Type Check
        if 'campaign_budget_type' in col_indices:
            cell = ws.cell(row=row, column=col_indices['campaign_budget_type'])
            if str(cell.value).lower().strip() == 'impressions':
                cell.fill = green_fill
            else:
                cell.fill = red_fill
            
        # Buffer Checks
        for buffer_col in ['Campaign Buffer %', 'Line Buffer %']:
            if buffer_col in col_indices:
                cell = ws.cell(row=row, column=col_indices[buffer_col])
                buffer_value = str(cell.value).strip()
                traffic_info = row_data.get('brief_traffic_info', '')
                
                if buffer_value != '' and buffer_value != 'Error':
                    try:
                        diff_value = float(buffer_value.strip('%').replace('+', ''))
                        expected_diff = 6.0 if str(traffic_info).lower() == 'yes' else 3.0
                        
                        if abs(diff_value - expected_diff) < 0.1:
                            cell.fill = green_fill
                        else:
                            cell.fill = red_fill
                    except ValueError:
                        cell.fill = red_fill
                else:
                    cell.fill = red_fill
    
    # Apply Budget Split % validation after all other checks
    if budget_split_col_idx and alt_id_col_idx:
        print("\n--- Validating Budget Split Percentages ---")
        for alt_id, rows in alt_id_to_rows.items():
            # Extract budget split values from all rows with this Alt ID
            split_values = []
            split_rows = []  # Store row indices along with their split values
            
            # Check LDA compliance - we'll need this for additional validation
            lda_compliant = "unknown"  # Default value to track if we found it
            brief_lda_compliant_col = None
            lda_col_name = find_col_name(qa_df.columns, ['brief_lda_compliant'])
            
            if lda_col_name:
                brief_lda_compliant_col = col_indices.get(lda_col_name)
                if brief_lda_compliant_col:
                    # Get LDA compliance from the first row (should be the same for all rows with same Alt ID)
                    first_row = rows[0]
                    lda_cell = ws.cell(row=first_row, column=brief_lda_compliant_col)
                    lda_value = str(lda_cell.value).strip()
                    lda_compliant = lda_value.lower()
                    print(f"  Found LDA value for {alt_id}: '{lda_value}' (normalized to '{lda_compliant}')")
                else:
                    print(f"  Warning: brief_lda_compliant column found but index not found in column indices")
            else:
                print(f"  Warning: brief_lda_compliant column not found for {alt_id}")
            
            for row in rows:
                split_cell = ws.cell(row=row, column=budget_split_col_idx)
                split_value = str(split_cell.value).strip()
                
                try:
                    # Extract numeric value from percentage
                    if split_value.endswith('%'):
                        numeric_value = float(split_value.rstrip('%'))
                        split_values.append(numeric_value)
                        split_rows.append((row, numeric_value))
                except ValueError:
                    # Skip non-numeric values
                    continue
            
            # Check if sum rounds to 100% (use tolerance for floating-point comparison)
            valid_sum = False
            if len(split_values) == 1:
                # If only one line item, it should be 100%
                valid_sum = abs(split_values[0] - 100.0) < 0.2  # Small tolerance
                print(f"Alt ID {alt_id}: Single line item with split {split_values[0]}% - {'Valid (100%)' if valid_sum else 'Invalid (should be 100%)'}")
            else:
                # Multiple line items - check if they sum to approximately 100%
                sum_value = sum(split_values)
                valid_sum = abs(sum_value - 100.0) < 0.2  # Small tolerance
                print(f"Alt ID {alt_id}: {len(split_values)} line items with splits {split_values} - Sum: {sum_value:.1f}% - {'Valid (≈100%)' if valid_sum else 'Invalid (not ≈100%)'}")
            
            # Apply LDA compliance check if applicable
            if lda_compliant == "yes":
                print(f"  LDA Compliant: Yes - Checking if splits are either 50% or 100%")
                lda_valid = True
                
                # For LDA-compliant items, splits should only be 50% or 100%
                for _, split_value in split_rows:
                    if not (abs(split_value - 50.0) < 0.2 or abs(split_value - 100.0) < 0.2):
                        lda_valid = False
                        print(f"  Invalid split value for LDA: {split_value}% (should be 50% or 100%)")
                
                if not lda_valid:
                    print("  LDA compliance check failed - marking as invalid")
                    valid_sum = False  # Override validation if LDA requirement fails
                else:
                    print("  All split values are valid for LDA compliance")
            
            # Apply formatting based on validation
            for row in rows:
                split_cell = ws.cell(row=row, column=budget_split_col_idx)
                if valid_sum:
                    split_cell.fill = green_fill
                    print(f"  Row {row}: VALID (green)")
                else:
                    split_cell.fill = red_fill
                    print(f"  Row {row}: INVALID (red)")

    # Set column widths
    for idx, col in enumerate(qa_df.columns, 1):
        max_length = max(
            qa_df[col].astype(str).apply(len).max(),
            len(str(col))
        )
        ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

    # Center align specific columns
    for align_col in ['Campaign Buffer %', 'Line Buffer %', 'Budget Split %']:
        if align_col in col_indices:
            for r_idx in range(2, ws.max_row + 1):
                cell = ws[f"{get_column_letter(col_indices[align_col])}{r_idx}"]
                cell.alignment = Alignment(horizontal='center')
                
    # FINAL CHECK: Make one more pass for LDA budget split validation
    if budget_split_col_idx and alt_id_col_idx:
        print("\n--- FINAL CHECK: Forcing Budget Split % validation with LDA rule ---")
        for alt_id, rows in alt_id_to_rows.items():
            # Extract values again
            split_values = []
            split_rows = []
            
            # Get LDA value
            first_row = rows[0]
            lda_compliant = "unknown"
            lda_col_name = find_col_name(qa_df.columns, ['brief_lda_compliant'])
            if lda_col_name:
                brief_lda_compliant_col = col_indices.get(lda_col_name)
                if brief_lda_compliant_col:
                    lda_cell = ws.cell(row=first_row, column=brief_lda_compliant_col)
                    lda_value = str(lda_cell.value).strip()
                    lda_compliant = lda_value.lower()
                    print(f"  FINAL CHECK: LDA value for {alt_id} is '{lda_compliant}'")
            
            # Collect split values
            for row in rows:
                cell = ws.cell(row=row, column=budget_split_col_idx)
                split_value = str(cell.value).strip()
                try:
                    if split_value.endswith('%'):
                        numeric_value = float(split_value.rstrip('%'))
                        split_values.append(numeric_value)
                        split_rows.append((row, numeric_value))
                except ValueError:
                    continue
            
            # Check sum to 100%
            valid_sum = False
            if len(split_values) == 1:
                valid_sum = abs(split_values[0] - 100.0) < 0.2
            else:
                sum_value = sum(split_values)
                valid_sum = abs(sum_value - 100.0) < 0.2
            
            # LDA validation (stricter enforcement)
            if lda_compliant == "yes":
                print(f"  FINAL CHECK: LDA validation for Alt ID {alt_id} - Splits: {split_values}")
                
                # Check each split value strictly
                lda_valid = True
                for row, value in split_rows:
                    # Must be exactly 50% or 100% (within tolerance)
                    is_valid_split = abs(value - 50.0) < 0.2 or abs(value - 100.0) < 0.2
                    
                    if not is_valid_split:
                        lda_valid = False
                        print(f"    Invalid LDA split at row {row}: {value}% (must be 50% or 100%)")
                
                # If any value fails LDA check, mark ALL rows in the group as invalid
                if not lda_valid:
                    valid_sum = False
                    print(f"    Group {alt_id} FAILS LDA check - All rows will be marked RED")
                else:
                    print(f"    Group {alt_id} PASSES LDA check - All rows will be marked GREEN")
            
            # Force apply formatting - use direct RGB hex codes to ensure consistency
            for row in rows:
                cell = ws.cell(row=row, column=budget_split_col_idx)
                if valid_sum:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    print(f"    Row {row}: FORCE GREEN")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                    print(f"    Row {row}: FORCE RED")

def main():
    print("Starting Targeting Data Check...")
    # --- Configuration ---
    brief_path, qa_report_path = load_env()
    
    # --- Load Data ---
    print(f"Loading QA Report: {qa_report_path}")
    if not os.path.exists(qa_report_path):
        print(f"Error: QA Report file not found at {qa_report_path}")
        return None
    try:
        # Load both Targeting Data and Consolidated Report tabs
        qa_df = pd.read_excel(qa_report_path, sheet_name='Targeting Data')
        consolidated_df = pd.read_excel(qa_report_path, sheet_name='Consolidated Report')
        
        print(f"QA Report Targeting Data loaded successfully. Shape: {qa_df.shape}")
        print(f"Consolidated Report loaded successfully. Shape: {consolidated_df.shape}")
        
        # Clean column names for both dataframes
        qa_df.columns = [str(col).strip() for col in qa_df.columns]
        consolidated_df.columns = [str(col).strip() for col in consolidated_df.columns]
        
        # Store original column order
        original_targeting_cols = list(qa_df.columns)
        
        # Extract campaign budget information from Consolidated Report
        campaign_cols = ['Campaign ID', 'campaign_budget_type', 'campaign_impressions_budget']
        campaign_data = None
        
        try:
            # Find the actual column names in consolidated report
            consolidated_campaign_id_col = find_col_name(consolidated_df.columns, ['Campaign ID', 'campaign_id'])
            consolidated_budget_type_col = find_col_name(consolidated_df.columns, ['campaign_budget_type', 'budget type'])
            consolidated_impressions_budget_col = find_col_name(consolidated_df.columns, ['campaign_impressions_budget', 'impressions budget'])
            
            if all([consolidated_campaign_id_col, consolidated_budget_type_col, consolidated_impressions_budget_col]):
                print("Found all required campaign columns in Consolidated Report.")
                
                # Create mapping dictionary for renaming
                rename_dict = {
                    consolidated_campaign_id_col: 'Campaign ID',
                    consolidated_budget_type_col: 'campaign_budget_type',
                    consolidated_impressions_budget_col: 'campaign_impressions_budget'
                }
                
                # Select and rename columns
                campaign_data = consolidated_df[[consolidated_campaign_id_col, 
                                              consolidated_budget_type_col,
                                              consolidated_impressions_budget_col]].copy()
                campaign_data.rename(columns=rename_dict, inplace=True)
                
                # Drop duplicates based on Campaign ID
                campaign_data = campaign_data.drop_duplicates(subset=['Campaign ID'], keep='first')
                
                print(f"Extracted campaign budget data. Shape: {campaign_data.shape}")
                
                # Merge campaign data with targeting data
                qa_campaign_col = find_col_name(qa_df.columns, ['Campaign ID', 'campaign_id'])
                if qa_campaign_col:
                    print("Merging campaign budget data with targeting data...")
                    qa_df = pd.merge(
                        qa_df,
                        campaign_data,
                        left_on=qa_campaign_col,
                        right_on='Campaign ID',
                        how='left'
                    )
                    print("Campaign budget data merged successfully.")
                else:
                    print("Warning: Could not find Campaign ID column in targeting data for merge.")
            else:
                print("Warning: Could not find all required campaign columns in Consolidated Report.")
                missing_cols = []
                if not consolidated_campaign_id_col: missing_cols.append("Campaign ID")
                if not consolidated_budget_type_col: missing_cols.append("campaign_budget_type")
                if not consolidated_impressions_budget_col: missing_cols.append("campaign_impressions_budget")
                print(f"Missing columns: {missing_cols}")
                
        except Exception as e:
            print(f"Error processing Consolidated Report: {e}")
            # Add empty columns to maintain structure
            qa_df['campaign_budget_type'] = 'N/A'
            qa_df['campaign_impressions_budget'] = 'N/A'
        
    except Exception as e:
        print(f"Error loading QA report: {e}")
        return None

    print(f"\nLoading and processing Campaign Brief: {brief_path}")
    if not os.path.exists(brief_path):
        print(f"Error: Campaign Brief file not found at {brief_path}")
        return None
    try:
        # Extract structured data using brief_extractor
        structured_brief_data = extract_structured_brief_data(brief_path)
        
        # Check if extraction was successful
        if not structured_brief_data or not any(df is not None and not df.empty for df in structured_brief_data.values()):
             print("Error: Failed to extract any structured data from the brief.")
             return None

        # Get individual dataframes
        account_data = structured_brief_data.get('account_data')
        campaign_data = structured_brief_data.get('campaign_data')
        placement_data = structured_brief_data.get('placement_data')
        target_data = structured_brief_data.get('target_data') 

        print("\nBrief data extracted successfully.")
        
    except Exception as e:
        print(f"Error processing Campaign Brief: {e}")
        return None

    # --- Extract Key Information from Brief ---
    # 1. Account Level Data
    product_type_str_brief = None
    if account_data is not None and not account_data.empty:
        product_type_col = find_col_name(account_data.columns, ['Product Type', 'product type', 'Campaign Type'])
        if product_type_col:
            product_type_str_brief = account_data.iloc[0].get(product_type_col)
            if pd.notna(product_type_str_brief):
                product_type_str_brief = str(product_type_str_brief).strip()
                print(f"Found Product Type string: {product_type_str_brief}")
            else:
                print("Warning: Product Type column found but value is empty/NaN in Account Data.")
                product_type_str_brief = None
        else:
            print(f"Warning: Product Type column not found in Account Data. Available columns: {list(account_data.columns)}")

    # 2. Campaign Level Data
    dairy_milk_restrictions = None
    lda_compliant = None
    viewability_contracted = None
    measurement_type_brief = None # Initialize Measurement Type

    if campaign_data is not None:
        dairy_milk_restrictions = get_field_value(campaign_data, 'Apply Dairy-Milk Restrictions')
        lda_compliant = get_field_value(campaign_data, 'LDA or Age Compliant')
        viewability_contracted = get_field_value(campaign_data, 'Viewability Contracted')
        measurement_type_brief = get_field_value(campaign_data, 'Measurement Type') # Extract Measurement Type
        print(f"Found Measurement Type: {measurement_type_brief}") # Log extraction

    # --- Prepare DataFrames for Merging ---
    print("\n--- Preparing Brief Data for Merge ---")
    # Find column names dynamically in Brief data
    bvp_col_pl, geo_col_pl, traffic_info_col_pl = None, None, None
    bvt_col_tg, bvp_col_tg, platform_media_col_tg, impressions_col_tg = None, None, None, None

    # 1. Target Level Data (Brief)
    target_data_to_merge = None # Initialize
    if target_data is not None and not target_data.empty:
        target_data.columns = [str(col).strip() for col in target_data.columns] # Clean target columns
        print(f"Processing Target Level Data (Brief)... Columns: {list(target_data.columns)}")
        # Prioritize 'bvt' or 'bvt id' over 'bv id'
        bvt_col_tg = find_col_name(target_data.columns, ['bvt', 'bvt id', 'bv id'])
        bvp_col_tg = find_col_name(target_data.columns, ['bvp', 'bvp id'])
        platform_media_col_tg = find_col_name(target_data.columns, ['platform/media type', 'platform / media type', 'platform', 'media type'])
        impressions_col_tg = find_col_name(target_data.columns, ['impressions', 'impression', 'imp'])
        
        print(f"Identified Target (Brief) columns: BVT='{bvt_col_tg}', BVP='{bvp_col_tg}', Platform/Media='{platform_media_col_tg}', Impressions='{impressions_col_tg}'")

        if bvt_col_tg:
            # Prepare the BVT column for case-insensitive merge
            target_data['merge_key_bvt_lower'] = target_data[bvt_col_tg].astype(str).str.strip().str.lower()
            
            columns_to_select_tg = ['merge_key_bvt_lower'] # Use the lower case key for selection
            # Keep original BVT col if needed: columns_to_select_tg.append(bvt_col_tg)
            rename_dict_tg = {'merge_key_bvt_lower': 'merge_key_bvt_tg'}

            if bvp_col_tg:
                target_data[bvp_col_tg] = target_data[bvp_col_tg].astype(str).str.strip()
                columns_to_select_tg.append(bvp_col_tg)
                rename_dict_tg[bvp_col_tg] = 'brief_bvp_id'
            if platform_media_col_tg:
                target_data[platform_media_col_tg] = target_data[platform_media_col_tg].astype(str).str.strip()
                columns_to_select_tg.append(platform_media_col_tg)
                rename_dict_tg[platform_media_col_tg] = 'brief_platform_media'
            if impressions_col_tg:
                target_data[impressions_col_tg] = target_data[impressions_col_tg].astype(str).str.strip()
                columns_to_select_tg.append(impressions_col_tg)
                rename_dict_tg[impressions_col_tg] = 'brief_impressions'

            # Select only needed columns and drop duplicates based on the lower case merge key
            target_data_to_merge = target_data[columns_to_select_tg].copy()
            target_data_to_merge.rename(columns=rename_dict_tg, inplace=True)
            # Important: Drop duplicates *after* renaming to ensure the merge key is correct
            target_data_to_merge = target_data_to_merge.drop_duplicates(subset=['merge_key_bvt_tg'], keep='first')
            print(f"Prepared Target (Brief) data for merge. Shape: {target_data_to_merge.shape}")

        else:
            print("Warning: Could not find BVT ID column in Brief Target data. Cannot merge Platform/Media, Impressions, BVP.")
    else:
        print("Warning: Target data not found or empty in brief.")

    # 2. Placement Level Data (Brief)
    placement_data_to_merge = None # Initialize
    if placement_data is not None and not placement_data.empty:
        placement_data.columns = [str(col).strip() for col in placement_data.columns] # Clean placement columns
        print(f"\nProcessing Placement Level Data (Brief)... Columns: {list(placement_data.columns)}")
        # Look for BVP, Geo Required, Traffic Info
        bvp_col_pl = find_col_name(placement_data.columns, ['bvp', 'bvp id', 'bv placement id', 'placement id'])
        # Add more variations for Geo Required including the one with newline
        geo_col_pl = find_col_name(placement_data.columns, ['geo required', 'geo targeting', 'geo', 'geo required? yes/no', 'geo required?yes/no', 'geo required yes no', 'Geo Required?\nYes/No'])
        traffic_info_col_pl = find_col_name(placement_data.columns, ['traffic information', 'traffic info', 'traffic'])
        
        print(f"Identified Placement (Brief) columns: BVP='{bvp_col_pl}', Geo='{geo_col_pl}', Traffic Info='{traffic_info_col_pl}'")

        if bvp_col_pl:
             # Prepare BVP column for case-insensitive merge
            placement_data['merge_key_bvp_lower'] = placement_data[bvp_col_pl].astype(str).str.strip().str.lower()
            columns_to_select_pl = ['merge_key_bvp_lower']
            rename_dict_pl = {'merge_key_bvp_lower': 'merge_key_bvp_pl'}

            if geo_col_pl:
                placement_data[geo_col_pl] = placement_data[geo_col_pl].astype(str).str.strip()
                columns_to_select_pl.append(geo_col_pl)
                rename_dict_pl[geo_col_pl] = 'brief_geo_required' # Ensure correct target name
            else:
                print("Warning: Geo Required column not found in Placement data using provided names.")
                
            if traffic_info_col_pl:
                placement_data[traffic_info_col_pl] = placement_data[traffic_info_col_pl].astype(str).str.strip()
                columns_to_select_pl.append(traffic_info_col_pl)
                rename_dict_pl[traffic_info_col_pl] = 'brief_traffic_info'

            # Select needed columns and drop duplicates based on lower case BVP merge key
            placement_data_to_merge = placement_data[columns_to_select_pl].copy()
            placement_data_to_merge.rename(columns=rename_dict_pl, inplace=True)
            placement_data_to_merge = placement_data_to_merge.drop_duplicates(subset=['merge_key_bvp_pl'], keep='first')
            print(f"Prepared Placement (Brief) data for merge. Shape: {placement_data_to_merge.shape}")
        else:
            print("Warning: Could not find BVP ID column in Brief Placement data. Cannot merge Geo Required, Traffic Info.")
    else:
        print("Warning: Placement data not found or empty in brief.")

    # --- Merge Data ---
    print("\n--- Merging QA data with Brief data ---")
    merged_df = qa_df.copy()  # Start with all targeting data

    # --- Identify QA Report Merge Keys ---
    qa_campaign_col = find_col_name(merged_df.columns, ['Campaign ID', 'campaign_id'])
    qa_alt_id_col = find_col_name(merged_df.columns, ['Alt ID', 'line_item_alternative_id', 'alternative id'])

    print(f"Identified QA Report columns for merging/insertion: Campaign='{qa_campaign_col}', AltID='{qa_alt_id_col}'")

    # Add campaign level data as new columns
    merged_df['brief_product_type'] = product_type_str_brief if pd.notna(product_type_str_brief) else 'N/A'
    merged_df['brief_dairy_milk_restrictions'] = dairy_milk_restrictions if pd.notna(dairy_milk_restrictions) else 'N/A'
    merged_df['brief_lda_compliant'] = lda_compliant if pd.notna(lda_compliant) else 'N/A'
    merged_df['brief_viewability_contracted'] = viewability_contracted if pd.notna(viewability_contracted) else 'N/A'
    merged_df['brief_measurement_type'] = measurement_type_brief if pd.notna(measurement_type_brief) else 'N/A'
    merged_df['brief_viewability_goal'] = get_field_value(campaign_data, 'Viewability Goal') if campaign_data is not None else 'N/A'

    # Merge Target Data (Brief -> QA)
    if target_data_to_merge is not None and qa_alt_id_col:
        print(f"Attempting to merge Target Data (Brief -> QA) using QA column '{qa_alt_id_col}'...")
        # Prepare QA Alt ID for case-insensitive merge
        merged_df['merge_key_altid_lower'] = merged_df[qa_alt_id_col].astype(str).str.strip().str.lower()
            
        # Check for matching keys before merge
        qa_keys = set(merged_df['merge_key_altid_lower'].unique())
        brief_keys = set(target_data_to_merge['merge_key_bvt_tg'].unique())
        matching_keys = qa_keys.intersection(brief_keys)
        print(f"Found {len(matching_keys)} matching keys between QA Alt ID and Brief BVT ID.")
        if not matching_keys: print("Warning: No matching keys found for Target Data merge!")
            
        merged_df = pd.merge(
            merged_df,
            target_data_to_merge,
            left_on='merge_key_altid_lower', # Use lower case QA key
            right_on='merge_key_bvt_tg',    # Use lower case Brief key
            how='left'
        )
        # Drop the temporary lower-case merge keys
        merged_df.drop(columns=['merge_key_altid_lower', 'merge_key_bvt_tg'], inplace=True, errors='ignore')
        print("Merged Target Data (BVP, Platform/Media, Impressions).")
    else:
        print("\nSkipped merging Target Data (Brief Target data missing or QA Alt ID column not found). Creating placeholder columns.")
        if 'brief_bvp_id' not in merged_df.columns: merged_df['brief_bvp_id'] = 'N/A'
        if 'brief_platform_media' not in merged_df.columns: merged_df['brief_platform_media'] = 'N/A'
        if 'brief_impressions' not in merged_df.columns: merged_df['brief_impressions'] = 'N/A'

    # Merge Placement Data (Brief -> QA) using the BVP obtained from Target merge
    # Ensure brief_bvp_id column exists before attempting merge
    if placement_data_to_merge is not None and 'brief_bvp_id' in merged_df.columns:
        print(f"Attempting to merge Placement Data (Brief -> QA) using derived column 'brief_bvp_id'...")
        # Prepare brief_bvp_id for case-insensitive merge, handle potential NAs from previous merge
        merged_df['merge_key_bvp_lower'] = merged_df['brief_bvp_id'].astype(str).str.strip().str.lower()
            
        # Check BVP keys
        qa_bvp_keys = set(merged_df['merge_key_bvp_lower'].unique()) - {'na'} # Exclude potential 'na' from failed previous merge
        brief_placement_keys = set(placement_data_to_merge['merge_key_bvp_pl'].unique())
        matching_bvp_keys = qa_bvp_keys.intersection(brief_placement_keys)
        print(f"Found {len(matching_bvp_keys)} matching keys between derived BVP ID and Brief Placement BVP ID.")
        if not matching_bvp_keys: print("Warning: No matching keys found for Placement Data merge! Check if BVP IDs were correctly derived from Target Data or if BVP IDs in Placement sheet match.")

        merged_df = pd.merge(
            merged_df,
            placement_data_to_merge,
            left_on='merge_key_bvp_lower', # Use lower case derived BVP key
            right_on='merge_key_bvp_pl',   # Use lower case Brief Placement key
            how='left'
        )
         # Drop the temporary lower-case merge keys
        merged_df.drop(columns=['merge_key_bvp_lower', 'merge_key_bvp_pl'], inplace=True, errors='ignore')
        print("Merged Placement Data (Geo Required, Traffic Info).")
    else:
        print("\nSkipped merging Placement Data (Brief Placement data missing or 'brief_bvp_id' column not derived/found). Creating placeholder columns.")
        # Ensure these columns exist even if merge skipped
        if 'brief_geo_required' not in merged_df.columns: merged_df['brief_geo_required'] = 'N/A'
        if 'brief_traffic_info' not in merged_df.columns: merged_df['brief_traffic_info'] = 'N/A'

    # Fill NA values in brief columns that might have been added or missing from merges
    brief_cols = [
        'brief_product_type', 'brief_dairy_milk_restrictions', 'brief_lda_compliant',
        'brief_viewability_contracted', 'brief_measurement_type', 'brief_viewability_goal'
    ]
    for col in brief_cols:
        if col in merged_df.columns:
            merged_df[col] = merged_df[col].fillna('N/A')

    # --- Pre-calculate Budget Check & Split ---
    print("\n--- Pre-calculating Budget Checks and Splits ---")
    budget_col = find_col_name(merged_df.columns, ['Budget'])
    alt_id_col = find_col_name(merged_df.columns, ['Alt ID', 'line_item_alternative_id'])
    brief_impressions_col = find_col_name(merged_df.columns, ['brief_impressions'])
    brief_traffic_info_col = find_col_name(merged_df.columns, ['brief_traffic_info'])
    brief_platform_media_col = find_col_name(merged_df.columns, ['brief_platform_media'])

    merged_df['Budget Check Status'] = 'N/A'
    merged_df['Budget Split %'] = 'N/A'
    merged_df['Target Required'] = np.nan
    merged_df['Line Buffer %'] = 'N/A'

    if all([budget_col, alt_id_col, brief_impressions_col, brief_traffic_info_col, brief_platform_media_col]):
        print("Found all required columns for budget calculation.")

        # Ensure numeric types
        merged_df[budget_col] = pd.to_numeric(merged_df[budget_col], errors='coerce')
        merged_df[brief_impressions_col] = pd.to_numeric(merged_df[brief_impressions_col], errors='coerce')

        # Identify mobile rows
        merged_df['is_mobile'] = merged_df[brief_platform_media_col].str.contains('mobile', case=False, na=False)

        # Group by Alt ID
        grouped = merged_df.groupby(alt_id_col)

        budget_check_results = {}
        budget_splits = {}
        target_budgets = {}
        budget_differences = {}

        for name, group in grouped:
            group_indices = group.index
            is_mobile_group = group['is_mobile'].any()

            # Get brief impressions and traffic info
            brief_imps_series = group[brief_impressions_col].dropna()
            brief_imps_val = brief_imps_series.iloc[0] if not brief_imps_series.empty else None

            traffic_info_series = group[brief_traffic_info_col].dropna().astype(str).str.strip()
            traffic_info_val = traffic_info_series.iloc[0] if not traffic_info_series.empty else None

            if brief_imps_val is None:
                print(f"Warning: Missing Brief Impressions for Alt ID {name}")
                for idx in group_indices:
                    budget_check_results[idx] = 'Error: Missing Brief Imps'
                    budget_splits[idx] = 'Error'
                    target_budgets[idx] = np.nan
                    budget_differences[idx] = 'Error'
                continue

            # Calculate target budget (no buffer)
            target_budget_total = brief_imps_val
            total_budget = group[budget_col].sum()
            
            if is_mobile_group:
                # For mobile groups, calculate splits and check total
                if total_budget > 0:
                    for idx, row_budget in group[budget_col].items():
                        # Calculate split percentage
                        split_pct = (row_budget / total_budget) * 100 if pd.notna(row_budget) else 0
                        budget_splits[idx] = f"{split_pct:.1f}%"
                        
                        # Apply same split percentage to target budget
                        split_target_budget = (split_pct / 100) * target_budget_total
                        target_budgets[idx] = split_target_budget
                        
                        # Calculate difference percentage
                        if split_target_budget > 0:
                            diff_pct = ((row_budget - split_target_budget) / split_target_budget) * 100
                            budget_differences[idx] = f"{diff_pct:+.1f}%"
                            # Check if difference is within 1% tolerance
                            budget_check_results[idx] = abs(diff_pct) <= 1
                        else:
                            budget_differences[idx] = "Error"
                            budget_check_results[idx] = False
                else:
                    for idx in group_indices:
                        budget_splits[idx] = "Error: Zero Total"
                        target_budgets[idx] = 0
                        budget_differences[idx] = "Error"
                        budget_check_results[idx] = False
            else:
                # For non-mobile, each line should match target exactly
                for idx, row_budget in group[budget_col].items():
                    budget_splits[idx] = "100%"
                    target_budgets[idx] = target_budget_total  # Full target budget for non-mobile
                    if pd.notna(row_budget):
                        diff_pct = ((row_budget - target_budget_total) / target_budget_total) * 100
                        budget_differences[idx] = f"{diff_pct:+.1f}%"
                        budget_check_results[idx] = abs(diff_pct) <= 1  # Use 1% tolerance
                    else:
                        budget_differences[idx] = "Error"
                        budget_check_results[idx] = False

        # Apply results back to DataFrame
        merged_df['Budget Check Status'] = pd.Series(budget_check_results)
        merged_df['Target Required'] = pd.Series(target_budgets)
        merged_df['Budget Split %'] = pd.Series(budget_splits)
        merged_df['Line Buffer %'] = pd.Series(budget_differences)

        # Format Target Required column
        merged_df['Target Required'] = merged_df['Target Required'].apply(
            lambda x: f"{round(x):,}" if pd.notna(x) and x != 0 else 'N/A'
        )

        # Drop helper column
        merged_df.drop(columns=['is_mobile'], inplace=True)

    # Pre-calculate Campaign Level Checks
    print("\n--- Pre-calculating Campaign Level Checks ---")
    campaign_budget_type_col = find_col_name(merged_df.columns, ['campaign_budget_type'])
    campaign_impressions_budget_col = find_col_name(merged_df.columns, ['campaign_impressions_budget'])
    
    # Add new columns for campaign checks
    merged_df['Campaign_Imps_Required'] = np.nan
    merged_df['Campaign Buffer %'] = 'N/A'

    if all([campaign_budget_type_col, campaign_impressions_budget_col, brief_impressions_col]):
        print("Found all required columns for campaign budget calculation.")

        # Calculate total brief impressions per campaign
        campaign_groups = merged_df.groupby(qa_campaign_col)
        for campaign_id, group in campaign_groups:
            # First deduplicate based on Alt ID
            unique_alt_id_group = group.drop_duplicates(subset=[qa_alt_id_col], keep='first')
            
            # Sum brief impressions for unique Alt IDs only
            total_brief_imps = pd.to_numeric(unique_alt_id_group[brief_impressions_col], errors='coerce').sum()
            print(f"Campaign {campaign_id}: Found {len(group)} total rows, {len(unique_alt_id_group)} unique Alt IDs")
            print(f"Total brief impressions (after deduplication): {total_brief_imps}")
            
            # Get campaign budget from JSON
            campaign_budget_str = group[campaign_impressions_budget_col].iloc[0]
            campaign_budget = extract_campaign_budget(str(campaign_budget_str))
            
            # Get traffic info for buffer calculation
            traffic_info = str(group[brief_traffic_info_col].iloc[0]).lower().strip()
            expected_buffer = 6.0 if traffic_info == 'yes' else 3.0

            # Calculate and store results
            if pd.notna(total_brief_imps) and campaign_budget is not None:
                # Set Campaign_Imps_Required for all rows in the campaign
                merged_df.loc[group.index, 'Campaign_Imps_Required'] = total_brief_imps
                
                # Calculate buffer percentage
                buffer_pct = ((campaign_budget - total_brief_imps) / total_brief_imps) * 100
                merged_df.loc[group.index, 'Campaign Buffer %'] = f"{buffer_pct:+.1f}%"
            else:
                merged_df.loc[group.index, 'Campaign_Imps_Required'] = 'Error'
                merged_df.loc[group.index, 'Campaign Buffer %'] = 'Error'

    # Format Campaign_Imps_Required to whole numbers
    merged_df['Campaign_Imps_Required'] = merged_df['Campaign_Imps_Required'].apply(
        lambda x: f"{int(x):,}" if pd.notna(x) and x != 'Error' else x
    )

    print("\nTargeting Data Check finished.")
    return merged_df

if __name__ == "__main__":
    main() 