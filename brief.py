import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re
import os
import sys
from datetime import datetime

def highlight_cell(sheet, row, col, color="FFFF00"):
    """Highlight a cell with the specified color
    Green (00FF00) = Pass
    Red (FF0000) = Fail
    Yellow (FFFF00) = Warning
    """
    cell = sheet.cell(row=row, column=col)
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill

def get_cell_value(sheet, row, col):
    """Get cell value, handling None"""
    if row is None or col is None:
        return ""
    
    # Support for both numeric and letter column references
    if isinstance(col, str) and len(col) == 1:
        col = ord(col.upper()) - ord('A') + 1
    
    value = sheet.cell(row=row, column=col).value
    return value if value is not None else ""

def format_date(date_value):
    """Format dates consistently for comparison"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%m/%d/%Y')
    elif isinstance(date_value, str):
        # Try to parse and standardize the date format
        try:
            # Try different date formats
            for fmt in ['%m/%d/%Y', '%#m/%#d/%Y', '%Y-%m-%d %H:%M:%S']:
                try:
                    date_obj = datetime.strptime(date_value, fmt)
                    return date_obj.strftime('%m/%d/%Y')
                except ValueError:
                    continue
        except Exception as e:
            print(f"Date parsing error: {e} for value: {date_value}")
    return str(date_value).strip()

def compare_dates(date1, date2):
    """Compare two date strings and return -1, 0, or 1"""
    try:
        date1_obj = datetime.strptime(date1, '%m/%d/%Y')
        date2_obj = datetime.strptime(date2, '%m/%d/%Y')
        
        if date1_obj < date2_obj:
            return -1
        elif date1_obj > date2_obj:
            return 1
        else:
            return 0
    except:
        # If conversion fails, fall back to string comparison
        return 0 if date1 == date2 else -1 if date1 < date2 else 1

def clean_numeric(value):
    """Convert string numbers with formatting to float values"""
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        return float(value)
    
    # Remove dollar signs, commas, and extra spaces
    clean_value = str(value).replace('$', '').replace(',', '').strip()
    
    # Try to convert to float
    try:
        return float(clean_value)
    except ValueError:
        print(f"WARNING: Could not convert '{value}' to a number")
        return 0.0

def find_row_containing(sheet, text, start_row, end_row=None):
    """Find the first row containing the specified text"""
    if end_row is None:
        end_row = sheet.max_row
    
    for row in range(start_row, end_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = get_cell_value(sheet, row, col)
            if isinstance(cell_value, str) and text.lower() in cell_value.lower():
                return row
    return None

def col_letter_to_number(col_letter):
    """Convert column letter to number (A=1, B=2, etc.)"""
    return ord(col_letter.upper()) - ord('A') + 1

def run_qa_checks(file_path):
    """Run QA checks on the campaign brief using specific cell references"""
    print(f"Running QA checks on {file_path}...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    issues = []
    
    # Define key cell references
    # Campaign info cells
    io_start_date_label_cell = "B15"
    io_start_date_value_cell = "C15"
    io_end_date_label_cell = "B16"
    io_end_date_value_cell = "C16"
    
    # Viewability cells
    viewability_contracted_label_cell = "F15"
    viewability_contracted_value_cell = "G15"
    viewability_goal_label_cell = "F16"
    viewability_goal_value_cell = "G16"
    
    # Budget cells
    bv_budget_label_cell = "B23"
    bv_budget_value_cell = "C23"
    
    # Compliance cells
    dairy_milk_restrictions_label_cell = "B20"
    dairy_milk_restrictions_value_cell = "C20"
    lda_age_compliant_label_cell = "B21"
    lda_age_compliant_value_cell = "C21"
    
    # Find the rows for placement and target data
    placement_header_row = find_row_containing(sheet, "BV Placement Name", 25, 35) or find_row_containing(sheet, "Placement Name", 25, 35)
    if not placement_header_row:
        print("WARNING: Could not find placement header row")
        placement_header_row = 27  # Default to row 27 based on example
    else:
        print(f"Found placement header row at {placement_header_row}")
    
    target_header_row = find_row_containing(sheet, "BV ID", placement_header_row + 1, placement_header_row + 15) or find_row_containing(sheet, "BVID", placement_header_row + 1, placement_header_row + 15)
    if not target_header_row:
        print("WARNING: Could not find target header row")
        # Try to find a blank row after placements, then the next row with content
        for row in range(placement_header_row + 1, placement_header_row + 15):
            if all(get_cell_value(sheet, row, col) == "" for col in range(1, 10)):
                # Found blank row, next non-blank row might be target header
                for check_row in range(row + 1, row + 5):
                    if any(get_cell_value(sheet, check_row, col) != "" for col in range(1, 10)):
                        target_header_row = check_row
                        print(f"Found target header row at {target_header_row}")
                        break
                break
        
        if not target_header_row:
            target_header_row = 31  # Default to row 31 based on example
    else:
        print(f"Found target header row at {target_header_row}")
    
    # Calculate the placement data range
    placement_data_start_row = placement_header_row + 1
    placement_data_end_row = target_header_row - 2  # Assuming one blank row between placements and targets
    
    print(f"Placement data range: rows {placement_data_start_row} to {placement_data_end_row}")
    print(f"Target header row: {target_header_row}")
    
    # Step 1: Get campaign information
    # Campaign dates
    io_start_date = get_cell_value(sheet, int(io_start_date_label_cell[1:]), col_letter_to_number(io_start_date_label_cell[0]))
    io_start_date_value = get_cell_value(sheet, int(io_start_date_value_cell[1:]), col_letter_to_number(io_start_date_value_cell[0]))
    io_end_date = get_cell_value(sheet, int(io_end_date_label_cell[1:]), col_letter_to_number(io_end_date_label_cell[0]))
    io_end_date_value = get_cell_value(sheet, int(io_end_date_value_cell[1:]), col_letter_to_number(io_end_date_value_cell[0]))
    
    io_start_formatted = format_date(io_start_date_value)
    io_end_formatted = format_date(io_end_date_value)
    
    print(f"IO Campaign Start Date: {io_start_date} = {io_start_date_value} (formatted: {io_start_formatted})")
    print(f"IO Campaign End Date: {io_end_date} = {io_end_date_value} (formatted: {io_end_formatted})")
    
    # Budget
    bv_budget_label = get_cell_value(sheet, int(bv_budget_label_cell[1:]), col_letter_to_number(bv_budget_label_cell[0]))
    bv_budget_value = get_cell_value(sheet, int(bv_budget_value_cell[1:]), col_letter_to_number(bv_budget_value_cell[0]))
    bv_budget = clean_numeric(bv_budget_value)
    
    print(f"BV Budget: {bv_budget_label} = ${bv_budget:.2f} (raw: {bv_budget_value})")
    
    # Viewability
    viewability_contracted = get_cell_value(sheet, int(viewability_contracted_value_cell[1:]), col_letter_to_number(viewability_contracted_value_cell[0]))
    viewability_goal = get_cell_value(sheet, int(viewability_goal_value_cell[1:]), col_letter_to_number(viewability_goal_value_cell[0]))
    viewability_h13_value = get_cell_value(sheet, 13, 8)  # Get value from H13
    
    print(f"Viewability Contracted: {viewability_contracted}")
    print(f"Viewability Goal: {viewability_goal}")
    print(f"Viewability H13 Reference Value: {viewability_h13_value}")
    
    # Step 2: Get placement column indexes
    geo_required_col = None
    geo_details_col = None
    proj_start_date_col = None
    end_date_col = None
    traffic_info_col = None
    third_party_vendor_col = None
    
    for col in range(1, sheet.max_column + 1):
        header_value = get_cell_value(sheet, placement_header_row, col)
        if isinstance(header_value, str):
            header_lower = header_value.lower()
            if "geo required" in header_lower:
                geo_required_col = col
                print(f"Found Geo Required column at column {col}")
            elif "geo details" in header_lower:
                geo_details_col = col
                print(f"Found Geo Details column at column {col}")
            elif "start date" in header_lower:
                proj_start_date_col = col
                print(f"Found Start Date column at column {col}")
            elif "end date" in header_lower:
                end_date_col = col
                print(f"Found End Date column at column {col}")
            elif "traffic information" in header_lower or "traffic info" in header_lower:
                traffic_info_col = col
                print(f"Found Traffic Information column at column {col}")
            elif "third party vendor" in header_lower:
                third_party_vendor_col = col
                print(f"Found Third Party Vendor column at column {col}")
                
    # Use default values if columns weren't found (based on provided screenshot)
    if traffic_info_col is None:
        traffic_info_col = 4  # Column D
        print(f"Using default Traffic Information column: {traffic_info_col}")
    
    if third_party_vendor_col is None:
        third_party_vendor_col = 5  # Column E
        print(f"Using default Third Party Vendor column: {third_party_vendor_col}")
    
    # Step 3: Get target column indexes
    print("\nSearching target columns in row", target_header_row)
    sell_side_cpm_col = None
    impressions_col = None
    hh_unique_col = None
    
    # Debug: Print all column headers in target row
    for col in range(1, sheet.max_column + 1):
        header_value = get_cell_value(sheet, target_header_row, col)
        print(f"Target column {col}: '{header_value}'")
        
        if isinstance(header_value, str):
            header_lower = header_value.lower()
            if ("sell-side" in header_lower and "cpm" in header_lower) or header_lower == "cpm upcharge":
                sell_side_cpm_col = col
                print(f"Found Sell-Side CPM column at column {col}")
            elif "impressions" in header_lower:
                impressions_col = col
                print(f"Found Impressions column at column {col}")
            elif "hh/unique" in header_lower or "hh" in header_lower or "reach" in header_lower:
                hh_unique_col = col
                print(f"Found HH/Unique column at column {col}")
    
    # If we still haven't found the columns, let's try some likely defaults
    if sell_side_cpm_col is None:
        # Try looking for columns that might contain CPM values
        for col in range(1, sheet.max_column + 1):
            header = get_cell_value(sheet, target_header_row, col)
            if isinstance(header, str) and "cpm" in header.lower():
                sell_side_cpm_col = col
                print(f"Found possible CPM column at column {col} with header '{header}'")
                break
    
    if impressions_col is None:
        # Look for columns with "impression" substring
        for col in range(1, sheet.max_column + 1):
            header = get_cell_value(sheet, target_header_row, col)
            if isinstance(header, str) and "impress" in header.lower():
                impressions_col = col
                print(f"Found possible Impressions column at column {col} with header '{header}'")
                break
    
    # Step 4: Check viewability settings
    if viewability_contracted:
        # Convert to lowercase for case-insensitive comparison
        viewability_contracted_lower = viewability_contracted.lower() if isinstance(viewability_contracted, str) else ""
        
        if "no" in viewability_contracted_lower:
            # If contracted is No, H13 should be empty
            if viewability_h13_value:
                issues.append(f"Viewability Contracted is 'No' but H13 cell has value '{viewability_h13_value}'")
                highlight_cell(sheet, 16, 7, "FF0000")  # G16 (Red)
                print(f"✗ Viewability Contracted is 'No' but H13 has value '{viewability_h13_value}'")
            else:
                highlight_cell(sheet, 16, 7, "00FF00")  # G16 (Green)
                print("✓ Viewability Contracted is 'No' and H13 is empty")
        
        elif "yes" in viewability_contracted_lower:
            # If contracted is Yes, H13 should have a meaningful value
            if not viewability_h13_value:
                issues.append("Viewability Contracted is 'Yes' but H13 cell is empty")
                highlight_cell(sheet, 16, 7, "FF0000")  # G16 (Red)
                print("✗ Viewability Contracted is 'Yes' but H13 is empty")
            else:
                highlight_cell(sheet, 16, 7, "00FF00")  # G16 (Green)
                print(f"✓ Viewability Contracted is 'Yes' and H13 has value '{viewability_h13_value}'")
    
    # Step 4.5: Check Dairy-Milk Restrictions and LDA Age Compliant fields
    print("\nChecking Compliance Fields...")
    
    # Check Dairy-Milk Restrictions
    dairy_milk_label = get_cell_value(sheet, int(dairy_milk_restrictions_label_cell[1:]), col_letter_to_number(dairy_milk_restrictions_label_cell[0]))
    dairy_milk_value = get_cell_value(sheet, int(dairy_milk_restrictions_value_cell[1:]), col_letter_to_number(dairy_milk_restrictions_value_cell[0]))
    
    print(f"Dairy-Milk Restrictions: '{dairy_milk_label}' = '{dairy_milk_value}'")
    
    if not dairy_milk_value:
        issues.append("Dairy-Milk Restrictions value is empty")
        highlight_cell(sheet, int(dairy_milk_restrictions_value_cell[1:]), col_letter_to_number(dairy_milk_restrictions_value_cell[0]), "FF0000")  # Red
        print("✗ Dairy-Milk Restrictions value is empty")
    elif isinstance(dairy_milk_value, str) and dairy_milk_value.lower() in ["yes", "no"]:
        highlight_cell(sheet, int(dairy_milk_restrictions_value_cell[1:]), col_letter_to_number(dairy_milk_restrictions_value_cell[0]), "00FF00")  # Green
        print(f"✓ Dairy-Milk Restrictions value is properly filled with '{dairy_milk_value}'")
    else:
        issues.append(f"Dairy-Milk Restrictions has unexpected value: '{dairy_milk_value}' (should be 'Yes' or 'No')")
        highlight_cell(sheet, int(dairy_milk_restrictions_value_cell[1:]), col_letter_to_number(dairy_milk_restrictions_value_cell[0]), "FFFF00")  # Yellow
        print(f"⚠ Dairy-Milk Restrictions has unexpected value: '{dairy_milk_value}'")
    
    # Check LDA or Age Compliant
    lda_age_label = get_cell_value(sheet, int(lda_age_compliant_label_cell[1:]), col_letter_to_number(lda_age_compliant_label_cell[0]))
    lda_age_value = get_cell_value(sheet, int(lda_age_compliant_value_cell[1:]), col_letter_to_number(lda_age_compliant_value_cell[0]))
    
    print(f"LDA or Age Compliant: '{lda_age_label}' = '{lda_age_value}'")
    
    if not lda_age_value:
        issues.append("LDA or Age Compliant value is empty")
        highlight_cell(sheet, int(lda_age_compliant_value_cell[1:]), col_letter_to_number(lda_age_compliant_value_cell[0]), "FF0000")  # Red
        print("✗ LDA or Age Compliant value is empty")
    elif isinstance(lda_age_value, str) and lda_age_value.lower() in ["yes", "no"]:
        highlight_cell(sheet, int(lda_age_compliant_value_cell[1:]), col_letter_to_number(lda_age_compliant_value_cell[0]), "00FF00")  # Green
        print(f"✓ LDA or Age Compliant value is properly filled with '{lda_age_value}'")
    else:
        issues.append(f"LDA or Age Compliant has unexpected value: '{lda_age_value}' (should be 'Yes' or 'No')")
        highlight_cell(sheet, int(lda_age_compliant_value_cell[1:]), col_letter_to_number(lda_age_compliant_value_cell[0]), "FFFF00")  # Yellow
        print(f"⚠ LDA or Age Compliant has unexpected value: '{lda_age_value}'")
    
    # Step 5: Check placement flight dates and geo requirements
    start_date_matches = 0
    end_date_matches = 0
    placement_count = 0
    date_outside_range_issues = []
    
    if geo_required_col and geo_details_col and proj_start_date_col and end_date_col and traffic_info_col and third_party_vendor_col:
        print("\nChecking placement data...")
        
        for row in range(placement_data_start_row, placement_data_end_row + 1):
            placement_name = get_cell_value(sheet, row, 2)  # Column B (2)
            if not placement_name:
                print(f"Row {row}: Empty placement name, skipping")
                continue
            
            placement_count += 1
            print(f"\nPlacement {placement_count}: {placement_name}")
            
            # Check flight dates
            placement_start = get_cell_value(sheet, row, proj_start_date_col)
            placement_end = get_cell_value(sheet, row, end_date_col)
            
            placement_start_formatted = format_date(placement_start)
            placement_end_formatted = format_date(placement_end)
            
            print(f"Placement dates: {placement_start_formatted} to {placement_end_formatted}")
            print(f"Campaign dates:  {io_start_formatted} to {io_end_formatted}")
            
            # Check start date match with IO Campaign Start Date
            if placement_start_formatted == io_start_formatted:
                start_date_matches += 1
                highlight_cell(sheet, row, proj_start_date_col, "00FF00")  # Green
                print(f"✓ Start date matches IO Campaign Start Date")
            else:
                # Check if placement start date is outside IO campaign date range
                if compare_dates(placement_start_formatted, io_start_formatted) < 0:
                    date_outside_range_issues.append(f"Placement '{placement_name}': Start date ({placement_start_formatted}) is before IO Campaign Start Date ({io_start_formatted})")
                    highlight_cell(sheet, row, proj_start_date_col, "FF0000")  # Red
                    print(f"✗ Start date is before IO Campaign Start Date")
                elif compare_dates(placement_start_formatted, io_end_formatted) > 0:
                    date_outside_range_issues.append(f"Placement '{placement_name}': Start date ({placement_start_formatted}) is after IO Campaign End Date ({io_end_formatted})")
                    highlight_cell(sheet, row, proj_start_date_col, "FF0000")  # Red
                    print(f"✗ Start date is after IO Campaign End Date")
                else:
                    # Start date is within range but doesn't match IO start date
                    highlight_cell(sheet, row, proj_start_date_col, "FFFF00")  # Yellow (warning)
                    print(f"⚠ Start date doesn't match IO Campaign Start Date but is within range")
            
            # Check end date match with IO Campaign End Date
            if placement_end_formatted == io_end_formatted:
                end_date_matches += 1
                highlight_cell(sheet, row, end_date_col, "00FF00")  # Green
                print(f"✓ End date matches IO Campaign End Date")
            else:
                # Check if placement end date is outside IO campaign date range
                if compare_dates(placement_end_formatted, io_start_formatted) < 0:
                    date_outside_range_issues.append(f"Placement '{placement_name}': End date ({placement_end_formatted}) is before IO Campaign Start Date ({io_start_formatted})")
                    highlight_cell(sheet, row, end_date_col, "FF0000")  # Red
                    print(f"✗ End date is before IO Campaign Start Date")
                elif compare_dates(placement_end_formatted, io_end_formatted) > 0:
                    date_outside_range_issues.append(f"Placement '{placement_name}': End date ({placement_end_formatted}) is after IO Campaign End Date ({io_end_formatted})")
                    highlight_cell(sheet, row, end_date_col, "FF0000")  # Red
                    print(f"✗ End date is after IO Campaign End Date")
                else:
                    # End date is within range but doesn't match IO end date
                    highlight_cell(sheet, row, end_date_col, "FFFF00")  # Yellow (warning)
                    print(f"⚠ End date doesn't match IO Campaign End Date but is within range")
            
            # Check Geo Requirements
            geo_required = get_cell_value(sheet, row, geo_required_col)
            geo_details = get_cell_value(sheet, row, geo_details_col)
            
            print(f"Geo Required: '{geo_required}', Geo Details: '{geo_details}'")
            
            # Convert to lowercase for case-insensitive comparison
            geo_required_lower = geo_required.lower() if isinstance(geo_required, str) else ""
            geo_details_lower = geo_details.lower() if isinstance(geo_details, str) else ""
            
            if "yes" in geo_required_lower:
                # If Yes, geo details should have meaningful content (not empty, NA, or National)
                if not geo_details or geo_details_lower in ["", "na", "national"]:
                    issues.append(f"Placement '{placement_name}': Geo Required is 'Yes' but Geo Details is empty/NA/National")
                    highlight_cell(sheet, row, geo_details_col, "FF0000")  # Red
                    print(f"✗ Geo Required is 'Yes' but Geo Details is '{geo_details}'")
                else:
                    highlight_cell(sheet, row, geo_details_col, "00FF00")  # Green
                    print(f"✓ Geo Required is 'Yes' and Geo Details is '{geo_details}'")
            
            elif "no" in geo_required_lower:
                # If No, geo details should be empty, NA, or National
                if geo_details and geo_details_lower not in ["", "na", "national"]:
                    issues.append(f"Placement '{placement_name}': Geo Required is 'No' but Geo Details has value '{geo_details}'")
                    highlight_cell(sheet, row, geo_details_col, "FF0000")  # Red
                    print(f"✗ Geo Required is 'No' but Geo Details has value '{geo_details}'")
                else:
                    highlight_cell(sheet, row, geo_details_col, "00FF00")  # Green
                    print(f"✓ Geo Required is 'No' and Geo Details is appropriate")
            
            else:
                # Geo Required field is empty or invalid
                issues.append(f"Placement '{placement_name}': Geo Required field is empty or invalid")
                highlight_cell(sheet, row, geo_required_col, "FF0000")  # Red
                print(f"✗ Geo Required field is empty or invalid")
                
            # Check Traffic Information
            traffic_info = get_cell_value(sheet, row, traffic_info_col)
            print(f"Traffic Information: '{traffic_info}'")
            
            # Check if Traffic Information is filled
            if not traffic_info:
                issues.append(f"Placement '{placement_name}': Traffic Information is empty")
                highlight_cell(sheet, row, traffic_info_col, "FF0000")  # Red
                print(f"✗ Traffic Information is empty")
            elif isinstance(traffic_info, str):
                traffic_info_lower = traffic_info.lower()
                if traffic_info_lower in ["yes", "no"]:
                    highlight_cell(sheet, row, traffic_info_col, "00FF00")  # Green
                    print(f"✓ Traffic Information is filled with '{traffic_info}'")
                else:
                    # Add warning for unexpected values
                    highlight_cell(sheet, row, traffic_info_col, "FFFF00")  # Yellow
                    print(f"⚠ Traffic Information has unexpected value: '{traffic_info}'")
            else:
                # Non-string value
                highlight_cell(sheet, row, traffic_info_col, "FFFF00")  # Yellow
                print(f"⚠ Traffic Information has non-text value: '{traffic_info}'")
            
            # Check Third Party Vendor - only required if Traffic Information is "Yes"
            third_party_vendor = get_cell_value(sheet, row, third_party_vendor_col)
            print(f"Third Party Vendor: '{third_party_vendor}'")
            
            if isinstance(traffic_info, str) and traffic_info.lower() == "yes":
                if not third_party_vendor:
                    issues.append(f"Placement '{placement_name}': Traffic Information is 'Yes' but Third Party Vendor is empty")
                    highlight_cell(sheet, row, third_party_vendor_col, "FF0000")  # Red
                    print(f"✗ Traffic Information is 'Yes' but Third Party Vendor is empty")
                else:
                    highlight_cell(sheet, row, third_party_vendor_col, "00FF00")  # Green
                    print(f"✓ Traffic Information is 'Yes' and Third Party Vendor is filled")
            else:
                # If Traffic Information is not "Yes", Third Party Vendor is optional
                if third_party_vendor:
                    highlight_cell(sheet, row, third_party_vendor_col, "00FF00")  # Green (filled but optional)
                    print(f"✓ Third Party Vendor is optional but filled: '{third_party_vendor}'")
                else:
                    # Empty but optional
                    highlight_cell(sheet, row, third_party_vendor_col, "FFFF00")  # Yellow (empty but optional)
                    print(f"⚠ Third Party Vendor is empty but optional for this placement")
    
    # Step 6: Check impressions and calculate budget
    total_calculated_budget = 0

    # Enhanced target data processing
    if target_header_row:
        print("\nChecking target data and calculating budget...")
        
        # Look for required target data in a wider range of rows
        max_target_row = min(target_header_row + 30, sheet.max_row)
        
        # If we don't have column indices yet, set them based on the template
        if sell_side_cpm_col is None:
            sell_side_cpm_col = 11  # Column K (Sell-side CPM)
            print(f"Using default CPM column: {sell_side_cpm_col}")
        
        if impressions_col is None:
            impressions_col = 9  # Column I (Impressions)
            print(f"Using default Impressions column: {impressions_col}")
        
        if hh_unique_col is None:
            hh_unique_col = 10  # Column J (HH/Unique Reach)
            print(f"Using default HH/Unique column: {hh_unique_col}")
        
        target_count = 0
        for row in range(target_header_row + 1, max_target_row):
            # Check if this row contains a BVT ID in column D
            bvt_id = get_cell_value(sheet, row, 4)  # Column D (4)
            
            # Continue if empty or not a string
            if not isinstance(bvt_id, str):
                continue
                
            # Check if the string starts with "BVT"
            if bvt_id.startswith("BVT"):
                target_count += 1
                print(f"\nTarget row {row} - BVT ID: {bvt_id}")
                
                # Get raw values for processing
                cpm_raw = get_cell_value(sheet, row, sell_side_cpm_col)
                impressions_raw = get_cell_value(sheet, row, impressions_col)
                reach_raw = get_cell_value(sheet, row, hh_unique_col) if hh_unique_col else 0
                
                # Clean and convert values
                cpm = clean_numeric(cpm_raw)
                impressions = clean_numeric(impressions_raw)
                reach = clean_numeric(reach_raw)
                
                print(f"Cleaned values - CPM: {cpm}, Impressions: {impressions}, Reach: {reach}")
                
                # Check impressions vs reach if both are available
                if impressions > 0 and reach > 0:
                    if impressions <= reach:
                        issues.append(f"Target {bvt_id}: Impressions ({impressions}) not greater than HH/Unique Reach ({reach})")
                        highlight_cell(sheet, row, impressions_col, "FF0000")  # Red
                        print(f"✗ Impressions ({impressions}) not greater than Reach ({reach})")
                    else:
                        highlight_cell(sheet, row, impressions_col, "00FF00")  # Green
                        print(f"✓ Impressions ({impressions}) greater than Reach ({reach})")
                
                # Calculate budget contribution
                if cpm > 0 and impressions > 0:
                    row_budget = (impressions * cpm) / 1000
                    total_calculated_budget += row_budget
                    print(f"Row budget: ${row_budget:.2f}, Running total: ${total_calculated_budget:.2f}")
        
        print(f"\nTotal targets processed: {target_count}")
    
    # Step 7: Check if calculated budget matches BV Budget
    if bv_budget > 0:
        budget_diff = abs(total_calculated_budget - bv_budget)
        budget_diff_pct = (budget_diff / bv_budget) * 100 if bv_budget > 0 else 0
        
        print(f"\nTotal Calculated Budget: ${total_calculated_budget:.2f}")
        print(f"BV Budget: ${bv_budget:.2f}")
        print(f"Difference: ${budget_diff:.2f} ({budget_diff_pct:.2f}%)")
        
        if budget_diff_pct > 1:  # Allow 1% tolerance
            issues.append(f"Budget Mismatch: Calculated (${total_calculated_budget:.2f}) vs. BV Budget (${bv_budget:.2f}), diff: ${budget_diff:.2f}")
            highlight_cell(sheet, int(bv_budget_value_cell[1:]), col_letter_to_number(bv_budget_value_cell[0]), "FF0000")  # Red
        else:
            highlight_cell(sheet, int(bv_budget_value_cell[1:]), col_letter_to_number(bv_budget_value_cell[0]), "00FF00")  # Green
    
    # Step 8: Check flight dates across placements - Updated logic
    print(f"\nPlacement count: {placement_count}")
    print(f"Start date matches: {start_date_matches}/{placement_count}")
    print(f"End date matches: {end_date_matches}/{placement_count}")
    
    # Add any date range issues
    if date_outside_range_issues:
        issues.extend(date_outside_range_issues)
        print("\nDate range issues:")
        for issue in date_outside_range_issues:
            print(f"- {issue}")
    
    # New date checking logic based on updated requirements
    date_issues = []
    
    # 1. Check if at least one placement's start date matches the IO Campaign Start Date
    if start_date_matches == 0:
        date_issues.append(f"No placement start date matches IO Campaign Start Date ({io_start_formatted})")
    
    # 2. Check if at least one placement's end date matches the IO Campaign End Date
    if end_date_matches == 0:
        date_issues.append(f"No placement end date matches IO Campaign End Date ({io_end_formatted})")
    
    # If there are date issues, add them to the main issues list
    if date_issues:
        issues.extend(date_issues)
    
    # Highlight IO Start/End Date cells based on whether at least one placement matches each
    highlight_cell(sheet, int(io_start_date_value_cell[1:]), col_letter_to_number(io_start_date_value_cell[0]), 
                  "00FF00" if start_date_matches > 0 else "FF0000")
    
    highlight_cell(sheet, int(io_end_date_value_cell[1:]), col_letter_to_number(io_end_date_value_cell[0]), 
                  "00FF00" if end_date_matches > 0 else "FF0000")
    
    # Save the highlighted file
    output_file = file_path.replace('.xlsx', '_QA_issues.xlsx')
    wb.save(output_file)
    
    print("\nQA ISSUES FOUND:")
    if issues:
        for i, issue in enumerate(issues, 1):
            print(f"{i}. {issue}")
    else:
        print("No issues found!")
    
    print(f"\nReport saved to {output_file}")
    
    return issues

if __name__ == "__main__":
    # Default file path
    default_file_path = r"C:\QA_auto_check\campaign_brief.xlsx"
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = default_file_path
    
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)
    
    run_qa_checks(file_path)