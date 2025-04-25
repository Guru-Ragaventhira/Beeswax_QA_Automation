import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import re
import os
import sys
from datetime import datetime

def highlight_cell(sheet, row, col, color="FFFF00"):
    """Highlight a cell with the specified color"""
    cell = sheet.cell(row=row, column=col)
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.fill = fill

def find_cell_position(sheet, text, start_row=1, end_row=None):
    """Find the position of a cell containing text"""
    if end_row is None:
        end_row = sheet.max_row
        
    for row in range(start_row, end_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value
            if cell_value and isinstance(cell_value, str) and text.lower() in cell_value.lower():
                print(f"Found '{text}' at row {row}, column {col}")
                return row, col
    print(f"Could not find '{text}' in rows {start_row}-{end_row}")
    return None, None

def find_header_row(sheet, text_marker):
    """Find a row containing a specific text marker (header row)"""
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str) and text_marker.lower() in cell_value.lower():
                return row
    return None

def get_cell_value(sheet, row, col):
    """Get cell value, handling None"""
    if row is None or col is None:
        return ""
    value = sheet.cell(row=row, column=col).value
    return value if value is not None else ""

def format_date(date_value):
    """Format dates consistently for comparison"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%#m/%#d/%Y')  # Windows format
    elif isinstance(date_value, str):
        # Try to parse and standardize the date format
        try:
            date_obj = datetime.strptime(date_value, '%m/%d/%Y')
            return date_obj.strftime('%#m/%#d/%Y')
        except ValueError:
            try:
                date_obj = datetime.strptime(date_value, '%#m/%#d/%Y')
                return date_obj.strftime('%#m/%#d/%Y')
            except ValueError:
                pass
    return str(date_value).strip()

def clean_numeric(value):
    """Convert string numbers with formatting to float values"""
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        return float(value)
    
    # Remove dollar signs, commas, and extra spaces
    clean_value = str(value).replace('$', '').replace(',', '').strip()
    
    # Try to convert to float
    try:
        return float(clean_value)
    except ValueError:
        print(f"WARNING: Could not convert '{value}' to a number")
        return 0.0

def find_column_in_row(sheet, row, text):
    """Find column in a specific row containing text"""
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value and isinstance(cell_value, str) and text.lower() in cell_value.lower():
            print(f"Found column '{text}' at column {col} in row {row}")
            return col
    print(f"Could not find column '{text}' in row {row}")
    return None

def run_qa_checks(file_path):
    """Run QA checks on the campaign brief"""
    print(f"Running QA checks on {file_path}...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    issues = []
    io_start_formatted = None
    io_end_formatted = None  # Initialize this variable
    
    # Finding key marker rows
    placement_header_row = None
    target_header_row = None
    
    # Find BV Placement Name row and Campaign ID row
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = get_cell_value(sheet, row, col)
            if isinstance(cell_value, str):
                if "BV Placement Name" in cell_value:
                    placement_header_row = row
                    print(f"Found BV Placement Name header at row {row}")
                elif "BV ID" in cell_value:
                    target_header_row = row
                    print(f"Found BV ID header at row {row}")
    
    # Find campaign dates
    io_start_row, io_start_col = find_cell_position(sheet, "IO Campaign Start Date")
    io_end_row, io_end_col = find_cell_position(sheet, "IO Campaign End Date")
    
    # Flight Dates Check
    if io_start_row and io_start_col:
        io_start_date = get_cell_value(sheet, io_start_row, io_start_col + 1)
        io_start_formatted = format_date(io_start_date)
        print(f"IO Start Date: {io_start_date} (formatted: {io_start_formatted})")
    else:
        issues.append("Could not find IO Campaign Start Date")
    
    if io_end_row and io_end_col:
        io_end_date = get_cell_value(sheet, io_end_row, io_end_col + 1)
        io_end_formatted = format_date(io_end_date)
        print(f"IO End Date: {io_end_date} (formatted: {io_end_formatted})")
    else:
        issues.append("Could not find IO Campaign End Date")
        io_end_formatted = None  # Set a default value
    
    # Find column indexes in placement section
    if placement_header_row:
        proj_start_col = find_column_in_row(sheet, placement_header_row, "Projected Start Date")
        end_date_col = find_column_in_row(sheet, placement_header_row, "End Date")
    
    # Find column indexes in target section (where BV ID is)
    if target_header_row:
        sell_side_cpm_col = find_column_in_row(sheet, target_header_row, "Sell-Side CPM")
        impressions_col = find_column_in_row(sheet, target_header_row, "Impressions")
        reach_col = find_column_in_row(sheet, target_header_row, "HH/Unique")
        
        print(f"Found columns in target section - CPM: {sell_side_cpm_col}, Impressions: {impressions_col}, Reach: {reach_col}")
    
    # Process each placement row
    total_calculated_budget = 0
    start_date_match_found = False
    end_date_match_found = False
    
    if placement_header_row and all([proj_start_col, end_date_col]):
        for row in range(placement_header_row + 1, sheet.max_row + 1):
            placement_name = get_cell_value(sheet, row, 1)
            if not placement_name:
                continue
            
            # Check flight dates
            placement_start = get_cell_value(sheet, row, proj_start_col)
            placement_end = get_cell_value(sheet, row, end_date_col)
            
            if placement_start and placement_end:
                placement_start_formatted = format_date(placement_start)
                placement_end_formatted = format_date(placement_end)
                
                # Check start date match
                if io_start_formatted and placement_start_formatted == io_start_formatted:
                    start_date_match_found = True
                    highlight_cell(sheet, row, proj_start_col, "00FF00")  # Green
                else:
                    highlight_cell(sheet, row, proj_start_col, "FFFF00")  # Yellow
                
                # Check end date match
                if io_end_formatted and placement_end_formatted == io_end_formatted:
                    end_date_match_found = True
                    highlight_cell(sheet, row, end_date_col, "00FF00")  # Green
                else:
                    highlight_cell(sheet, row, end_date_col, "FFFF00")  # Yellow
    
    # Check impressions and calculate budget only if we found the required columns
    if target_header_row and all([sell_side_cpm_col, impressions_col, reach_col]):
        for row in range(target_header_row + 1, sheet.max_row + 1):
            placement_name = get_cell_value(sheet, row, 1)
            if not placement_name:
                continue
            
            # Check impressions vs reach
            impressions = clean_numeric(get_cell_value(sheet, row, impressions_col))
            reach = clean_numeric(get_cell_value(sheet, row, reach_col))
            cpm = clean_numeric(get_cell_value(sheet, row, sell_side_cpm_col))
            
            if impressions > 0 and reach > 0:
                if impressions <= reach:
                    issues.append(f"Impressions ({impressions}) not greater than HH/Unique Reach ({reach}) for placement '{placement_name}'")
                    highlight_cell(sheet, row, impressions_col, "FF0000")  # Red
                else:
                    highlight_cell(sheet, row, impressions_col, "00FF00")  # Green
            
            # Calculate budget
            if impressions > 0 and cpm > 0:
                row_budget = (impressions * cpm) / 1000
                total_calculated_budget += row_budget
    
    # Flight dates summary
    if io_start_formatted and not start_date_match_found:
        issues.append(f"No placement start date matches IO Campaign Start Date ({io_start_formatted})")
    if io_end_formatted and not end_date_match_found:
        issues.append(f"No placement end date matches IO Campaign End Date ({io_end_formatted})")
    
    # Save the highlighted file
    output_file = file_path.replace('.xlsx', '_QA_issues.xlsx')
    wb.save(output_file)
    print(f"\nReport saved to {output_file}")
    
    return issues

if __name__ == "__main__":
    # Default file path
    default_file_path = r"C:\QA_auto_check\campaign_brief.xlsx"
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = default_file_path
    
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)
    
    run_qa_checks(file_path)

def run_qa_checks(sheet):
    issues = []
    header_row = find_header_row(sheet, "Placement")
    headers = {sheet.cell(header_row, col).value.strip(): col for col in range(1, sheet.max_column + 1) if sheet.cell(header_row, col).value}

    # Required columns
    required_fields = ["Impressions", "HH/Unique Reach", "Sell-Side CPM", "Geo Required", "Geo Details",
                       "Viewability Contracted", "Viewability Goal", "Projected Start Date", "End Date"]
    col_idx = {field: headers.get(field) for field in required_fields if headers.get(field)}

    io_start_row, io_start_col = find_cell_position(sheet, "IO Campaign Start Date")
    io_end_row, io_end_col = find_cell_position(sheet, "IO Campaign End Date")

    io_start = format_date(sheet.cell(io_start_row, io_start_col+1).value)
    io_end = format_date(sheet.cell(io_end_row, io_end_col+1).value)

    placement_dates = []
    start_match, end_match = False, False

    for row in range(header_row+1, sheet.max_row + 1):
        def val(field): return sheet.cell(row, col_idx.get(field)).value if col_idx.get(field) else None

        # Impression vs Reach
        try:
            imp = int(val("Impressions") or 0)
            reach = int(val("HH/Unique Reach") or 0)
            if imp <= reach:
                highlight_cell(sheet, row, col_idx["Impressions"], "FFC7CE")
                highlight_cell(sheet, row, col_idx["HH/Unique Reach"], "FFC7CE")
                issues.append(f"Row {row}: Impressions not greater than Reach")
            else:
                highlight_cell(sheet, row, col_idx["Impressions"], "C6EFCE")
                highlight_cell(sheet, row, col_idx["HH/Unique Reach"], "C6EFCE")
        except: pass

        # Spend check - store for BV Budget comparison
        try:
            cpm = float(val("Sell-Side CPM") or 0)
            spend = (imp * cpm) / 1000
            sheet.cell(row, sheet.max_column + 1).value = spend  # Temporary storage
        except: pass

        # Geo Required check
        geo_required = str(val("Geo Required")).strip().lower()
        geo_detail = str(val("Geo Details")).strip().lower()
        if geo_required == "no":
            if geo_detail not in ["", "na", "national"]:
                highlight_cell(sheet, row, col_idx["Geo Details"], "FFC7CE")
                issues.append(f"Row {row}: Geo Required is 'No' but Geo Details is not empty/NA/National")
        elif geo_required == "yes":
            if geo_detail in ["", "na", "national"]:
                highlight_cell(sheet, row, col_idx["Geo Details"], "FFC7CE")
                issues.append(f"Row {row}: Geo Required is 'Yes' but Geo Details is empty or NA/National")

        # Viewability check
        view_contracted = str(val("Viewability Contracted")).strip().lower()
        view_goal = str(val("Viewability Goal")).strip()
        if view_contracted == "no":
            if re.search(r"[a-zA-Z]", view_goal):
                highlight_cell(sheet, row, col_idx["Viewability Goal"], "FFC7CE")
                issues.append(f"Row {row}: Viewability is No but Goal contains text")
        elif view_contracted == "yes":
            if not re.search(r"\d+%", view_goal):
                highlight_cell(sheet, row, col_idx["Viewability Goal"], "FFC7CE")
                issues.append(f"Row {row}: Viewability is Yes but Goal not like '70%'")

        # Flight Dates check
        try:
            start = format_date(val("Projected Start Date"))
            end = format_date(val("End Date"))
            placement_dates.append((start, end))

            if start == io_start:
                start_match = True
            if end == io_end:
                end_match = True
            if start < io_start or end > io_end:
                highlight_cell(sheet, row, col_idx["Projected Start Date"], "FFC7CE")
                highlight_cell(sheet, row, col_idx["End Date"], "FFC7CE")
                issues.append(f"Row {row}: Placement dates out of IO range")
        except: pass

    # Overall flight date validation for multiple placements
    if len(placement_dates) > 1:
        if not start_match:
            issues.append("No placement start date matches IO campaign start date")
        if not end_match:
            issues.append("No placement end date matches IO campaign end date")
    elif len(placement_dates) == 1:
        s, e = placement_dates[0]
        if s != io_start or e != io_end:
            issues.append("Single placement dates do not match IO campaign dates")

    print("\n".join(issues))
    return issues
