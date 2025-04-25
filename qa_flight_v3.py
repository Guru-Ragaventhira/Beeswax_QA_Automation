import pandas as pd
import numpy as np
from datetime import datetime
import os
import re
import glob
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import the function from brief_extractor
from brief_extractor import extract_structured_brief_data

def find_latest_qa_report(output_dir):
    """Find the latest QA report file in the output directory"""
    qa_report_files = glob.glob(os.path.join(output_dir, "qa_report_*.xlsx"))
    if not qa_report_files:
        return None
    
    # Sort by modification time, newest first
    latest_file = max(qa_report_files, key=os.path.getmtime)
    return latest_file

def safe_date_convert(date_val):
    """Safely convert various date formats to pandas datetime"""
    if pd.isna(date_val):
        return None
    
    if isinstance(date_val, (datetime, pd.Timestamp)):
        return date_val.replace(tzinfo=None) # Remove timezone if present
    
    # Try parsing date string with various formats
    try:
        # First try pandas to_datetime with default parser
        dt = pd.to_datetime(date_val)
        return dt.replace(tzinfo=None) # Remove timezone
    except Exception as e1:
        # print(f"Debug: Failed default parse for {date_val}: {e1}") # Keep commented unless debugging
        # Handle Excel date as float (days since 1900-01-01)
        try:
            if isinstance(date_val, (int, float)):
                # Excel's epoch starts on 1900-01-01, but treats 1900 as a leap year incorrectly.
                # pandas uses 1899-12-30 as the base for 'excel' origin.
                dt = pd.to_datetime('1899-12-30') + pd.Timedelta(days=float(date_val))
                return dt.replace(tzinfo=None) # Remove timezone
        except Exception as e2:
            # print(f"Debug: Failed Excel float parse for {date_val}: {e2}") # Keep commented unless debugging
            pass
            
        # Try common date formats explicitly
        date_formats = [
            '%Y-%m-%d %H:%M:%S', # Handle timestamp format if present
            '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y', '%m-%d-%Y',
            '%Y/%m/%d', '%b %d, %Y', '%d %b %Y',
            '%m/%d/%y', # Handle short year format
        ]
        
        original_val_str = str(date_val) # Convert to string for parsing

        for fmt in date_formats:
            try:
                dt = pd.to_datetime(original_val_str, format=fmt)
                return dt.replace(tzinfo=None) # Remove timezone
            except Exception as e3:
                # print(f"Debug: Failed format '{fmt}' for {original_val_str}: {e3}") # Keep commented unless debugging
                continue
                
    # If all attempts fail, return None
    # print(f"Warning: Could not convert '{date_val}' to datetime.") # Keep commented unless debugging
    return None

def main():
    print("Starting flight date verification...")
    
    # Load environment variables
    env_path = os.getenv("ENV_PATH", "./input_folder/beeswax_input_qa.env")
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        load_dotenv(env_path)
    else:
        print(f"Warning: Environment file {env_path} not found. Using default environment.")
        load_dotenv()  # Try default locations
    
    # Get paths from environment or use defaults
    brief_path = os.path.abspath(os.getenv("BRIEF_PATH", "./Brief/Campaign_Brief.xlsx"))
    output_dir = os.path.abspath(os.getenv("OUTPUT_DIR", "./output_folder"))
    qa_flight_output_path = os.path.abspath(os.getenv("QA_FLIGHT_OUTPUT_PATH", "./qa_flight_v3_output.xlsx"))
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Determine QA report path - first check env variable, then find latest
    qa_report_path = os.getenv("QA_REPORT_PATH")
    if not qa_report_path or not os.path.exists(qa_report_path):
        print(f"QA report not found at {qa_report_path}, looking for the latest report...")
        latest_report = find_latest_qa_report(output_dir)
        if latest_report:
            qa_report_path = latest_report
            print(f"Using latest QA report: {qa_report_path}")
        else:
            # Fallback to default path in current directory
            qa_report_path = "./qa_report.xlsx"
            print(f"No reports found in output directory, using default path: {qa_report_path}")
    
    # Resolve to absolute paths
    qa_report_path = os.path.abspath(qa_report_path)
    output_path = qa_flight_output_path  # Use the output path from env variables
    
    print(f"Using paths:")
    print(f"  QA Report: {qa_report_path}")
    print(f"  Brief: {brief_path}")
    print(f"  Output: {output_path}")
    
    # Check if files exist
    if not os.path.exists(qa_report_path):
        print(f"Error: QA report file not found at {qa_report_path}")
        return
    if not os.path.exists(brief_path):
        print(f"Error: Campaign brief file not found at {brief_path}")
        return
    
    # --- Load QA Report ---
    print(f"Loading QA report from {qa_report_path}")
    try:
        qa_df_full = pd.read_excel(qa_report_path)
        # Define required columns from QA report
        qa_cols_needed = [
            'campaign_id', 'campaign_name', 'campaign_start_date', 'campaign_end_date', 
            'line_item_id', 'line_item_name', 'line_item_start_date', 'line_item_end_date', 
            'line_item_alternative_id' # This is the BVT ID
        ]
        # Check if all needed columns exist
        missing_qa_cols = [col for col in qa_cols_needed if col not in qa_df_full.columns]
        if missing_qa_cols:
             print(f"Error: Missing required columns in QA report: {', '.join(missing_qa_cols)}")
             return
        # Select only the needed columns
        qa_df = qa_df_full[qa_cols_needed].copy()
        print(f"QA Report loaded. Shape: {qa_df.shape}")
    except Exception as e:
        print(f"Error loading QA report: {e}")
        return
    
    # --- Extract Structured Data from Brief ---
    print(f"Extracting structured data from brief: {brief_path}")
    try:
        brief_data = extract_structured_brief_data(brief_path)
        if not brief_data:
            print("Error: Failed to extract any data from the brief.")
            return
    except Exception as e:
        print(f"Error extracting brief data: {e}")
        return
    
    # --- Process Campaign Level Data ---
    campaign_data = brief_data.get('campaign_data') # campaign_data can be a DataFrame or None
    sf_campaign_start_date = None
    sf_campaign_end_date = None
    # Check if campaign_data is a non-empty DataFrame
    if isinstance(campaign_data, pd.DataFrame) and not campaign_data.empty:
        try:
            # Set 'Field' column as index for easy lookup
            campaign_data_indexed = campaign_data.set_index('Field')
            # Use .loc to get the 'Value' for specific fields, handle potential KeyError
            sf_campaign_start_date_raw = campaign_data_indexed.loc['IO Campaign Start Date', 'Value']
            sf_campaign_end_date_raw = campaign_data_indexed.loc['IO Campaign End Date', 'Value']
            
            sf_campaign_start_date = safe_date_convert(sf_campaign_start_date_raw)
            sf_campaign_end_date = safe_date_convert(sf_campaign_end_date_raw)
            
            print(f"Brief Campaign Start Date: {sf_campaign_start_date}")
            print(f"Brief Campaign End Date: {sf_campaign_end_date}")
        except KeyError as e:
            print(f"Warning: Could not find key {e} in campaign data DataFrame.")
        except Exception as e:
            print(f"Error processing campaign data DataFrame: {e}")
    elif campaign_data is not None:
        # Handle the case where it might be something else (e.g., empty dict/list) or log a warning
        print(f"Warning: Campaign data extracted but is not a valid DataFrame. Type: {type(campaign_data)}")
    else:
        print("Warning: Could not extract campaign data from brief.")
        
    # Add brief campaign dates to the QA DataFrame
    qa_df['sf_campaign_start_date'] = sf_campaign_start_date
    qa_df['sf_campaign_end_date'] = sf_campaign_end_date
    
    # --- Process Target Level Data (BVT -> BVP Mapping) ---
    target_data = brief_data.get('target_data') # target_data can be list or DataFrame
    bvt_bvp_map = {}
    # Check if target_data is a non-empty DataFrame
    if isinstance(target_data, pd.DataFrame) and not target_data.empty:
        print("Processing Target Data as DataFrame...")
        # Assuming columns are named 'BVT' and 'BVP' from extractor
        expected_cols = ['BVT', 'BVP']
        if all(col in target_data.columns for col in expected_cols):
            for row in target_data.itertuples(index=False):
                try:
                    bvt = str(row.BVT).strip()
                    bvp = str(row.BVP).strip()
                    # Check BVT looks like a BVT ID (optional but good practice)
                    if bvt and bvp and re.match(r'BVT\d+', bvt, re.IGNORECASE): 
                        bvt_bvp_map[bvt] = bvp
                except AttributeError as e:
                    print(f"Warning: Missing expected column in target_data row: {e}")
                    continue # Skip this row
        else:
            print(f"Warning: Target data DataFrame missing expected columns ('{expected_cols}'). Columns found: {list(target_data.columns)}")
    # Check if target_data is a non-empty list
    elif isinstance(target_data, list) and target_data:
        print("Processing Target Data as List...")
        for target_item in target_data:
            if isinstance(target_item, dict) and 'BVT' in target_item and 'BVP' in target_item:
                bvt = str(target_item.get('BVT', '')).strip()
                bvp = str(target_item.get('BVP', '')).strip()
                if bvt and bvp and re.match(r'BVT\d+', bvt, re.IGNORECASE):
                    bvt_bvp_map[bvt] = bvp
    else:
        print("Warning: Could not extract valid Target Data (list or DataFrame) from brief.")
        
    print(f"Found {len(bvt_bvp_map)} BVT to BVP mappings.")
         
    # --- Process Placement Level Data (BVP -> Dates Mapping) ---
    placement_data = brief_data.get('placement_data') # placement_data can be list or DataFrame
    bvp_date_map = {}
    processed_bvps = set() # Track processed BVPs to avoid overwriting

    # Check if placement_data is a non-empty DataFrame
    if isinstance(placement_data, pd.DataFrame) and not placement_data.empty:
        print("Processing Placement Data as DataFrame...")
        # --- Debug: Print available columns ---
        print(f"Placement DataFrame Columns: {list(placement_data.columns)}") 
        # --- End Debug ---
        
        # Find column names precisely (case-insensitive search)
        bvp_col_name = next((col for col in placement_data.columns if col.strip().upper() == 'BVP'), None)
        proj_start_col_name = next((col for col in placement_data.columns if col.strip().upper() == 'PROJECTED START DATE'), None)
        end_col_name = next((col for col in placement_data.columns if col.strip().upper() == 'END DATE'), None)
        
        print(f"Found Columns - BVP: '{bvp_col_name}', Projected Start: '{proj_start_col_name}', End: '{end_col_name}'") # Debug print

        if bvp_col_name:
            # Get column indices *before* the loop
            try:
                bvp_col_idx = list(placement_data.columns).index(bvp_col_name)
                proj_start_col_idx = list(placement_data.columns).index(proj_start_col_name) if proj_start_col_name else -1
                end_col_idx = list(placement_data.columns).index(end_col_name) if end_col_name else -1
            except ValueError as e:
                print(f"Error finding index for columns: {e}")
                return # Cannot proceed without column indices

            # Use itertuples(index=False) - tuple indices match column indices
            for row_tuple in placement_data.itertuples(index=False):
                try:
                    bvp = str(row_tuple[bvp_col_idx]).strip()
                    if bvp and bvp not in processed_bvps:
                        start_date_raw = None
                        end_date_raw = None
                        
                        # Access by index
                        if proj_start_col_idx != -1:
                            start_date_raw = row_tuple[proj_start_col_idx]
                            # print(f"Raw Start Date Value (Index {proj_start_col_idx} - '{proj_start_col_name}'): {repr(start_date_raw)}") # Debug removed
                        # else:
                            # print(f"Projected Start Date column '{proj_start_col_name}' not found.") # Debug removed
                            
                        if end_col_idx != -1:
                             end_date_raw = row_tuple[end_col_idx]
                             # print(f"Raw End Date Value (Index {end_col_idx} - '{end_col_name}'): {repr(end_date_raw)}") # Debug removed
                        # else:
                             # print(f"End Date column '{end_col_name}' not found.") # Debug removed

                        start_date = safe_date_convert(start_date_raw)
                        end_date = safe_date_convert(end_date_raw)
                        
                        # print(f"Converted Start Date: {start_date}") # Debug removed
                        # print(f"Converted End Date: {end_date}") # Debug removed
                        
                        if start_date or end_date:
                            bvp_date_map[bvp] = (start_date, end_date)
                            processed_bvps.add(bvp)
                # Handle potential IndexError if indices are wrong, though finding them first should prevent this
                except IndexError as e: 
                    print(f"Warning: Error accessing data by index in placement_data row: {e}")
                    continue 
                except Exception as e: # Catch other potential errors in the loop
                    print(f"Warning: An unexpected error occurred processing row for BVP '{bvp}': {e}")
                    continue
        else:
             print(f"Warning: Placement data DataFrame missing essential 'BVP' column.")

    # Check if placement_data is a non-empty list
    elif isinstance(placement_data, list) and placement_data:
        print("Processing Placement Data as List...")
        # --- Debug: Print keys of the first item ---
        # if isinstance(placement_data[0], dict):
            # print(f"Placement List Item Keys (first item): {list(placement_data[0].keys())}") # Debug removed
        # --- End Debug ---
        
        for placement_item in placement_data:
            if isinstance(placement_item, dict):
                # Use exact keys as requested
                bvp_key = 'BVP'
                proj_start_key = 'Projected Start Date'
                end_key = 'End Date'
                
                if bvp_key in placement_item:
                    bvp = str(placement_item.get(bvp_key, '')).strip()
                    if bvp and bvp not in processed_bvps:
                        
                        start_date_raw = placement_item.get(proj_start_key) # Get value for 'Projected Start Date'
                        end_date_raw = placement_item.get(end_key)         # Get value for 'End Date'
                        
                        start_date = safe_date_convert(start_date_raw)
                        end_date = safe_date_convert(end_date_raw)
                        
                        if start_date or end_date:
                            bvp_date_map[bvp] = (start_date, end_date)
                            processed_bvps.add(bvp)
    else:
        print("Warning: Could not extract valid Placement Data (list or DataFrame) from brief.")
        
    print(f"Found {len(bvp_date_map)} BVP date mappings.")
        
    # --- Map Line Items (BVT) to BVP and then to Dates ---
    sf_li_starts = []
    sf_li_ends = []
    matched_bvps = []
    
    for index, row in qa_df.iterrows():
        bvt = str(row['line_item_alternative_id']).strip() if pd.notna(row['line_item_alternative_id']) else ""
        bvp = bvt_bvp_map.get(bvt)
        
        current_start = None
        current_end = None
        
        if bvp:
            matched_bvps.append(bvp)
            dates = bvp_date_map.get(bvp)
            if dates:
                current_start = dates[0]
                current_end = dates[1]
            else:
                # This warning is still useful
                print(f"Warning: No dates found in Placement Data for BVP '{bvp}' (mapped from BVT '{bvt}')")
        else:
            matched_bvps.append(None)
            if bvt: # Only warn if there was a BVT ID to look up
                 # This warning is still useful
                 print(f"Warning: Could not find BVP mapping in Target Data for BVT '{bvt}'")

        sf_li_starts.append(current_start)
        sf_li_ends.append(current_end)
            
    qa_df['matched_bvp'] = matched_bvps
    qa_df['sf_li_start_date'] = sf_li_starts
    qa_df['sf_li_end_date'] = sf_li_ends
    
    # --- Perform Date Comparisons ---
    # Convert QA report dates safely first
    qa_df['campaign_start_date_dt'] = qa_df['campaign_start_date'].apply(safe_date_convert)
    qa_df['campaign_end_date_dt'] = qa_df['campaign_end_date'].apply(safe_date_convert)
    qa_df['line_item_start_date_dt'] = qa_df['line_item_start_date'].apply(safe_date_convert)
    qa_df['line_item_end_date_dt'] = qa_df['line_item_end_date'].apply(safe_date_convert)
    
    # Ensure SF dates are also proper datetimes (already done by safe_date_convert)
    qa_df['sf_campaign_start_date_dt'] = qa_df['sf_campaign_start_date'] # Already datetime or None
    qa_df['sf_campaign_end_date_dt'] = qa_df['sf_campaign_end_date']     # Already datetime or None
    qa_df['sf_li_start_date_dt'] = qa_df['sf_li_start_date']           # Already datetime or None
    qa_df['sf_li_end_date_dt'] = qa_df['sf_li_end_date']             # Already datetime or None

    # Compare campaign dates (only the date part)
    qa_df['c_start_date_match'] = qa_df.apply(
        lambda row: (row['campaign_start_date_dt'].date() == row['sf_campaign_start_date_dt'].date()) 
                    if pd.notna(row['campaign_start_date_dt']) and pd.notna(row['sf_campaign_start_date_dt']) else False,
        axis=1
    )
    qa_df['c_end_date_match'] = qa_df.apply(
        lambda row: (row['campaign_end_date_dt'].date() == row['sf_campaign_end_date_dt'].date()) 
                    if pd.notna(row['campaign_end_date_dt']) and pd.notna(row['sf_campaign_end_date_dt']) else False,
        axis=1
    )
    
    # Compare line item dates (only the date part)
    qa_df['li_start_date_match'] = qa_df.apply(
        lambda row: (row['line_item_start_date_dt'].date() == row['sf_li_start_date_dt'].date()) 
                    if pd.notna(row['line_item_start_date_dt']) and pd.notna(row['sf_li_start_date_dt']) else False,
        axis=1
    )
    qa_df['li_end_date_match'] = qa_df.apply(
        lambda row: (row['line_item_end_date_dt'].date() == row['sf_li_end_date_dt'].date()) 
                    if pd.notna(row['line_item_end_date_dt']) and pd.notna(row['sf_li_end_date_dt']) else False,
        axis=1
    )
    
    # Add overall summary column
    qa_df['all_dates_match'] = (
        qa_df['c_start_date_match'] & 
        qa_df['c_end_date_match'] & 
        qa_df['li_start_date_match'] & 
        qa_df['li_end_date_match']
    )
    
    # --- Prepare Final Output ---
    # Define final column order
    cols_order = [
        'campaign_id', 'campaign_name', 
        'campaign_start_date', 'sf_campaign_start_date', 'c_start_date_match',
        'campaign_end_date', 'sf_campaign_end_date', 'c_end_date_match',
        'line_item_id', 'line_item_name', 'line_item_alternative_id', 'matched_bvp',
        'line_item_start_date', 'sf_li_start_date', 'li_start_date_match',
        'line_item_end_date', 'sf_li_end_date', 'li_end_date_match', 
        'all_dates_match'
    ]
    
    # Select and reorder columns for the DataFrame
    # Convert datetime columns to string in desired format for Excel to avoid timezone issues
    output_df = qa_df.copy()
    date_cols_to_format = [
        'campaign_start_date', 'sf_campaign_start_date', 
        'campaign_end_date', 'sf_campaign_end_date',
        'line_item_start_date', 'sf_li_start_date', 
        'line_item_end_date', 'sf_li_end_date'
    ]
    for col in date_cols_to_format:
         # Use the original columns if dt conversion failed, otherwise use dt columns
         dt_col = col + '_dt' if col + '_dt' in output_df.columns else col
         # Format as YYYY-MM-DD
         output_df[col] = pd.to_datetime(output_df[dt_col]).dt.strftime('%Y-%m-%d').replace('NaT', '') 

    # Ensure boolean columns are actual booleans for formatting logic later
    bool_cols = ['c_start_date_match', 'c_end_date_match', 'li_start_date_match', 'li_end_date_match', 'all_dates_match']
    for col in bool_cols:
        output_df[col] = output_df[col].astype(bool)

    # Select final columns
    output_df = output_df[cols_order]

    # Deduplicate based on line_item_id to remove duplicate entries from different creatives
    print(f"Original row count before deduplication: {len(output_df)}")
    output_df_deduplicated = output_df.drop_duplicates(subset=['line_item_id'], keep='first')
    print(f"Row count after deduplication: {len(output_df_deduplicated)}")

    # Use the deduplicated DataFrame for output
    output_df = output_df_deduplicated

    # --- Save Output with Formatting ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Flight Check Results"
        
        # Write header row
        ws.append(cols_order)
        header_font = Font(bold=True, color="FF000000") # Black text
        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid") # Light Grey
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting to header (Row 1)
        for col_idx, col_name in enumerate(cols_order, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        ws.row_dimensions[1].height = 30 # Set height for header
        
        # Define fills for data cells
        true_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid") # Light Green
        false_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid") # Light Red
        
        # Write data rows with formatting
        for r_idx, row in enumerate(output_df.itertuples(index=False), 2): # Start writing from row 2
            ws.append(row)
            # Apply formatting to boolean columns for this row
            for col_idx, col_name in enumerate(cols_order, 1):
                if col_name in bool_cols:
                    cell = ws.cell(row=r_idx, column=col_idx)
                    # Use the boolean value from the tuple `row`
                    # The index in the tuple corresponds to the index in cols_order
                    bool_value = getattr(row, col_name) 
                    if bool_value: # Check if True
                        cell.fill = true_fill
                    else:
                         cell.fill = false_fill

        # Adjust column widths automatically (based on header and sample data)
        for col_idx, column_title in enumerate(cols_order, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            # Check header length
            max_length = max(len(str(column_title)), max_length)
            # Check data length in the column (sample first 100 rows + header for performance)
            for i in range(1, min(ws.max_row + 1, 102)): 
                cell_value = ws[f"{column_letter}{i}"].value
                if cell_value:
                    # Consider max length for boolean "FALSE"
                    if isinstance(cell_value, bool):
                        max_length = max(max_length, 5) 
                    else:
                        max_length = max(max_length, len(str(cell_value)))
            
            # Add padding, set min/max width
            adjusted_width = min(max((max_length + 2), 10), 50) # Min width 10, Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print(f"\nFlight check complete. Formatted output saved to {output_path}")
        
    except Exception as e:
        print(f"Error writing formatted Excel output: {e}")
        print("Attempting to save raw data without formatting...")
        try:
            # Ensure boolean columns are strings for raw export if openpyxl fails
            output_df_raw = output_df.copy()
            for col in bool_cols:
                 output_df_raw[col] = output_df_raw[col].astype(str) # Convert bool back to string
                 
            output_df_raw.to_excel(output_path, index=False)
            print(f"Raw data saved successfully to {output_path}")
        except Exception as e2:
            print(f"Failed to save raw data: {e2}")
    
    # Print summary
    total_line_items = len(output_df)
    if total_line_items > 0:
        matched_dates = output_df['all_dates_match'].sum()
        match_percentage = (matched_dates / total_line_items) * 100
        print(f"Summary: {matched_dates} out of {total_line_items} line items have all dates matching ({match_percentage:.1f}%)")
    else:
        print("Summary: No line items found in the QA report to process.")

if __name__ == "__main__":
    main() 