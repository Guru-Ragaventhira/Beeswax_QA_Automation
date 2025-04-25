"""
Name Assignment and Verification Tool

This script analyzes campaign briefs and QA reports to verify naming conventions
for campaigns, line items, and creatives.

Features:
- Robust campaign brief parsing with multiple fallback mechanisms
- Intelligent column detection for varying brief formats
- Flexible date and viewability extraction
- Comprehensive naming convention checks
- Clear output with formatting and cell comments

Usage:
1. Place your Campaign Brief Excel file in the Brief directory
2. Run this script
3. Review results in the output file (defined in .env)

Environment variables:
- BRIEF_PATH: Path to Campaign Brief Excel file
- QA_REPORT_PATH: Path to the QA report Excel file
- NAME_ASSIGN_OUTPUT_PATH: Path for the output file
"""

import pandas as pd
import numpy as np
import re
import os
import glob
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import the function from brief_extractor
from brief_extractor import extract_structured_brief_data

def find_latest_qa_report(output_dir):
    """Find the latest QA report file in the output directory"""
    qa_report_files = glob.glob(os.path.join(output_dir, "qa_report_*.xlsx"))
    if not qa_report_files:
        return None
    
    # Sort by modification time, newest first
    latest_file = max(qa_report_files, key=os.path.getmtime)
    return latest_file

def safe_date_convert(date_val):
    """Safely convert various date formats to pandas datetime"""
    if pd.isna(date_val) or date_val == '':
        return None
    
    if isinstance(date_val, (datetime, pd.Timestamp)):
        return pd.to_datetime(date_val) # Ensure it's pandas Timestamp
    
    # Try parsing date string with various formats
    try:
        # First try pandas to_datetime with default parser
        return pd.to_datetime(date_val)
    except Exception as e1:
        # Handle Excel date as float (days since 1900-01-01)
        try:
            if isinstance(date_val, (int, float)):
                # Check if it's a reasonable Excel date number
                if 30000 < date_val < 70000: 
                    return pd.to_datetime('1899-12-30') + pd.Timedelta(days=float(date_val))
        except Exception as e2:
            print(f"Excel date conversion error for {date_val}: {e2}")
            pass
            
        # Try common date formats explicitly
        date_formats = [
            '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y', '%m-%d-%Y',
            '%Y/%m/%d', '%b %d, %Y', '%d %b %Y',
            '%m/%d/%y', '%d/%m/%y', '%y-%m-%d',
            '%m.%d.%Y', '%d.%m.%Y', '%Y.%m.%d'
        ]
        
        for fmt in date_formats:
            try:
                # Ensure input is string before parsing
                return pd.to_datetime(str(date_val).strip(), format=fmt)
            except (ValueError, TypeError):
                continue
                
    # If all attempts fail, return None
    print(f"Warning: Could not convert '{date_val}' to datetime after trying multiple formats. Error: {e1}")
    return None

def get_field_value(df, field_pattern, default=None):
    """
    Extract value for a specific field pattern from a DataFrame with 'Field'/'Value' columns.
    Case-insensitive search. Returns the first match or default.
    """
    if df is None or df.empty or 'Field' not in df.columns or 'Value' not in df.columns:
        # print(f"Warning: Cannot get field value for '{field_pattern}'. DataFrame is invalid or empty.") # Reduced verbosity
        return default
        
    try:
        # Ensure Field column is string type for reliable comparison
        df['Field'] = df['Field'].astype(str) 
        
        # Find rows matching the field pattern (case-insensitive)
        # Use regex=True for more flexible pattern matching if needed, but False is safer for exact strings
        matching_rows = df[df['Field'].str.contains(field_pattern, case=False, na=False, regex=False)]
        
        if not matching_rows.empty:
            # Return the value from the first matching row
            value = matching_rows.iloc[0]['Value']
            # Convert potential numpy types to standard Python types
            if isinstance(value, np.generic):
                 value = value.item()
            # print(f"Found value for '{field_pattern}': {value}") # Reduced verbosity
            return value if pd.notna(value) else default # Return default if value is NaN/None
        else:
            print(f"Field containing pattern '{field_pattern}' not found. Available fields: {list(df['Field'])}")
            return default
            
    except KeyError:
        # This shouldn't happen due to the check at the start, but good practice
        print(f"Error: 'Field' or 'Value' column not found in DataFrame while searching for '{field_pattern}'.")
        return default
    except Exception as e:
        print(f"Error getting field value for '{field_pattern}': {e}")
        return default

def extract_product_type_shortform(product_type):
    """Get short form for product type"""
    if pd.isna(product_type) or not isinstance(product_type, str):
        return None

    product_type = product_type.strip().lower()
    
    short_forms = {
        "all outlet rewards": "AOR",
        "ad2ecomm": "A2E",
        "ad2survey": "A2S",
        "connected tv": "CTV",
        "sequential": "SQ",
        "volume maximizer": "VMR",
        "standard bv": "SBV",
        "post campaign measurement": "PCM",
        "circular personalizer": "CircP",
        "price promoter": "CircP_PP", # Assuming nested under CircP
        "trip driver": "CircP_TD"     # Assuming nested under CircP
    }
    
    # Check for CircP variations first
    if "price promoter" in product_type:
        return "CircP_PP"
    if "trip driver" in product_type:
        return "CircP_TD"
        
    # Check for exact matches in keys or significant parts
    for key, form in short_forms.items():
        if key in product_type:
            return form
            
    # Handle potential variations like "BV - Standard" -> SBV
    if "bv" in product_type and "standard" in product_type:
        return "SBV"

    print(f"Warning: No short form found for product type: {product_type}")
    return None

def extract_viewability_percentage(viewability_text):
    """Extract percentage value from viewability text"""
    if pd.isna(viewability_text) or not isinstance(viewability_text, str):
        return None
    
    input_str = str(viewability_text).strip()
    if not input_str:
        return None

    # --- Priority 1: Check for decimal representation (0.0 to 1.0) --- 
    try:
        float_val = float(input_str)
        # Check if it's a percentage expressed as a decimal (e.g., 0.7 for 70%)
        # Allow range (0, 1] - 0% doesn't usually need naming, 1 means 100%
        if 0 < float_val <= 1:
            percentage = int(float_val * 100)
            print(f"Extracted viewability percentage from decimal: {percentage}")
            return percentage
        # Handle if user entered 70 instead of 0.7 or 70%
        elif 1 < float_val <= 100:
            # Check if it looks like a whole percentage number (e.g., 70.0)
            if float_val == int(float_val):
                 percentage = int(float_val)
                 print(f"Extracted viewability percentage from whole number float: {percentage}")
                 return percentage

    except (ValueError, TypeError):
        # Not a simple float, proceed to regex checks
        pass 
    # --- End Priority 1 ---

    # Regex to find numbers followed by '%'
    match = re.search(r'(\d{1,3})\s*%', viewability_text)
    if match:
        # Fall through to try the fallback method
        try:
            percentage = int(match.group(1))
            # Basic sanity check
            if 0 <= percentage <= 100:
                 print(f"Extracted viewability percentage: {percentage}")
                 return percentage
            else:
                 print(f"Warning: Extracted percentage {percentage} is out of range (0-100).")
                 # Consider returning None if out of range, based on requirements
                 # return None 
        except ValueError:
            print(f"Warning: Could not convert extracted percentage '{match.group(1)}' to int.")
            # Fall through to try the fallback method

    # Fallback: Look for just a number if '%' is missing, assuming it's percentage
    # Use the cleaned input_str here
    match_fallback = re.search(r'\b(\d{1,3})\b', input_str)
    if match_fallback:
        try:
            percentage = int(match_fallback.group(1))
            if 0 <= percentage <= 100:
                print(f"Extracted viewability percentage (fallback): {percentage}")
                return percentage
            else:
                print(f"Warning: Extracted fallback percentage {percentage} is out of range (0-100).")
                # Consider returning None if out of range
                # return None
        except ValueError:
            print(f"Warning: Could not convert extracted fallback percentage '{match_fallback.group(1)}' to int.")
            # Continue to final warning if fallback also fails

    print(f"Warning: Could not extract percentage from viewability text: {viewability_text}")
    return None

def extract_platform_media_type(platform_media_text):
    """Extract platform and media type"""
    if pd.isna(platform_media_text) or not isinstance(platform_media_text, str):
        return None, None
    
    text = platform_media_text.strip()
    
    # Common pattern: Platform/MediaType
    if '/' in text:
        parts = text.split('/', 1)
        platform = parts[0].strip().lower()
        media_type = parts[1].strip().lower()
        print(f"Extracted platform='{platform}', media_type='{media_type}' from '{text}'")
        return platform, media_type
        
    # Handle cases where only one is present or format is different
    text_lower = text.lower()
    platform = None
    media_type = None
    
    # Simple keyword checks
    if 'mobile' in text_lower: platform = 'mobile'
    elif 'desktop' in text_lower: platform = 'desktop'
    elif 'ctv' in text_lower or 'connected tv' in text_lower: platform = 'ctv'
    
    if 'banner' in text_lower: media_type = 'banner'
    elif 'rich media' in text_lower: media_type = 'rich media'
    elif 'video' in text_lower: media_type = 'video'
    
    # If only one was identified, return it
    if platform and not media_type:
        print(f"Extracted platform='{platform}', media_type=None from '{text}'")
        return platform, None
    if not platform and media_type:
        print(f"Extracted platform=None, media_type='{media_type}' from '{text}'")
        return None, media_type
    # If both were found via keywords (unlikely but possible)
    if platform and media_type:
         print(f"Extracted platform='{platform}', media_type='{media_type}' (keyword search) from '{text}'")
         return platform, media_type

    print(f"Warning: Could not reliably extract platform/media type from: {platform_media_text}")
    return None, None

def get_platform_prefix(platform):
    """Get naming prefix based on platform"""
    if pd.isna(platform): return None
    platform = platform.lower()
    if 'mobile' in platform: return ('MOA_', 'MOW_', 'MO_') # Return tuple of possibilities
    if 'desktop' in platform: return ('DE_',)
    if 'ctv' in platform: return ('CTV_',)
    return None

def get_media_type_code(media_type):
    """Get naming code based on media type"""
    if pd.isna(media_type): return None
    media_type = media_type.lower()
    if 'banner' in media_type: return '_BA_'
    if 'rich media' in media_type: return '_RM_'
    if 'video' in media_type: return '_VI_'
    return None

def check_naming_format(name, checks):
    """
    Check a name against a list of format requirements.
    Returns a dictionary of boolean check results (True=Issue Found).
    """
    results = {
        'has_issues': False, # Overall flag
        'has_spaces': False,
        'has_special_chars': False,
        'missing_quarter': False,
        'missing_year': False,
        'missing_product_type': False, # Campaign specific
        'missing_hub_ifo_tag': set(),   # Campaign specific (Combined, stores missing tags like INFMT, IFO)
        'missing_lda': False,          # Campaign specific (NEW)
        'missing_viewability': False,  # Line Item / Campaign specific
        'geo_mismatch': False,         # Line Item & Creative specific
        'platform_mismatch': False,    # Line Item & Creative specific
        'media_type_mismatch': False   # Line Item & Creative specific
    }
    error_messages = [] # Keep track of specific errors for comments if needed

    if pd.isna(name) or not isinstance(name, str) or not name.strip():
        error_messages.append("Name is missing or empty.")
        results['has_issues'] = True
        # Return early if name is invalid
        return results, error_messages # Return tuple: (results_dict, error_list)
        
    name_original = name # Keep original case if needed later
    name = name.strip()
    name_upper = name.upper() # Use uppercase for most checks

    # Common checks for all names
    # 1. No spaces
    if ' ' in name:
        results['has_spaces'] = True
        error_messages.append("Name contains spaces.")
        
    # 2. No special characters except '_'
    # Use the strict pattern for all entity types
    allowed_pattern_re = r'[a-zA-Z0-9_ ]+'
    invalid_chars_pattern_re = r'[^a-zA-Z0-9_ ]'

    # --- Debug: Print details before the fullmatch check ---
    match_result = re.fullmatch(allowed_pattern_re, name)
    # print(f"Debug CheckFormat - Name: {repr(name)}, Pattern: '{allowed_pattern_re}', Fullmatch Result: {bool(match_result)}") # Keep commented unless debugging
    # --- End Debug ---

    if not re.fullmatch(allowed_pattern_re, name):
        results['has_special_chars'] = True
        # Find characters *not* matching the allowed pattern for this type
        invalid_chars = set(re.findall(invalid_chars_pattern_re, name))
        error_messages.append(f"Name contains invalid characters: {', '.join(invalid_chars)}.")
         
    # 3. Quarter and Year check (_Q[1-4]_YYYY or _Q[1-4]_YY) - Simplified independent checks
    year_pattern_str = checks.get('year_pattern') # e.g., '_2024|_24'
    quarter_required = checks.get('quarter_required', True) 
    year_required = bool(year_pattern_str) # Year is required if pattern exists

    # Check for Quarter (_Q[1-4]_)
    if quarter_required:
       if not re.search(r'_Q[1-4]_', name_upper):
            results['missing_quarter'] = True
            error_messages.append("Name missing Quarter format (e.g., _Q1_).")

    # Check for Year (_YYYY or _YY)
    # Ensure the year pattern requires an underscore before it
    if year_required:
         # Example year_pattern_str: _2024|_24
         # We need to match _Qx_YYYY or _Qx_YY, OR just _YYYY or _YY if Q isn't required/present
         # Let's check for the direct year pattern first
         if not re.search(year_pattern_str.upper(), name_upper):
             results['missing_year'] = True
             year_display = year_pattern_str.replace('|',' or ')
             error_messages.append(f"Name missing Year format (e.g., {year_display}).")

    # Get check type for specific rules
    check_type = checks.get('type')

    # Campaign specific checks
    if check_type == 'campaign':
        # a. Product Type Short Form
        product_short_forms = checks.get('product_short_forms')
        if product_short_forms: # Might be a list
             found_product = False
             for form in product_short_forms:
                 # Check with underscores around the form, case insensitive
                 if f'_{form.upper()}_' in name_upper:
                     found_product = True
                     break
             if not found_product:
                  results['missing_product_type'] = True
                  error_messages.append(f"Name missing Product Type code (e.g., _{ ' or _'.join(product_short_forms)}_).")
        
        # b. Combined HUB/IFO Tag Check
        is_hub = checks.get('is_hub')
        is_ifo = checks.get('is_ifo')
        if is_hub and '_INFMT_' not in name_upper:
            results['missing_hub_ifo_tag'].add('INFMT')
            error_messages.append("Name missing HUB indicator ('_INFMT_') when required.")
        if is_ifo and '_IFO_' not in name_upper:
            results['missing_hub_ifo_tag'].add('IFO')
            error_messages.append("Name missing IFO indicator ('_IFO_') when required.")
            
        # c. LDA Tag Check
        is_lda_required = checks.get('is_lda_required')
        if is_lda_required and '_LDA_' not in name_upper:
            results['missing_lda'] = True
            error_messages.append("Name missing LDA indicator ('_LDA_') when required by brief.")

        # d. Viewability Percentage (Campaign Level)
        viewability_perc = checks.get('viewability_perc')
        if viewability_perc is not None and viewability_perc != 0:
            pct_str = str(viewability_perc)
            # Allow _XX_Viewability_, _XXViewability_, or just _XX_ (case insensitive)
            view_pattern1 = rf'_{pct_str}_VIEWABILITY_'
            view_pattern2 = rf'_{pct_str}VIEWABILITY_'
            view_pattern3 = rf'_{pct_str}_'
            if not (re.search(view_pattern1, name_upper, re.IGNORECASE) or
                    re.search(view_pattern2, name_upper, re.IGNORECASE) or
                    re.search(view_pattern3, name_upper, re.IGNORECASE)):
                results['missing_viewability'] = True
                error_messages.append(f"Campaign Name missing Viewability ({viewability_perc}%) indicator.")

    # Line Item specific checks
    elif check_type == 'line_item':
        # a. Viewability Percentage (Line Item Level)
        viewability_perc = checks.get('viewability_perc')
        if viewability_perc is not None and viewability_perc != 0:
            pct_str = str(viewability_perc)
            # Allow _XX_Viewability_, _XXViewability_, or just _XX_ (case insensitive)
            view_pattern1 = rf'_{pct_str}_VIEWABILITY_'
            view_pattern2 = rf'_{pct_str}VIEWABILITY_'
            view_pattern3 = rf'_{pct_str}_'
            if not (re.search(view_pattern1, name_upper, re.IGNORECASE) or
                    re.search(view_pattern2, name_upper, re.IGNORECASE) or
                    re.search(view_pattern3, name_upper, re.IGNORECASE)):
                results['missing_viewability'] = True
                error_messages.append(f"Line Item Name missing Viewability ({viewability_perc}%) indicator.")
        
        # b. Geo Targeting (_Geo_ or _GEO_)
        is_geo_required = checks.get('is_geo_required')
        has_geo_in_name = '_GEO_' in name_upper # Check uppercase
        if is_geo_required is True and not has_geo_in_name:
             results['geo_mismatch'] = True
             error_messages.append("Name missing Geo indicator ('_Geo_') but brief requires it.")
        elif is_geo_required is False and has_geo_in_name:
             results['geo_mismatch'] = True
             error_messages.append("Name includes Geo indicator ('_Geo_') but brief does not require it.")
             
        # c. Platform Prefix (Starts with MOA_, MOW_, MO_, DE_, CTV_)
        platform_prefixes = checks.get('platform_prefixes') # Tuple of valid prefixes
        if platform_prefixes:
             # Check if name starts with any of the prefixes (case insensitive)
             if not any(name_upper.startswith(pfx.upper()) for pfx in platform_prefixes):
                 results['platform_mismatch'] = True
                 error_messages.append(f"Name does not start with expected Platform prefix ({' or '.join(platform_prefixes)}).")
        elif checks.get('platform') is not None: 
             pass 

        # d. Media Type Code (_BA_, _RM_, _VI_)
        media_type_code = checks.get('media_type_code')
        if media_type_code:
            # Check if code exists anywhere in the name (case insensitive)
            if media_type_code.upper() not in name_upper:
                results['media_type_mismatch'] = True
                error_messages.append(f"Name missing expected Media Type code ('{media_type_code}').")
        elif checks.get('media_type') is not None: 
             pass

    # Creative specific checks
    elif check_type == 'creative':
        # Creative checks often depend on the associated Line Item's structure
        
        # a. Geo Targeting (matches Line Item's expected Geo status)
        li_has_geo = checks.get('li_has_geo', False) # From the associated LI check
        creative_has_geo = '_GEO_' in name_upper
        if li_has_geo and not creative_has_geo:
            results['geo_mismatch'] = True
            error_messages.append("Creative name missing Geo indicator ('_Geo_') expected from Line Item.")
        elif not li_has_geo and creative_has_geo:
            results['geo_mismatch'] = True
            error_messages.append("Creative name has Geo indicator ('_Geo_') but Line Item does not.")
            
        # b. Platform Prefix (matches Line Item's prefix, with special mobile non-HUB case)
        li_platform_prefix = checks.get('li_platform_prefix')
        li_platform = checks.get('li_platform') # e.g., 'mobile'
        measurement_type = checks.get('measurement_type_str_brief')

        # Special rule applies if platform is mobile AND measurement type is NOT a string containing 'HUB'
        is_non_hub_mobile = False
        if li_platform == 'mobile' and (not isinstance(measurement_type, str) or 'HUB' not in measurement_type.upper()):
             is_non_hub_mobile = True
             print(f"Debug: Applying Non-HUB Mobile rule for Creative Check (LI Platform: {li_platform}, Measurement: '{measurement_type}')") # Debug Print
        # else: # Optional: Add else for clarity on when standard rule applies
        #     print(f"Debug: Applying Standard Mobile rule for Creative Check (LI Platform: {li_platform}, Measurement: '{measurement_type}')")

        # Apply check
        if li_platform_prefix:
            creative_starts_with_li_prefix = name_upper.startswith(li_platform_prefix.upper())
            creative_starts_with_mo = name_upper.startswith('MO_') # Check for generic MO_

            if is_non_hub_mobile and li_platform_prefix in ['MOA_', 'MOW_']:
                # Special Case: Non-HUB Mobile, LI is MOA_ or MOW_
                # Creative must start with LI prefix OR MO_
                if not (creative_starts_with_li_prefix or creative_starts_with_mo):
                    results['platform_mismatch'] = True
                    error_messages.append(f"Creative name does not start with expected LI Platform prefix ('{li_platform_prefix}') or generic 'MO_' for non-HUB mobile.")
            else:
                # Standard Case: Must start with exact LI prefix
                if not creative_starts_with_li_prefix:
                  results['platform_mismatch'] = True
                  error_messages.append(f"Creative name does not start with expected Line Item Platform prefix ('{li_platform_prefix}').")
        
        # c. Media Type Code (matches Line Item's code)
        li_media_type_code = checks.get('li_media_type_code')
        if li_media_type_code:
            if li_media_type_code.upper() not in name_upper:
                 results['media_type_mismatch'] = True
                 error_messages.append(f"Creative name missing expected Line Item Media Type code ('{li_media_type_code}').")

    # Update overall 'has_issues' flag
    results['has_issues'] = any(
        v for k, v in results.items() 
        if k != 'has_issues' and k != 'missing_hub_ifo_tag' and v is True
    ) or bool(results['missing_hub_ifo_tag']) # Check if the tag set is non-empty

    return results, error_messages


def add_comment_to_cell(worksheet, cell_coord, comment_text):
    """Adds a comment to a specific cell"""
    if comment_text:
        # Ensure comment text is not excessively long
        max_len = 32767 # Excel limit, but keep it reasonable
        if len(comment_text) > max_len:
            comment_text = comment_text[:max_len-3] + "..."
            
        # Remove invalid characters if any (rare but possible)
        comment_text = re.sub(r'[\\x00-\\x08\\x0B\\x0C\\x0E-\\x1F]', '', comment_text) 
            
        try:
            from openpyxl.comments import Comment # Import here to avoid top-level if unused
            # Use cell coordinate directly
            cell = worksheet[cell_coord]
            # Remove existing comment first to avoid appending issues with formatting
            if cell.comment:
                 cell.comment = None 
                 
            # Create and assign the new comment
            # Use the original comment text directly, let Excel handle wrapping
            comment_obj = Comment(comment_text, "NameCheck Bot") 
            
            # Set comment box size
            comment_obj.width = 300  # Keep reduced width
            comment_obj.height = 150 # Keep reduced height
            
            # Assign the comment
            cell.comment = comment_obj
            
            # Removed: worksheet._comments.append(cell.comment) - not standard/reliable

        except Exception as e:
            print(f"Error adding comment to {cell_coord}: {e}")
            # Fallback: Print comment to console if adding fails
            print(f"Comment for {cell_coord}: {comment_text}")

# Define column descriptions for header comments
# Use simple strings, let Excel wrap the text in the comment box
column_descriptions = {
    'has_issues': 'Overall check result. TRUE if any issue found.',
    'has_spaces': 'Checks if name contains spaces. Rule: No spaces allowed.',
    'has_special_chars': 'Checks for special characters except underscores. Rule: Only alphanumeric and underscores allowed.',
    'missing_quarter': 'Checks if name includes a quarter format (_Q1_, _Q2_, etc.). Rule: Must include quarter.',
    'missing_year': 'Checks if name includes year format (_YYYY or _YY). Rule: Must include campaign year.',
    'missing_product_type': '[Campaign Only] Checks if name includes product type shortform (_SBV_). Rule: Must include shortform.',
    'missing_hub_ifo_tag': '[Campaign Only] Checks if name includes _INFMT_ (if HUB) or _IFO_ (if IFO). Rule: Must include required tag.',
    'missing_lda': '[Campaign Only] Checks if name includes _LDA_ (if LDA). Rule: Must include if required by brief.',
    'missing_viewability': '[Campaign/LI] Checks if name includes viewability % (_70_, _70_VIEWABILITY_). Rule: Must include if specified in brief.',
    'geo_mismatch': '[LI/Creative] Checks if name includes _GEO_ tag appropriately based on brief/LI. Rule: _GEO_ required if brief says Yes, prohibited if No.',
    'platform_mismatch': '[LI/Creative] Checks if name starts with platform prefix (MOA_, DE_) or matches LI prefix. Rule: Must match brief/LI.',
    'media_type_mismatch': '[LI/Creative] Checks if name includes media type code (_BA_, _RM_) or matches LI code. Rule: Must match brief/LI.',
    'check_active_status': 'Checks if Campaign (C), LI (Li), Creative Active (Cr_A), Creative Secure (Cr_S) are correct. Rules: C=TRUE, Li=FALSE, Cr_A=TRUE, Cr_S=1.',
    'hub_creative_sharing': '[HUB Campaign Only] Checks if a Creative ID is used under multiple Line Item IDs. Rule: Each Creative must belong to only one LI.',
    # Brief Context Columns Descriptions
    'brief_product_type': 'Product Type string from Brief (Account Level).',
    'brief_measurement_type': 'Measurement Type string from Brief (Campaign Level).',
    'brief_viewability': 'Viewability Goal string from Brief (Campaign Level).',
    'brief_lda_compliant': 'LDA or Age Compliant string from Brief (Campaign Level).',
    'brief_bvt_id': 'BVT ID from QA Report (used for merging Target Data).',
    'brief_bvp_id': 'BVP ID derived from Brief (Target Level via BVT).',
    'brief_geo_required': 'Geo Required value from Brief (Placement Level via BVP).',
    'brief_platform_media': 'Platform/Media Type from Brief (Target Level via BVT).',
}

def main():
    print("Starting Name Assignment Check...")
    
    # Load environment variables
    env_path = os.getenv("ENV_PATH", "./input_folder/beeswax_input_qa.env")
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        load_dotenv(env_path)
    else:
        print(f"Warning: Environment file {env_path} not found. Using default environment.")
        load_dotenv()  # Try default locations
    
    # Get paths from environment or use defaults
    brief_path = os.path.abspath(os.getenv("BRIEF_PATH", "./Brief/Campaign_Brief.xlsx"))
    output_dir = os.path.abspath(os.getenv("OUTPUT_DIR", "./output_folder"))
    name_assign_output_path = os.path.abspath(os.getenv("NAME_ASSIGN_OUTPUT_PATH", "./name_assign_output_v2.xlsx"))
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Determine QA report path - first check env variable, then find latest
    qa_report_path = os.getenv("QA_REPORT_PATH")
    if not qa_report_path or not os.path.exists(qa_report_path):
        print(f"QA report not found at {qa_report_path}, looking for the latest report...")
        latest_report = find_latest_qa_report(output_dir)
        if latest_report:
            qa_report_path = latest_report
            print(f"Using latest QA report: {qa_report_path}")
        else:
            # Fallback to default path in current directory
            qa_report_path = "./qa_report.xlsx"
            print(f"No reports found in output directory, using default path: {qa_report_path}")
    
    # Resolve to absolute paths
    qa_report_path = os.path.abspath(qa_report_path)
    output_path = name_assign_output_path
    
    print(f"Using paths:")
    print(f"  QA Report: {qa_report_path}")
    print(f"  Brief: {brief_path}")
    print(f"  Output: {output_path}")
    
    # Define required columns from QA report
    qa_cols_input = [
        'campaign_id', 'campaign_name', 
        'line_item_id', 'line_item_name', 'line_item_alternative_id',
        'creative_id', 'creative_name', 'creative_alternative_id',
        'campaign_active', 'line_item_active', 'creative_active', 'creative_secure'
    ]
    
    # --- Load Data ---
    print(f"Loading QA Report: {qa_report_path}")
    if not os.path.exists(qa_report_path):
        print(f"Error: QA Report file not found at {qa_report_path}")
        return
    try:
        qa_df = pd.read_excel(qa_report_path)
        # Select and rename columns to ensure consistency
        qa_df = qa_df[[col for col in qa_cols_input if col in qa_df.columns]].copy()
        print(f"QA Report loaded successfully. Shape: {qa_df.shape}")
        # Ensure all required columns exist
        missing_qa_cols = [col for col in qa_cols_input if col not in qa_df.columns]
        if missing_qa_cols:
             print(f"Error: Missing required columns in QA report: {', '.join(missing_qa_cols)}")
             return
    except Exception as e:
        print(f"Error loading QA report: {e}")
        return

    # --- Debug: Inspect names immediately after reading QA Report --- 
    print("\n--- Debug: Raw data read from QA Report --- ")
    # Print head to see multiple rows
    print(qa_df[['line_item_name', 'creative_name']].head().to_string()) 
    # Print repr of a specific problematic cell if possible (e.g., first row)
    if not qa_df.empty:
        print(f"Debug: Repr of first LI Name read: {repr(qa_df.iloc[0].get('line_item_name', 'N/A'))}")
        print(f"Debug: Repr of first Creative Name read: {repr(qa_df.iloc[0].get('creative_name', 'N/A'))}")
    print("-------------------------------------------"
)
    # --- End Debug ---

    print(f"Loading and processing Campaign Brief: {brief_path}")
    if not os.path.exists(brief_path):
        print(f"Error: Campaign Brief file not found at {brief_path}")
        return
    try:
        # Extract structured data using brief_extractor
        structured_brief_data = extract_structured_brief_data(brief_path)
        
        # Check if extraction was successful
        if not structured_brief_data or not any(df is not None and not df.empty for df in structured_brief_data.values()):
             print("Error: Failed to extract any structured data from the brief.")
             error_df = pd.DataFrame({'Error': ["Failed to extract structured data from Campaign Brief."]})
             error_df.to_excel(output_path, index=False)
             print(f"Error report saved to {output_path}")
             return

        # Get individual dataframes
        account_data = structured_brief_data.get('account_data')
        campaign_data = structured_brief_data.get('campaign_data')
        placement_data = structured_brief_data.get('placement_data')
        target_data = structured_brief_data.get('target_data') 
        
        print("Brief data extracted successfully.")
        
    except Exception as e:
        print(f"Error processing Campaign Brief: {e}")
        return

    # --- Extract Key Information from Brief (used for checks) ---
    campaign_year = None
    product_type_str_brief = None # Store raw product type string
    product_short_forms = []
    measurement_type_str_brief = None # Store raw measurement type string
    is_hub = False
    is_ifo = False # Added for IFO check
    is_lda_required = False # Added for LDA check
    viewability_goal_str_brief = None # Store raw viewability goal string
    viewability_perc = None
    lda_compliant_str_brief = None # Store raw LDA compliant string

    # 1. Campaign Year (from IO Campaign Start Date in Campaign Level)
    if campaign_data is not None:
        start_date_str = get_field_value(campaign_data, 'IO Campaign Start Date')
        if start_date_str:
            start_date = safe_date_convert(start_date_str)
            if start_date:
                campaign_year = start_date.year
                print(f"Determined Campaign Year: {campaign_year}")
            else:
                print("Warning: Could not parse IO Campaign Start Date to determine year.")
        
             
    # 2. Product Type(s) (from Product Type in Account Level)
    if account_data is not None and not account_data.empty:
        # Account data is likely a single row DataFrame, access by column name
        product_type_col = None
        possible_pt_cols = ['Product Type', 'product type', 'Campaign Type'] # Add variations if needed
        for col in possible_pt_cols:
             # Check case-insensitively
             if any(existing_col.lower() == col.lower() for existing_col in account_data.columns):
                 # Find the original case column name
                 product_type_col = next((existing_col for existing_col in account_data.columns if existing_col.lower() == col.lower()), None)
                 break
        
        if product_type_col:
            product_type_str_brief = account_data.iloc[0].get(product_type_col)
            if pd.notna(product_type_str_brief):
                 product_type_str_brief = str(product_type_str_brief).strip()
                 print(f"Found Product Type string: {product_type_str_brief}")
                 delimiters = r'[;,|/+]' 
                 product_types = re.split(delimiters, product_type_str_brief)
                 for pt in product_types:
                     pt = pt.strip()
                     if pt:
                         short_form = extract_product_type_shortform(pt)
                         if short_form and short_form not in product_short_forms:
                             product_short_forms.append(short_form)
                 print(f"Determined Product Type Short Forms: {product_short_forms}")
            else:
                 print("Warning: Product Type column found but value is empty/NaN in Account Data.")
                 product_type_str_brief = None # Ensure it's None if empty
        else:
            print(f"Warning: Product Type column not found in Account Data. Available columns: {list(account_data.columns)}")
            
    # 3. Measurement Type & HUB/IFO Check (from Measurement Type in Campaign Level)
    if campaign_data is not None:
        measurement_type_str_brief = get_field_value(campaign_data, 'Measurement Type')
        if measurement_type_str_brief and isinstance(measurement_type_str_brief, str):
             print(f"Found Measurement Type string: {measurement_type_str_brief}")
             measurement_upper = measurement_type_str_brief.upper()
             if 'HUB:' in measurement_upper:
                  is_hub = True
                  print("HUB indicator found in Measurement Type.")
             if 'IFO:' in measurement_upper or 'IN-FLIGHT OPTIMIZATION' in measurement_upper:
                  is_ifo = True
                  print("IFO indicator found in Measurement Type.")
             
    # 4. Viewability Percentage (from Viewability Goal in Campaign Level)
    if campaign_data is not None:
        viewability_goal_str_brief = get_field_value(campaign_data, 'Viewability Goal')
        if viewability_goal_str_brief:
            viewability_perc = extract_viewability_percentage(viewability_goal_str_brief)
            if viewability_perc is not None:
                print(f"Determined Viewability Percentage: {viewability_perc}%")
                
    # 5. LDA Requirement (from LDA or Age Compliant in Campaign Level)
    if campaign_data is not None:
        lda_compliant_str = get_field_value(campaign_data, 'LDA or Age Compliant')
        lda_compliant_str_brief = lda_compliant_str # Store the raw value
        if lda_compliant_str and isinstance(lda_compliant_str, str):
            print(f"Found LDA or Age Compliant string: {lda_compliant_str}")
            if lda_compliant_str.strip().lower() == 'yes':
                is_lda_required = True
                print("LDA requirement found ('_LDA_' tag needed in campaign name).")
        else:
            print("Warning: 'LDA or Age Compliant' field not found or invalid in Campaign Data.")

    # --- Prepare DataFrames for Merging ---
    # Standardize key columns and relevant data columns

    # Find column names dynamically (keep existing logic)
        bvp_col_pl, geo_col_pl = None, None 
    bvt_col_tg, bvp_col_tg, platform_media_col_tg = None, None, None

    if placement_data is not None and not placement_data.empty:
        print(f"Preparing Placement Level Data... Columns: {list(placement_data.columns)}")
        possible_bvp_cols = ['bvp', 'bvp id', 'bv placement id', 'placement id', 'bvp placement id']
        possible_geo_cols = ['geo required', 'geo targeting', 'geo', 'geo required?yes/no', 'geo required? yes/no', 'geo required yes no']
        col_map_pl = {str(col).lower().strip(): str(col) for col in placement_data.columns} 
        
        for possible in possible_bvp_cols:
             if possible in col_map_pl: bvp_col_pl = col_map_pl[possible]; break
        for possible in possible_geo_cols:
             if possible in col_map_pl: geo_col_pl = col_map_pl[possible]; break
        
        if geo_col_pl is None and len(placement_data.columns) > 3: 
             print("Geo column not found by name, checking fallback column D (index 3).")
             geo_col_pl = placement_data.columns[3] 
             print(f"Using fallback Geo column: '{geo_col_pl}'")

        print(f"Identified Placement columns: BVP='{bvp_col_pl}', Geo='{geo_col_pl}'")

        if bvp_col_pl:
            placement_data[bvp_col_pl] = placement_data[bvp_col_pl].astype(str).str.strip()
            if geo_col_pl:
                 placement_data[geo_col_pl] = placement_data[geo_col_pl].astype(str).str.strip()
            # Select and rename columns for clarity before merge
            placement_data_to_merge = placement_data[[col for col in [bvp_col_pl, geo_col_pl] if col]].copy()
            rename_dict_pl = {}
            if bvp_col_pl: rename_dict_pl[bvp_col_pl] = 'merge_key_bvp_pl'
            if geo_col_pl: rename_dict_pl[geo_col_pl] = 'brief_geo_required'
            placement_data_to_merge.rename(columns=rename_dict_pl, inplace=True)
            # Keep only unique BVP keys to avoid merge duplication if brief has duplicate BVPs
            placement_data_to_merge = placement_data_to_merge.drop_duplicates(subset=['merge_key_bvp_pl'])

        else:
            print("Warning: Could not find BVP ID column in placement data. Cannot merge Geo details by BVP.")
            placement_data_to_merge = None # Ensure it's None if key column missing
    else:
        print("Warning: Placement data not found or empty in brief.")
        placement_data_to_merge = None

    if target_data is not None and not target_data.empty:
        print(f"Preparing Target Level Data... Columns: {list(target_data.columns)}")
        possible_bvt_cols = ['bv id', 'bvt id', 'bvt'] # BVT is usually labeled 'BV ID'
        possible_bvp_cols = ['bvp', 'bvp id'] 
        possible_platform_media_cols = ['platform/media type', 'platform / media type', 'platform', 'media type', 'platform media type']
        # **** Add specific search for 'BVT' column ****
        possible_real_bvt_cols = ['bvt', 'bv target id']
        
        col_map_tg = {str(col).lower().strip(): str(col) for col in target_data.columns}
        
        # Identify the primary ID column (likely 'BV ID') - Keep this logic
        primary_id_col_tg = None
        for possible in possible_bvt_cols: 
             if possible in col_map_tg: primary_id_col_tg = col_map_tg[possible]; break

        # **** Explicitly find the BVT column to use for merging ****
        merge_key_col_tg = None
        for possible in possible_real_bvt_cols:
            if possible in col_map_tg: merge_key_col_tg = col_map_tg[possible]; break

        # Fallback: If specific BVT column not found, revert to using the primary ID (old behavior, less likely correct)
        if not merge_key_col_tg:
            print(f"Warning: Explicit 'BVT' column not found in Target data, falling back to primary ID '{primary_id_col_tg}' for merge key.")
            merge_key_col_tg = primary_id_col_tg # Use the found 'BV ID' or similar as fallback
        else:
             print(f"Using column '{merge_key_col_tg}' from Target data as the merge key (matching QA Alt ID)." )

        # Find BVP and Platform/Media columns (keep this logic)
        bvp_col_tg = None
        platform_media_col_tg = None
        for possible in possible_bvp_cols:
             if possible in col_map_tg: bvp_col_tg = col_map_tg[possible]; break
        for possible in possible_platform_media_cols:
             if possible in col_map_tg: platform_media_col_tg = col_map_tg[possible]; break
             
        if platform_media_col_tg is None and len(target_data.columns) > 6:
             print("Platform/Media column not found by name, checking fallback column G (index 6).")
             platform_media_col_tg = target_data.columns[6]
             print(f"Using fallback Platform/Media column: '{platform_media_col_tg}'")
             
        print(f"Identified Target columns: MergeKey='{merge_key_col_tg}', BVP='{bvp_col_tg}', Platform/Media='{platform_media_col_tg}'")

        # **** Use the identified merge_key_col_tg for merging ****
        if merge_key_col_tg:
            target_data[merge_key_col_tg] = target_data[merge_key_col_tg].astype(str).str.strip()
            columns_to_select_tg = [merge_key_col_tg]
            # Rename the actual BVT column to the merge key name
            rename_dict_tg = {merge_key_col_tg: 'merge_key_bvt_tg'}

            if bvp_col_tg:
                 target_data[bvp_col_tg] = target_data[bvp_col_tg].astype(str).str.strip()
                 columns_to_select_tg.append(bvp_col_tg)
                 rename_dict_tg[bvp_col_tg] = 'brief_bvp_id'
            if platform_media_col_tg:
                 target_data[platform_media_col_tg] = target_data[platform_media_col_tg].astype(str).str.strip()
                 columns_to_select_tg.append(platform_media_col_tg)
                 rename_dict_tg[platform_media_col_tg] = 'brief_platform_media'

            # Select and rename columns for clarity before merge
            target_data_to_merge = target_data[columns_to_select_tg].copy()
            target_data_to_merge.rename(columns=rename_dict_tg, inplace=True)
            # Keep only unique BVT keys to avoid merge duplication
            # **** Drop duplicates based on the renamed merge key ****
            target_data_to_merge = target_data_to_merge.drop_duplicates(subset=['merge_key_bvt_tg'])
        else:
            print("Warning: Could not find BVT ID (using BVT column) column in target data. Cannot merge target details.")
            target_data_to_merge = None # Ensure it's None if key column missing
    else:
        print("Warning: Target data not found or empty in brief.")
        target_data_to_merge = None

    # --- Merge Brief Data with QA Data ---
    print("Merging QA data with extracted brief data...")
    # Clean QA report's merge key
    qa_df['line_item_alternative_id'] = qa_df['line_item_alternative_id'].astype(str).str.strip().str.lower() # Also convert to lower case
    merged_df = qa_df.copy()

    # Merge Target Data (BVT -> BVP, Platform/Media)
    if target_data_to_merge is not None and 'merge_key_bvt_tg' in target_data_to_merge.columns:
        # --- Debug Print 1: Keys for Target Merge ---
        print("\n--- Debug: Keys for Target Merge ---")
        print("QA Report Key ('line_item_alternative_id'):")
        print(merged_df['line_item_alternative_id'].head())
        print("\nTarget Data Key ('merge_key_bvt_tg'):")
        # Also convert target key to lower case for merge
        target_data_to_merge['merge_key_bvt_tg'] = target_data_to_merge['merge_key_bvt_tg'].astype(str).str.strip().str.lower()
        print(target_data_to_merge['merge_key_bvt_tg'].head())
        print("------------------------------------\n")
        # --- End Debug Print ---

        merged_df = pd.merge(
            merged_df,
            target_data_to_merge,
            left_on='line_item_alternative_id',
            right_on='merge_key_bvt_tg',
            how='left'
        )
        # Drop the merge key column from target
        merged_df.drop(columns=['merge_key_bvt_tg'], inplace=True)
        print("Merged Target Data (BVP, Platform/Media).")

        # --- Debug Print 2: Info after Target Merge ---
        print("\n--- Debug: Info after Target Merge ---")
        merged_df.info()
        print(merged_df[['line_item_alternative_id', 'brief_bvp_id', 'brief_platform_media']].head())
        print("--------------------------------------\n")
        # --- End Debug Print ---
    else:
        # Add placeholder columns if merge didn't happen
        if 'brief_bvp_id' not in merged_df.columns: merged_df['brief_bvp_id'] = pd.NA
        if 'brief_platform_media' not in merged_df.columns: merged_df['brief_platform_media'] = pd.NA
        print("Skipped merging Target Data (missing key column or data).")


    # Merge Placement Data (BVP -> Geo Required) using the BVP obtained from Target merge
    if placement_data_to_merge is not None and 'merge_key_bvp_pl' in placement_data_to_merge.columns and 'brief_bvp_id' in merged_df.columns:
         # Ensure the BVP key in merged_df is string for matching
         merged_df['brief_bvp_id'] = merged_df['brief_bvp_id'].astype(str).str.strip().str.lower() # Also convert to lower case
         # Make sure BVP key in placement_data_to_merge is also string and lower case
         placement_data_to_merge['merge_key_bvp_pl'] = placement_data_to_merge['merge_key_bvp_pl'].astype(str).str.strip().str.lower()

         # --- Debug Print 3: Keys for Placement Merge ---
         print("\n--- Debug: Keys for Placement Merge ---")
         print("Merged Data Key ('brief_bvp_id'):")
         print(merged_df['brief_bvp_id'].head())
         print("\nPlacement Data Key ('merge_key_bvp_pl'):")
         print(placement_data_to_merge['merge_key_bvp_pl'].head())
         print("---------------------------------------"
)
         # --- End Debug Print ---

         merged_df = pd.merge(
            merged_df,
            placement_data_to_merge,
            left_on='brief_bvp_id',
            right_on='merge_key_bvp_pl',
            how='left'
         )
         # Drop the merge key column from placement
         merged_df.drop(columns=['merge_key_bvp_pl'], inplace=True)
         print("Merged Placement Data (Geo Required).")

         # --- Debug Print 4: Info after Placement Merge ---
         print("\n--- Debug: Info after Placement Merge ---")
         merged_df.info()
         print(merged_df[['line_item_alternative_id', 'brief_bvp_id', 'brief_platform_media', 'brief_geo_required']].head())
         print("-----------------------------------------"
)
         # --- End Debug Print ---
    else:
        # Add placeholder column if merge didn't happen
        if 'brief_geo_required' not in merged_df.columns: merged_df['brief_geo_required'] = pd.NA
        print("Skipped merging Placement Data (missing key columns or data).")

    # Fill NA values introduced by merges with 'N/A' for final output consistency
    merge_cols_to_fill = ['brief_bvp_id', 'brief_platform_media', 'brief_geo_required']
    for col in merge_cols_to_fill:
        if col in merged_df.columns:
             merged_df[col] = merged_df[col].fillna('N/A')
        else:
             # Ensure column exists even if merge failed completely
             merged_df[col] = 'N/A'

    # --- Prepare Checks ---
    year_pattern_str = None
    if campaign_year:
        year_str_full = str(campaign_year)
        year_str_short = year_str_full[-2:]
        # Ensure the pattern looks for underscore before the year
        year_pattern_str = f'_{year_str_full}|_{year_str_short}' 
        print(f"Using year pattern for checks: {year_pattern_str}")

    # --- Perform Checks ---
    results_data = [] # Rename to avoid conflict with check_results dictionary
    # Define which columns represent checks that should be True/False and colored
    boolean_check_cols = [
        'has_issues', # Overall flag (simple bool)
        'has_spaces', 'has_special_chars', 'missing_quarter', 'missing_year',
        'missing_product_type',
        'missing_hub_ifo_tag',
        'missing_lda', # Added LDA check column
        'missing_viewability',
        'geo_mismatch', 'platform_mismatch', 'media_type_mismatch',
        'check_active_status', # Consolidated active status check (will hold sets)
        'hub_creative_sharing' # Added for HUB creative uniqueness
    ]
    # Add brief_bvt_id to the context columns
    brief_context_cols = [
        'brief_product_type', 'brief_measurement_type', 'brief_viewability',
        'brief_lda_compliant', # Added LDA value from brief
        'brief_bvt_id', 'brief_bvp_id', 'brief_geo_required', 'brief_platform_media' 
    ] 

    # Define all output columns in the desired order
    output_cols = (
        qa_cols_input + # Original QA cols (now includes active statuses)
        brief_context_cols + # Brief context (now includes LDA value)
        boolean_check_cols # All boolean check results
    )

    print(f"\nProcessing {len(merged_df)} rows from QA report...") # Use merged_df length

    for index, row in merged_df.iterrows(): # Iterate over merged_df
        # Get original QA columns (ensure all defined in qa_cols_input are present)
        result_row_base = {col: row.get(col) for col in qa_cols_input}
        # Store extracted viewability for use in checks
        viewability_perc_for_row = viewability_perc # Use the globally extracted value

        # Initialize brief context for this row - now directly from merged row
        result_row_brief = {
             'brief_product_type': product_type_str_brief if pd.notna(product_type_str_brief) else 'N/A',
             'brief_measurement_type': measurement_type_str_brief if pd.notna(measurement_type_str_brief) else 'N/A',
             'brief_viewability': viewability_goal_str_brief if pd.notna(viewability_goal_str_brief) else 'N/A',
             'brief_lda_compliant': lda_compliant_str_brief if pd.notna(lda_compliant_str_brief) else 'N/A', # Add LDA string
             # Get these directly from the merged row, defaulting to 'N/A' if they somehow got lost (shouldn't happen with fillna)
             'brief_bvt_id': row.get('line_item_alternative_id', 'N/A'), # Use the Alt ID as BVT
             'brief_bvp_id': row.get('brief_bvp_id', 'N/A'),
             'brief_geo_required': row.get('brief_geo_required', 'N/A'),
             'brief_platform_media': row.get('brief_platform_media', 'N/A')
        }

        # Initialize check results for this row with empty sets for checks that track entities
        entity_tracking_checks = [
            'has_spaces', 'has_special_chars', 'missing_quarter', 'missing_year',
            'missing_product_type',
            'missing_hub_ifo_tag',
            'missing_lda', # Added LDA check column
            'missing_viewability', 'geo_mismatch',
            'platform_mismatch', 'media_type_mismatch',
            'check_active_status'
            # Note: hub_creative_sharing is a simple boolean, not entity tracking
        ]
        result_row_checks = {col: set() for col in entity_tracking_checks}
        # Initialize simple boolean checks
        result_row_checks['has_issues'] = False
        result_row_checks['hub_creative_sharing'] = False # Initialize new simple bool check

        campaign_id = row.get('campaign_id')
        line_item_id = row.get('line_item_id')
        creative_id = row.get('creative_id')
        
        campaign_name = str(row.get('campaign_name', ''))
        line_item_name = str(row.get('line_item_name', ''))
        creative_name = str(row.get('creative_name', ''))
        
        # BVT ID already captured in result_row_brief
        bvt_id_qa = result_row_brief['brief_bvt_id']
        
        creative_alt_id = str(row.get('creative_alternative_id', '')).strip() 

        row_type = 'Creative' 
        li_props_for_creative = {} # Store props needed for creative check on this row

        if pd.isna(creative_id) and not pd.isna(line_item_id):
            row_type = 'Line Item'
        elif pd.isna(creative_id) and pd.isna(line_item_id) and not pd.isna(campaign_id):
             row_type = 'Campaign'

        result_row_base['type'] = row_type

        # --- Perform Active Status Checks --- 
        is_campaign_active = row.get('campaign_active')
        is_line_item_active = row.get('line_item_active') # Get LI active status
        is_creative_active = row.get('creative_active')
        is_creative_secure = row.get('creative_secure')

        # Populate the check_active_status set
        active_status_set = result_row_checks['check_active_status']
        if not bool(is_campaign_active) if pd.notna(is_campaign_active) else True:
            active_status_set.add('C')
        if bool(is_line_item_active) if pd.notna(is_line_item_active) else False:
            active_status_set.add('Li')
        if not bool(is_creative_active) if pd.notna(is_creative_active) else True:
             active_status_set.add('Cr_A')
        if is_creative_secure != 1 if pd.notna(is_creative_secure) else True:
             active_status_set.add('Cr_S')

        # --- Check Campaign Name (only once per campaign ID) ---
        campaign_errors = []
        if campaign_name:
            campaign_checks_config = {
                'type': 'campaign', 'year_pattern': year_pattern_str,
                'product_short_forms': product_short_forms, 
                'is_hub': is_hub,
                'is_ifo': is_ifo,
                'is_lda_required': is_lda_required, # Pass LDA flag
                 'viewability_perc': viewability_perc_for_row, # Pass viewability
                'quarter_required': True 
            }
            check_results, campaign_errors = check_naming_format(campaign_name, campaign_checks_config)
            for check_key, has_issue_or_set in check_results.items():
                # Handle the combined HUB/IFO tag check
                if check_key == 'missing_hub_ifo_tag':
                    if has_issue_or_set: # Check if the set is not empty
                        result_row_checks[check_key] = has_issue_or_set # Store the set of missing tags
                # Handle other checks (simple booleans or sets of entity types)
                elif has_issue_or_set is True and check_key in result_row_checks and check_key != 'has_issues':
                    # This now includes missing_lda
                    result_row_checks[check_key].add('C')

        # --- Check Line Item Name (only once per line item ID) ---
        li_errors = []
        if line_item_name:
            # --- Derive Platform/Media Type for this specific Line Item's row ---
            platform_media_raw_brief_li = row.get('brief_platform_media', 'N/A')
            platform = None
            media_type = None
            if isinstance(platform_media_raw_brief_li, str) and platform_media_raw_brief_li != 'N/A':
                 platform, media_type = extract_platform_media_type(platform_media_raw_brief_li)
            # --- End Derivation ---

            # Now calculate prefixes/codes using the derived platform/media_type
            platform_prefixes = get_platform_prefix(platform)
            media_type_code = get_media_type_code(media_type)
            
            # Get Geo boolean derived earlier in the main loop for this row
            geo_required_bool_li = None # Default to None
            geo_raw_li = row.get('brief_geo_required', 'N/A')
            if isinstance(geo_raw_li, str):
                 geo_text_li = geo_raw_li.strip().lower()
                 if geo_text_li in ['yes', 'true', 'y', '1']: geo_required_bool_li = True
                 elif geo_text_li in ['no', 'false', 'n', '0', '']: geo_required_bool_li = False
            
            li_checks_config = {
                'type': 'line_item', 'year_pattern': year_pattern_str,
                'viewability_perc': viewability_perc_for_row,
                'is_geo_required': geo_required_bool_li, # Use boolean derived for this row
                'platform': platform, # Pass the derived platform
                'media_type': media_type, # Pass the derived media_type
                'platform_prefixes': platform_prefixes, # Pass the calculated prefixes
                'media_type_code': media_type_code, # Pass the calculated code
                'viewability_perc': viewability_perc_for_row, # Pass viewability
                'quarter_required': True
            }
            check_results, li_errors = check_naming_format(line_item_name, li_checks_config)
            for check_key, has_issue in check_results.items():
                if has_issue and check_key in result_row_checks:
                    # Only add to sets, skip the summary 'has_issues' key
                    if check_key != 'has_issues':
                        result_row_checks[check_key].add('Li')
            
            # Store results and derived info for creative checks
            li_has_geo_in_name = '_GEO_' in line_item_name.upper()
            li_derived_platform_prefix = next((pfx for pfx in platform_prefixes if line_item_name.upper().startswith(pfx.upper())), None) if platform_prefixes else None
            li_derived_media_type_code = media_type_code if media_type_code and media_type_code.upper() in line_item_name.upper() else None

            # Store properties needed for creative check on this row
            li_props_for_creative = {
                'li_has_geo': li_has_geo_in_name,
                'li_platform_prefix': li_derived_platform_prefix,
                'li_platform': platform,
                'li_media_type_code': li_derived_media_type_code
            }

        # --- Check Creative Name (only once per creative ID) ---
        creative_errors = []
        if creative_name:
             # Get overall measurement type for HUB check
             measurement_type_str_brief_for_creative = result_row_brief['brief_measurement_type']
             
             # Get associated LI's derived properties for comparison
             li_has_geo = li_props_for_creative.get('li_has_geo', False)
             li_platform_prefix = li_props_for_creative.get('li_platform_prefix')
             li_platform = li_props_for_creative.get('li_platform')
             li_media_type_code = li_props_for_creative.get('li_media_type_code')

             creative_checks_config = {
                 'type': 'creative', 'year_pattern': year_pattern_str, 
                 'li_has_geo': li_has_geo, 
                 'li_platform_prefix': li_platform_prefix, 
                 'li_platform': li_platform, # Pass LI platform string
                 'measurement_type_str_brief': measurement_type_str_brief_for_creative, # Pass measurement type
                 'li_media_type_code': li_media_type_code, 
                 'viewability_perc': viewability_perc_for_row, # Pass viewability
                 'quarter_required': True
             }
             check_results, creative_errors = check_naming_format(creative_name, creative_checks_config)
             for check_key, has_issue in check_results.items():
                 if has_issue and check_key in result_row_checks:
                     # Only add to sets, skip the summary 'has_issues' key
                     if check_key != 'has_issues':
                         result_row_checks[check_key].add('Cr')

        # Combine all data for the row
        final_result_row = {**result_row_base, **result_row_brief, **result_row_checks}
        # Recalculate overall 'has_issues' based on any check column having failing entities or being True
        has_entity_issues = any(bool(final_result_row.get(k)) for k in entity_tracking_checks)
        # Only check entity tracking columns now, as simple bools are consolidated or implicitly covered
        final_result_row['has_issues'] = has_entity_issues

        results_data.append(final_result_row)

    # --- Perform HUB Creative Uniqueness Check (after processing all rows) ---
    if is_hub:
        print("\nPerforming HUB Creative Uniqueness Check...")
        # Extract relevant columns, drop missing IDs, keep unique LI/Creative pairs
        creative_li_map_df = merged_df[['line_item_id', 'creative_id']].dropna().drop_duplicates()
        
        # Count unique Line Items per Creative
        creative_li_counts = creative_li_map_df.groupby('creative_id')['line_item_id'].nunique()
        
        # Find creatives linked to more than one LI
        shared_creative_ids = set(creative_li_counts[creative_li_counts > 1].index)
        
        if shared_creative_ids:
            print(f"Warning: Found {len(shared_creative_ids)} Creative IDs used across multiple Line Item IDs: {shared_creative_ids}")
            # Update the results_data for the affected creatives
            for result_row in results_data:
                # Check if the row has a creative ID and if it's in the shared set
                row_creative_id = result_row.get('creative_id')
                if pd.notna(row_creative_id) and row_creative_id in shared_creative_ids:
                    result_row['hub_creative_sharing'] = True
                    # Also update the overall issue flag for this row
                    result_row['has_issues'] = True 
        else:
            print("HUB Creative Uniqueness Check passed: No creatives found shared across multiple Line Items.")
    # --- End HUB Check --- 

    # --- Create Output ---
    print(f"\nGenerating output file: {output_path}")
    # Create DataFrame with the exact column order using the collected results list
    output_df = pd.DataFrame(results_data)
    # Ensure all expected columns exist, fill NaNs appropriately
    for col in output_cols:
        if col not in output_df.columns:
            # Default boolean checks to False, others to None/NaN
            default_val = False if col in boolean_check_cols else None
            # For entity tracking columns, default to empty set
            if col in entity_tracking_checks:
                output_df[col] = [set() for _ in range(len(output_df))]
            # Handle hub_creative_sharing initialization if column was missing
            elif col == 'hub_creative_sharing':
                 output_df[col] = False 
            else:
                output_df[col] = default_val
            
    # Fill NaN values appropriately
    for col in boolean_check_cols:
         if col not in entity_tracking_checks:
             # Fill hub_creative_sharing if it exists
             if col == 'hub_creative_sharing' and col in output_df.columns:
                 output_df[col] = output_df[col].fillna(False).astype(bool)
             elif col != 'hub_creative_sharing': # Avoid double handling
                 output_df[col] = output_df[col].fillna(False).astype(bool)
         
    # Fill NaN in context columns with 'N/A' for clarity in output
    # Also fill Alt IDs if missing
    context_and_alt_id_cols = brief_context_cols + ['line_item_alternative_id', 'creative_alternative_id']
    for col in context_and_alt_id_cols:
         if col in output_df.columns:
              # Special handling for lda_compliant: Fillna with N/A if it's missing
              if col == 'brief_lda_compliant':
                  output_df[col] = output_df[col].fillna('N/A')
              # Keep existing fillna for other context/alt_id columns
              elif col not in ['brief_lda_compliant']:
                  output_df[col] = output_df[col].fillna('N/A')
         
    # Reorder columns to the desired final structure
    output_df = output_df[output_cols]
    
    # Use openpyxl to write formatted Excel
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Naming Check Results"

        # Write header row
        ws.append(output_cols)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid") # Light grey
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting to header (Row 1)
        for col_idx, col_name in enumerate(output_cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # Set row height for header
            ws.row_dimensions[1].height = 30
            
        # --- Add Description Row (Row 2) ---
        description_row_values = [column_descriptions.get(col, '') for col in output_cols]
        ws.append(description_row_values) # Appends to the next available row (which is 2)
        desc_font = Font(italic=True, size=9)
        desc_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        desc_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid") # Lighter grey
        
        # Apply formatting to description row (Row 2)
        for col_idx, desc in enumerate(description_row_values, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = desc_font
            cell.fill = desc_fill
            cell.alignment = desc_alignment
        # Set row height for descriptions
        ws.row_dimensions[2].height = 60 # Increased height for wrapped text
        # --- End Description Row --- 

        # Define fills for data cells (True=Red, False=Green)
        true_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid") # Light red
        false_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid") # Light green
        
        # Write data rows and apply conditional formatting
        col_indices = {name: i + 1 for i, name in enumerate(output_cols)}

        # Start data rows from Row 3 now
        for r_idx, row_data in enumerate(output_df.to_dict(orient='records'), start=3): 
            # Write row data - ensure boolean values are written as TRUE/FALSE
            excel_row = []
            row_comments = {} # Store comments for this row
            for col_name in output_cols:
                 value = row_data.get(col_name)
                 output_value = value # Default output is the value itself

                 # Handle boolean specifically for Excel TRUE/FALSE
                 if isinstance(value, (bool, np.bool_)):
                      output_value = bool(value)
                 elif pd.isna(value) or value == 'N/A': # Check for our filled N/A too
                      output_value = '' # Write empty string for NaN/None/N/A
                 elif col_name in entity_tracking_checks and isinstance(value, set) and value:
                      # --- Special formatting for combined HUB/IFO tag check ---
                      if col_name == 'missing_hub_ifo_tag':
                          # Format as TRUE - TAG1, TAG2
                          missing_tags = ", ".join(sorted(list(value)))
                          output_value = f'TRUE - {missing_tags}'
                      # --- Formatting for other entity tracking checks ---
                      else:
                          # Format entity check: TRUE - C, Li, Cr (or Cr_A, Cr_S for active status)
                          sorted_entities = ", ".join(sorted(list(value)))
                          output_value = f'TRUE - {sorted_entities}'
                 elif col_name in entity_tracking_checks and isinstance(value, set) and not value:
                      # Empty set means FALSE
                      output_value = False
                      
                 # --- Convert specific ID columns to uppercase for display ---
                 display_upper_cols = ['line_item_alternative_id', 'creative_alternative_id', 'brief_bvt_id', 'brief_bvp_id']
                 if col_name in display_upper_cols and isinstance(output_value, str) and output_value:
                     output_value = output_value.upper()
                 # --- End Uppercase Conversion ---

                 excel_row.append(output_value)

            # Append the formatted row to the worksheet
            ws.append(excel_row)

            # Apply fill based on boolean check columns
            center_align_cols = ['line_item_alternative_id', 'creative_alternative_id', 'brief_bvt_id', 'brief_bvp_id']
            center_alignment = Alignment(horizontal='center', vertical='center') # Define center alignment once
            
            for check_col in boolean_check_cols:
                if check_col in col_indices:
                    col_letter = get_column_letter(col_indices[check_col])
                    cell_coord = f"{col_letter}{r_idx}"
                    cell = ws[cell_coord]
                    # Check if the check failed (True for simple bools, non-empty set for entity checks)
                    check_failed = False
                    if check_col in entity_tracking_checks:
                        check_failed = bool(row_data.get(check_col))
                    else:
                        check_failed = row_data.get(check_col) is True

                    if check_failed:
                        cell.fill = true_fill
                    else:
                         cell.fill = false_fill
                    # Optional: Add alignment to boolean columns
                    cell.alignment = Alignment(horizontal='center')
                    
            # --- Apply center alignment to specific ID columns ---
            for id_col_name in center_align_cols:
                if id_col_name in col_indices:
                    col_letter = get_column_letter(col_indices[id_col_name])
                    cell = ws[f"{col_letter}{r_idx}"]
                    # Apply center alignment, keeping existing font/fill
                    current_font = cell.font.copy()
                    current_fill = cell.fill.copy()
                    cell.alignment = center_alignment
                    cell.font = current_font # Reapply font
                    cell.fill = current_fill # Reapply fill (might not be needed but safe)
            # --- End center alignment --- 

        # Auto-adjust column widths (optional, can be slow for large files)
        print("Adjusting column widths...")
        for col_idx, column_title in enumerate(output_cols, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            # Check header length (Row 1)
            max_length = max(len(str(ws[f"{column_letter}1"].value)), max_length)
            # Check description length (Row 2)
            # Estimate wrapped length based on width - this is tricky
            # Let's prioritize data length and header/desc will wrap
            # max_length = max(len(str(ws[f"{column_letter}2"].value)) * 0.5, max_length) # Rough estimate
            
            # Check data length (sample a few rows for speed?)
            # for r_idx in range(3, min(ws.max_row + 1, 103)): # Check desc + 100 data rows
            for r_idx in range(3, ws.max_row + 1): # Check all data rows
                cell_value = ws[f"{column_letter}{r_idx}"].value
                if cell_value:
                    # For boolean TRUE/FALSE, consider fixed width?
                    if isinstance(cell_value, bool):
                         max_length = max(max_length, 5) # Length of 'FALSE'
                    else:
                         max_length = max(max_length, len(str(cell_value)))
            
            # Add padding, cap width
            adjusted_width = min(max((max_length + 4), 15), 50) # Min width 15, Max width 50, more padding
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print("Output file saved successfully with formatting.")

    except Exception as e:
        print(f"Error writing formatted Excel output: {e}")
        print("Attempting to save raw data without formatting...")
        try:
            # Ensure boolean columns are strings for raw export if openpyxl fails
            output_df_raw = output_df.copy()
            for col in boolean_check_cols:
                 output_df_raw[col] = output_df_raw[col].astype(str)
            output_df_raw.to_excel(output_path, index=False)
            print("Raw data saved successfully.")
        except Exception as e2:
            print(f"Failed to save raw data: {e2}")

    print("\nName Assignment Check finished.")


if __name__ == "__main__":
    main() 