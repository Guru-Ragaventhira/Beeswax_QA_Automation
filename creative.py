"""
Creative QA Analysis Tool

This script analyzes creative data from QA reports to verify various aspects of creative implementations.

Features:
- Imports creative data from consolidated QA report
- Performs targeted checks on creatives:
  * Creative naming validation
  * Click URL and landing page URL matching
  * Technology vendor validation (must be 632)
  * Size validation based on naming conventions
  * Video-specific attribute validation
  * Creative addons validation ([4] for banner, [7] for video)
  * URL security validation (ensuring all URLs use https)
  * Dimension/duration presence in name and other fields
  * LDA or Age Compliance information from brief

Usage:
The script uses paths defined in the environment file (.env):
- QA_REPORT_PATH: Path to the QA report Excel file
- CREATIVE_OUTPUT_PATH: Path for the output of this script
- BRIEF_PATH: Path to the campaign brief file (for LDA data)

If no QA report is found at the specified path, the script will look for the latest report.
"""

import pandas as pd
import numpy as np
import re
import os
import glob
import json
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import brief extraction function
from brief_extractor import extract_structured_brief_data

def find_latest_qa_report(output_dir):
    """
    Find the latest QA report file in the output directory 
    based on file naming convention with timestamp.
    """
    # Look for QA reports with timestamp in name (qa_report_YYYYMMDD_HHMMSS.xlsx)
    qa_report_files = glob.glob(os.path.join(output_dir, "qa_report_*.xlsx"))
    
    if not qa_report_files:
        print("No QA report files found in output directory")
        return None
    
    # Get the latest one based on modification time
    latest_file = max(qa_report_files, key=os.path.getmtime)
    print(f"Found latest QA report: {latest_file}")
    return latest_file

def check_creative_naming(creative_name, line_item_name, campaign_name=None):
    """
    Check if the creative name contains the line item name or campaign name.
    
    Args:
        creative_name (str): The name of the creative.
        line_item_name (str): The name of the line item.
        campaign_name (str, optional): The name of the campaign.
        
    Returns:
        tuple: (bool, str) - Whether the naming is valid and an issue message if invalid.
    """
    # Handle missing values
    if not creative_name:
        return False, "Creative name is missing."
    
    if not line_item_name and not campaign_name:
        return False, "Both line item name and campaign name are missing."
    
    # Check if line item name is contained within creative name (case-insensitive)
    if line_item_name and line_item_name.lower() in creative_name.lower():
        return True, "Creative name contains line item name."
    
    # Fallback: Check if campaign name is contained within creative name (case-insensitive)
    if campaign_name and campaign_name.lower() in creative_name.lower():
        return True, "Creative name contains campaign name."
    
    # If neither line item name nor campaign name is contained in creative name
    message = "Creative name does not contain either line item or campaign name."
    if line_item_name:
        message += f" Line Item: '{line_item_name}'"
    if campaign_name:
        message += f" Campaign: '{campaign_name}'"
    message += f" Creative: '{creative_name}'"
    
    return False, message

def check_landing_page_url(click_url, landing_page_url, template_id=None):
    """
    Check if the landing page URL matches the click URL.
    
    Args:
        click_url (str): The click URL of the creative.
        landing_page_url (str): The landing page URL extracted from creative attributes.
        template_id (int, optional): The creative template ID.
        
    Returns:
        tuple: (bool, str) - Whether the URLs match and an issue message if invalid.
    """
    # Special handling for template IDs 4 and 6
    if template_id is not None and template_id in [4, 6]:
        # For these templates, click_url might be empty but landing_page_url should exist
        if not landing_page_url or pd.isna(landing_page_url):
            return False, "Landing page URL is missing for template ID 4 or 6."
        return True, f"Template ID {template_id} - Only checking landing page URL exists"
    
    # Standard handling for other templates
    # Handle missing values
    if not click_url or pd.isna(click_url):
        return False, "Click URL is missing."
    
    if not landing_page_url or pd.isna(landing_page_url):
        return False, "Landing page URL is missing."
    
    # Extract domain from URLs for comparison
    def extract_domain(url):
        if not isinstance(url, str):
            return ""
        
        url = url.lower().strip()
        # Remove protocol
        if '://' in url:
            url = url.split('://', 1)[1]
        # Get domain (everything before first slash or query parameter)
        if '/' in url:
            url = url.split('/', 1)[0]
        if '?' in url:
            url = url.split('?', 1)[0]
        return url
    
    click_domain = extract_domain(click_url)
    landing_domain = extract_domain(landing_page_url)
    
    # For tracking/redirect URLs, we may not be able to match directly
    if 'redirect' in click_url.lower() or 'track' in click_url.lower():
        return True, "Click URL contains tracking/redirect - domain matching skipped"
    
    # Check if domains match
    if click_domain and landing_domain and click_domain == landing_domain:
        return True, ""
    else:
        return False, f"Click URL domain and landing page URL domain do not match. Click: '{click_domain}', Landing: '{landing_domain}'"

def check_technology_vendor(vendor_id):
    """
    Check if the creative technology vendor ID is 632.
    
    Args:
        vendor_id: The technology vendor ID extracted from creative attributes.
        
    Returns:
        tuple: (bool, str) - Whether the vendor ID is valid and an issue message if invalid.
    """
    # Handle missing values
    if not vendor_id or pd.isna(vendor_id):
        return False, "Technology vendor ID is missing."
    
    # Convert to string for comparison (in case it's stored as a number)
    vendor_str = str(vendor_id).strip()
    
    if vendor_str == "632":
        return True, ""
    else:
        return False, f"Technology vendor ID should be 632, but found '{vendor_str}'."

def check_creative_size_by_naming(creative_name, width, height, creative_type=None):
    """
    Check if the creative size is allowed based on the creative name prefix.
    Skips check for video creatives.
    
    Args:
        creative_name (str): The name of the creative.
        width: The width of the creative.
        height: The height of the creative.
        creative_type (str, optional): The type of creative.
        
    Returns:
        tuple: (bool, str) - Whether the size is valid and an issue message if invalid.
    """
    # Skip check for video creatives
    if creative_type and isinstance(creative_type, str) and 'video' in creative_type.lower():
        return True, "Size check skipped for video creative"
    
    # Handle missing values
    if not creative_name or pd.isna(creative_name):
        return False, "Creative name is missing."
    
    if pd.isna(width) or pd.isna(height):
        return False, "Creative dimensions are missing."
    
    try:
        width_val = int(width)
        height_val = int(height)
    except (ValueError, TypeError):
        return False, f"Invalid dimensions format: width={width}, height={height}"
    
    # Define allowed sizes by prefix
    mobile_prefixes = ["MOA_", "MOW_", "MO_"]
    mobile_allowed_sizes = [(320, 50), (728, 90), (300, 250)]
    
    desktop_prefixes = ["DE_"]
    desktop_allowed_sizes = [(300, 600), (160, 600), (300, 250), (728, 90)]
    
    creative_name = creative_name.strip()
    
    # Check mobile prefixes
    for prefix in mobile_prefixes:
        if creative_name.startswith(prefix):
            if (width_val, height_val) in mobile_allowed_sizes:
                return True, f"Valid mobile size {width_val}x{height_val} for prefix {prefix}"
            else:
                return False, f"Invalid size {width_val}x{height_val} for mobile prefix {prefix}. Allowed sizes: 320x50, 728x90, 300x250"
    
    # Check desktop prefixes
    for prefix in desktop_prefixes:
        if creative_name.startswith(prefix):
            if (width_val, height_val) in desktop_allowed_sizes:
                return True, f"Valid desktop size {width_val}x{height_val} for prefix {prefix}"
            else:
                return False, f"Invalid size {width_val}x{height_val} for desktop prefix {prefix}. Allowed sizes: 300x600, 160x600, 300x250, 728x90"
    
    # If no prefix match, return True as this check doesn't apply
    return True, "No prefix match for size validation"

def check_video_attributes(creative_type, video_duration, skippable):
    """
    Check video-specific attributes when creative type is 'video'.
    
    Args:
        creative_type (str): The type of creative.
        video_duration: The duration of the video.
        skippable: Whether the video is skippable.
        
    Returns:
        tuple: (bool, str) - Whether the video attributes are valid and an issue message if invalid.
    """
    # If not a video creative, this check doesn't apply
    if not creative_type or pd.isna(creative_type) or 'video' not in creative_type.lower():
        return True, "Not a video creative"
    
    issues = []
    
    # Check video duration
    if not video_duration or pd.isna(video_duration) or video_duration == '':
        issues.append("Video duration is missing for video creative")
    
    # Check skippable flag
    if pd.isna(skippable) or skippable == '':
        issues.append("Skippable flag is missing for video creative")
    elif str(skippable).lower() != 'false':
        issues.append(f"Skippable should be 'False' for video creative, found '{skippable}'")
    
    if issues:
        return False, "; ".join(issues)
    else:
        return True, "Valid video attributes"

def check_creative_addons(creative_type, creative_addons):
    """
    Check if creative_addons match the creative_type.
    
    Args:
        creative_type (str): The type of creative.
        creative_addons (str): The creative addons value.
        
    Returns:
        tuple: (bool, str) - Whether the addons are valid and an issue message if invalid.
    """
    if pd.isna(creative_type) or not isinstance(creative_type, str):
        return False, "Creative type is missing"
    
    if pd.isna(creative_addons) or not isinstance(creative_addons, str):
        return False, "Creative addons value is missing"
    
    creative_type_lower = creative_type.lower()
    
    # Strip any whitespace and check for exact match
    creative_addons = creative_addons.strip()
    
    # Banner type creatives should have creative_addons = "[4]"
    if 'banner' in creative_type_lower or 'display' in creative_type_lower:
        if creative_addons == "[4]":
            return True, "Valid addons for banner creative"
        else:
            return False, f"Banner creative should have addons value '[4]', found '{creative_addons}'"
    
    # Video type creatives should have creative_addons = "[7]"
    elif 'video' in creative_type_lower:
        if creative_addons == "[7]":
            return True, "Valid addons for video creative"
        else:
            return False, f"Video creative should have addons value '[7]', found '{creative_addons}'"
    
    # For other types, just return valid
    return True, "Creative addons check not applicable for this creative type"

def check_all_urls_secure(row):
    """
    Check that all URLs in various fields don't contain 'http:' (only https).
    
    Args:
        row: Row of data with URL fields to check.
        
    Returns:
        tuple: (bool, str) - Whether all URLs are secure and an issue message if any are not.
    """
    # Fields to check for non-secure URLs
    url_fields = [
        'creative_click_url', 
        'creative_pixels', 
        'creative_scripts', 
        'creative_click_trackers', 
        'creative_content_munge',
        'landing_page_url'
    ]
    
    insecure_fields = []
    
    for field in url_fields:
        value = row.get(field)
        if isinstance(value, str) and 'http:' in value:
            insecure_fields.append(field)
    
    if insecure_fields:
        return False, f"Non-secure URLs found in: {', '.join(insecure_fields)}"
    else:
        return True, "All URLs are secure (https)"

def check_dimension_or_duration_in_fields(row):
    """
    Check if creative dimensions (for banners) or duration (for videos) appear in specified fields.
    
    For banners: Check if 'widthxheight' (e.g., 300x250) appears in fields
    For videos: Check if '_duration' (e.g., _15) appears in fields
    
    Special handling for:
    - Template ID 4 with celtra.com in content_munge: Skip size check in content_munge
    - For video creatives: Only check creative_name, skip content_munge and thumbnail_url
    
    Args:
        row: Row of data with fields to check.
        
    Returns:
        tuple: (bool, str) - Whether dimensions/duration are found in fields and details about the check.
    """
    creative_type = row.get('creative_type', '')
    template_id = row.get('creative_template_id')
    
    # Default fields to check
    fields_to_check = ['creative_name', 'creative_content_munge', 'creative_thumbnail_url']
    
    # Adjust fields to check based on template_id and creative_type
    template_specific_fields = []
    skip_fields = []
    
    # Handle template ID 4 with celtra.com in content_munge (for banners)
    if template_id == 4 and row.get('creative_content_munge') and 'celtra.com' in str(row.get('creative_content_munge')):
        skip_fields.append('creative_content_munge')
        template_specific_fields = ['creative_name', 'creative_thumbnail_url']
    
    # Use template specific fields if available, otherwise use all fields
    fields_to_check_actual = template_specific_fields if template_specific_fields else fields_to_check
    
    missing_fields = []
    search_pattern = ''
    
    # For banner/display type creatives, check for dimensions (width x height)
    if isinstance(creative_type, str) and ('banner' in creative_type.lower() or 'display' in creative_type.lower()):
        width = row.get('creative_width')
        height = row.get('creative_height')
        
        if pd.isna(width) or pd.isna(height):
            return False, "Missing width or height dimensions for banner creative"
            
        try:
            width_val = int(width)
            height_val = int(height)
            # Generate patterns to search for: both 300x250 and 300X250
            dimension_pattern_lower = f"{width_val}x{height_val}"
            dimension_pattern_upper = f"{width_val}X{height_val}"
            
            for field in fields_to_check_actual:
                field_value = row.get(field)
                if not isinstance(field_value, str) or (dimension_pattern_lower not in field_value.lower() and dimension_pattern_upper not in field_value.lower()):
                    missing_fields.append(field)
                    
            if missing_fields:
                search_pattern = f"{width_val}x{height_val}"
                skipped_note = f" (Skipped checking {', '.join(skip_fields)})" if skip_fields else ""
                return False, f"Banner dimensions {search_pattern} not found in: {', '.join(missing_fields)}{skipped_note}"
            else:
                skipped_note = f" (Skipped checking {', '.join(skip_fields)})" if skip_fields else ""
                return True, f"Banner dimensions found in all checked fields{skipped_note}"
                
        except (ValueError, TypeError):
            return False, f"Invalid dimensions format: width={width}, height={height}"
            
    # For video type creatives, check for duration (only in creative_name)
    elif isinstance(creative_type, str) and 'video' in creative_type.lower():
        duration = row.get('video_duration')
        
        if pd.isna(duration) or duration == '':
            return False, "Missing duration for video creative"
            
        try:
            duration_val = str(int(duration))
            duration_pattern = f"_{duration_val}"
            
            # For videos, only check creative_name
            video_fields_to_check = ['creative_name']
            skip_fields = ['creative_content_munge', 'creative_thumbnail_url']
            
            for field in video_fields_to_check:
                field_value = row.get(field)
                if not isinstance(field_value, str) or duration_pattern not in field_value:
                    missing_fields.append(field)
                    
            if missing_fields:
                search_pattern = duration_pattern
                return False, f"Video duration {search_pattern} not found in creative_name (Skipped checking content_munge, thumbnail_url)"
            else:
                return True, f"Video duration found in creative_name (Skipped checking content_munge, thumbnail_url)"
                
        except (ValueError, TypeError):
            return False, f"Invalid duration format: {duration}"
    
    # If neither banner nor video, or type not specified
    else:
        return True, "Not a banner or video creative, dimension/duration check not applicable"

def extract_creative_attributes(creative_attributes_str):
    """Extract specific attributes from creative_attributes JSON string
    
    Extracts key attributes from the JSON-like string in creative_attributes
    and returns them as individual values for separate columns
    """
    # Initialize default values
    extracted = {
        'creative_advertiser_category': '',
        'advertiser_domain': '',
        'landing_page_url': '',
        'creative_technology_vendor': '',
        'video_duration': '',
        'skippable': ''
    }
    
    # Return empty values if attributes are missing
    if pd.isna(creative_attributes_str) or not isinstance(creative_attributes_str, str):
        return extracted
    
    try:
        # Clean up the string to make it valid JSON if possible
        # Replace single quotes with double quotes for JSON parsing
        json_str = creative_attributes_str.replace("'", '"')
        
        # Try to parse as JSON
        try:
            attr_dict = json.loads(json_str)
        except json.JSONDecodeError:
            # If standard JSON parsing fails, try regex-based extraction
            attr_dict = {}
            
            # Extract advertiser category using regex
            category_match = re.search(r"'advertiser_category':\s*\[([^\]]+)\]", creative_attributes_str)
            if category_match:
                # Extract text between quotes after removing any extra characters
                category_text = category_match.group(1).strip().strip("'\"")
                attr_dict['advertiser_category'] = category_text
            
            # Extract advertiser domain
            domain_match = re.search(r"'advertiser_domain':\s*\['([^']+)'", creative_attributes_str)
            if domain_match:
                attr_dict['advertiser_domain'] = domain_match.group(1)
                
            # Extract landing page URL
            url_match = re.search(r"'landing_page_url':\s*\['([^']+)'", creative_attributes_str)
            if url_match:
                attr_dict['landing_page_url'] = url_match.group(1)
                
            # Extract creative technology vendor 
            vendor_match = re.search(r"'creative_technology_vendor':\s*\[(\d+)\]", creative_attributes_str)
            if vendor_match:
                attr_dict['creative_technology_vendor'] = vendor_match.group(1)
                
            # Extract video duration
            duration_match = re.search(r"'video_duration':\s*\[(\d+)\]", creative_attributes_str)
            if duration_match:
                attr_dict['video_duration'] = duration_match.group(1)
                
            # Extract skippable status
            skippable_match = re.search(r"'skippable':\s*\[(True|False)\]", creative_attributes_str)
            if skippable_match:
                attr_dict['skippable'] = skippable_match.group(1)
        
        # Extract specific attributes if they exist in the parsed dictionary
        if 'advertiser_category' in attr_dict:
            if isinstance(attr_dict['advertiser_category'], list):
                extracted['creative_advertiser_category'] = ', '.join(str(x) for x in attr_dict['advertiser_category'])
            else:
                extracted['creative_advertiser_category'] = str(attr_dict['advertiser_category'])
                
        if 'advertiser_domain' in attr_dict:
            if isinstance(attr_dict['advertiser_domain'], list):
                extracted['advertiser_domain'] = ', '.join(str(x) for x in attr_dict['advertiser_domain'])
            else:
                extracted['advertiser_domain'] = str(attr_dict['advertiser_domain'])
                
        if 'landing_page_url' in attr_dict:
            if isinstance(attr_dict['landing_page_url'], list):
                extracted['landing_page_url'] = ', '.join(str(x) for x in attr_dict['landing_page_url'])
            else:
                extracted['landing_page_url'] = str(attr_dict['landing_page_url'])
                
        if 'creative_technology_vendor' in attr_dict:
            if isinstance(attr_dict['creative_technology_vendor'], list):
                extracted['creative_technology_vendor'] = ', '.join(str(x) for x in attr_dict['creative_technology_vendor'])
            else:
                extracted['creative_technology_vendor'] = str(attr_dict['creative_technology_vendor'])
                
        if 'video_duration' in attr_dict:
            if isinstance(attr_dict['video_duration'], list):
                extracted['video_duration'] = ', '.join(str(x) for x in attr_dict['video_duration'])
            else:
                extracted['video_duration'] = str(attr_dict['video_duration'])
                
        if 'skippable' in attr_dict:
            if isinstance(attr_dict['skippable'], list):
                extracted['skippable'] = ', '.join(str(x) for x in attr_dict['skippable'])
            else:
                extracted['skippable'] = str(attr_dict['skippable'])
    
    except Exception as e:
        print(f"Error extracting creative attributes: {e}")
    
    return extracted

def check_lda_compliance(lda_compliant, advertiser_category):
    """
    Check if LDA (Legal Drinking Age) compliance is properly set in the advertiser category.
    
    Args:
        lda_compliant (str): LDA compliance value from brief (typically 'Yes' or 'No')
        advertiser_category (str): The advertiser category value from creative attributes
        
    Returns:
        tuple: (bool, str) - Whether the LDA compliance is correct and an issue message if invalid.
    """
    print(f"Checking LDA compliance: LDA={lda_compliant}, Category={advertiser_category}")
    
    if not isinstance(lda_compliant, str) or lda_compliant.strip().upper() != 'YES':
        return True, "Not LDA compliant, no category check needed"
    
    if not isinstance(advertiser_category, str) or not advertiser_category.strip():
        return False, "LDA compliance required but advertiser category is missing"
    
    # Check if advertiser category contains required categories for LDA compliance
    advertiser_category = advertiser_category.upper()
    if 'IAB8_18' in advertiser_category or 'IAB8_5' in advertiser_category:
        return True, "LDA compliance properly set with correct IAB category"
    else:
        return False, "LDA compliance required but advertiser category does not contain IAB8_18 or IAB8_5"

def check_rm_creative_coppa_tag(creative_name, content_munge):
    """
    Check if creatives with "_RM_" in their name have the required COPPA tag.
    
    Args:
        creative_name (str): The name of the creative.
        content_munge (str): The content munge field of the creative.
        
    Returns:
        tuple: (bool, str) - Whether the COPPA tag is present when required and an issue message if invalid.
    """
    # If creative name doesn't contain _RM_, this check is not applicable
    if not creative_name or not isinstance(creative_name, str) or "_RM_" not in creative_name:
        return True, "Not an _RM_ creative, COPPA tag check not applicable."
    
    # Check if content_munge exists
    if not content_munge or not isinstance(content_munge, str):
        return False, "Creative with _RM_ in name is missing content_munge field."
    
    # Check for the required COPPA tag
    required_tag = "<!-- coppa                = raw %%TFCD%% -->"
    if required_tag in content_munge:
        return True, "Required COPPA tag found in _RM_ creative."
    else:
        return False, f"Creative with _RM_ in name is missing required COPPA tag: '{required_tag}'"

def main():
    print("Starting Creative QA Analysis...")
    
    # Load environment variables
    env_path = os.getenv("ENV_PATH", "./input_folder/beeswax_input_qa.env")
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        load_dotenv(env_path)
    else:
        print(f"Warning: Environment file {env_path} not found. Using default environment.")
        load_dotenv()  # Try default locations
    
    # Get paths from environment or use defaults
    brief_path = os.path.abspath(os.getenv("BRIEF_PATH", "./Brief/Campaign_Brief.xlsx"))
    output_dir = os.path.abspath(os.getenv("OUTPUT_DIR", "./output_folder"))
    creative_output_path = os.path.abspath(os.getenv("CREATIVE_OUTPUT_PATH", "./output_raw/creative_qa_output.xlsx"))
    
    # Create output directory if it doesn't exist
    os.makedirs(os.path.dirname(creative_output_path), exist_ok=True)
    
    # Determine QA report path - first check env variable, then find latest
    qa_report_path = os.getenv("QA_REPORT_PATH")
    if not qa_report_path or not os.path.exists(qa_report_path):
        print(f"QA report not found at {qa_report_path}, looking for the latest report...")
        latest_report = find_latest_qa_report(output_dir)
        if latest_report:
            qa_report_path = latest_report
            print(f"Using latest QA report: {qa_report_path}")
        else:
            # Fallback to default path in current directory
            qa_report_path = "./qa_report.xlsx"
            print(f"No reports found in output directory, using default path: {qa_report_path}")
    
    # Resolve to absolute paths
    qa_report_path = os.path.abspath(qa_report_path)
    output_path = creative_output_path
    
    print(f"Using paths:")
    print(f"  QA Report: {qa_report_path}")
    print(f"  Brief: {brief_path}")
    print(f"  Output: {output_path}")
    
    # --- Load Brief Data using brief_extractor.py ---
    lda_compliance = "Unknown"
    lda_compliance_notes = ""
    
    if os.path.exists(brief_path):
        print(f"Loading brief data from: {brief_path}")
        try:
            # Extract structured data from brief using brief_extractor
            structured_brief_data = extract_structured_brief_data(brief_path)
            
            if structured_brief_data and structured_brief_data['campaign_data'] is not None:
                campaign_df = structured_brief_data['campaign_data']
                
                # Try to find LDA compliance information
                if 'Field' in campaign_df.columns and 'Value' in campaign_df.columns:
                    # Search for LDA or Age Compliant fields
                    lda_field_row = campaign_df[campaign_df['Field'].str.contains('LDA or Age Compliant', case=False, na=False)]
                    if not lda_field_row.empty:
                        lda_compliance = lda_field_row.iloc[0]['Value']
                        print(f"Found LDA compliance in brief: {lda_compliance}")
                    
                    # Search for LDA or Age Compliant Notes
                    lda_notes_row = campaign_df[campaign_df['Field'].str.contains('LDA or Age Compliant Notes', case=False, na=False)]
                    if not lda_notes_row.empty:
                        lda_compliance_notes = lda_notes_row.iloc[0]['Value']
                        print(f"Found LDA compliance notes in brief: {lda_compliance_notes}")
            else:
                print("No campaign data found in structured brief")
        except Exception as e:
            print(f"Error loading brief data: {e}")
    else:
        print(f"Brief file not found at: {brief_path}")
    
    # Define required columns from QA report
    qa_cols_input = [
        'campaign_id', 'campaign_name', 'line_item_id', 'line_item_name', 'line_item_alternative_id',
        'creative_id', 'creative_name', 'creative_template_id', 'creative_advertiser_id',
        'creative_alternative_id', 'creative_type', 'creative_width', 'creative_height',
        'creative_click_url', 'creative_attributes', 'creative_addons', 'creative_content',
        'creative_pixels', 'creative_scripts', 'creative_click_trackers', 'creative_content_munge',
        'creative_thumbnail_url', 'creative_start_date', 'creative_end_date'
    ]
    
    # --- Load Data ---
    print(f"Loading QA Report: {qa_report_path}")
    if not os.path.exists(qa_report_path):
        print(f"Error: QA Report file not found at {qa_report_path}")
        return
    
    try:
        # First try to load 'Consolidated Report' tab specifically
        try:
            qa_df = pd.read_excel(qa_report_path, sheet_name='Consolidated Report')
            print("Successfully loaded 'Consolidated Report' tab.")
        except Exception as e:
            print(f"Could not load 'Consolidated Report' tab: {e}")
            print("Trying to load the default sheet instead...")
            qa_df = pd.read_excel(qa_report_path)
        
        # Select only required columns and filter for rows with creative data
        available_cols = [col for col in qa_cols_input if col in qa_df.columns]
        missing_cols = [col for col in qa_cols_input if col not in qa_df.columns]
        
        if missing_cols:
            print(f"Warning: The following requested columns are missing: {', '.join(missing_cols)}")
        
        if not available_cols:
            print("Error: None of the required columns found in the QA report.")
            return
            
        # Create dataframe with available columns
        qa_df = qa_df[available_cols].copy()
        
        # Filter for rows that have creative_id (only process creative data)
        qa_df = qa_df.dropna(subset=['creative_id'])
        
        # Extract attributes from creative_attributes column
        print("Extracting attributes from creative_attributes...")
        
        # Initialize new columns
        qa_df['creative_advertiser_category'] = ''
        qa_df['advertiser_domain'] = ''
        qa_df['landing_page_url'] = ''
        qa_df['creative_technology_vendor'] = ''
        qa_df['video_duration'] = ''
        qa_df['skippable'] = ''
        
        # Add LDA compliance columns
        qa_df['LDA_or_Age_Compliant'] = lda_compliance
        qa_df['LDA_or_Age_Compliant_Notes'] = lda_compliance_notes
        
        # Process each row to extract attributes
        for idx, row in qa_df.iterrows():
            if 'creative_attributes' in row:
                extracted = extract_creative_attributes(row['creative_attributes'])
                for key, value in extracted.items():
                    qa_df.at[idx, key] = value
        
        print(f"QA Report loaded successfully. Processing {len(qa_df)} creatives.")
    except Exception as e:
        print(f"Error loading QA report: {e}")
        return
    
    # --- Perform Checks ---
    print("Performing creative checks...")
    
    # Initialize only the requested result columns
    qa_df['naming_valid'] = False
    qa_df['ClickUrl_LP_match'] = False
    qa_df['technology_vendor_valid'] = False
    qa_df['size_by_naming_valid'] = False
    qa_df['video_attributes_valid'] = False
    qa_df['creative_addons_valid'] = False
    qa_df['all_urls_secure_valid'] = False
    qa_df['dimension_duration_valid'] = False
    qa_df['lda_compliance_valid'] = False
    qa_df['rm_creative_coppa_valid'] = False  # New column for RM creative COPPA tag check
    qa_df['has_issues'] = True  # Default to True, will set to False if all checks pass
    
    # Process each creative
    for idx, row in qa_df.iterrows():
        # Creative naming check
        naming_valid, _ = check_creative_naming(
            row.get('creative_name'),
            row.get('line_item_name'),
            row.get('campaign_name')
        )
        qa_df.at[idx, 'naming_valid'] = naming_valid
        
        # ClickUrl LP match check
        clickurl_lp_match, _ = check_landing_page_url(
            row.get('creative_click_url'),
            row.get('landing_page_url'),
            row.get('creative_template_id')
        )
        qa_df.at[idx, 'ClickUrl_LP_match'] = clickurl_lp_match
        
        # Technology vendor check
        technology_vendor_valid, _ = check_technology_vendor(
            row.get('creative_technology_vendor')
        )
        qa_df.at[idx, 'technology_vendor_valid'] = technology_vendor_valid
        
        # Creative size by naming check
        size_by_naming_valid, _ = check_creative_size_by_naming(
            row.get('creative_name'),
            row.get('creative_width'),
            row.get('creative_height'),
            row.get('creative_type')
        )
        qa_df.at[idx, 'size_by_naming_valid'] = size_by_naming_valid
        
        # Video attributes check
        video_attributes_valid, _ = check_video_attributes(
            row.get('creative_type'),
            row.get('video_duration'),
            row.get('skippable')
        )
        qa_df.at[idx, 'video_attributes_valid'] = video_attributes_valid
        
        # Creative addons check
        creative_addons_valid, _ = check_creative_addons(
            row.get('creative_type'),
            row.get('creative_addons')
        )
        qa_df.at[idx, 'creative_addons_valid'] = creative_addons_valid
        
        # Check all URLs are secure
        all_urls_secure_valid, _ = check_all_urls_secure(row)
        qa_df.at[idx, 'all_urls_secure_valid'] = all_urls_secure_valid
        
        # Check dimension or duration in fields
        dimension_duration_valid, _ = check_dimension_or_duration_in_fields(row)
        qa_df.at[idx, 'dimension_duration_valid'] = dimension_duration_valid
        
        # Check LDA compliance
        lda_compliance_valid, _ = check_lda_compliance(
            row.get('LDA_or_Age_Compliant'),
            row.get('creative_advertiser_category')
        )
        qa_df.at[idx, 'lda_compliance_valid'] = lda_compliance_valid
        
        # Check RM creative COPPA tag
        rm_creative_coppa_valid, _ = check_rm_creative_coppa_tag(
            row.get('creative_name'),
            row.get('creative_content_munge')
        )
        qa_df.at[idx, 'rm_creative_coppa_valid'] = rm_creative_coppa_valid
        
        # Update overall issue flag - only check the requested validations
        qa_df.at[idx, 'has_issues'] = not (
            naming_valid and
            clickurl_lp_match and
            technology_vendor_valid and
            size_by_naming_valid and
            video_attributes_valid and
            creative_addons_valid and
            all_urls_secure_valid and
            dimension_duration_valid and
            lda_compliance_valid and
            rm_creative_coppa_valid
        )
    
    # --- Create Output ---
    print(f"Generating output file: {output_path}")
    
    # Define check result columns for formatting
    check_columns = [
        'has_issues',
        'naming_valid',
        'ClickUrl_LP_match',
        'technology_vendor_valid',
        'size_by_naming_valid',
        'video_attributes_valid',
        'creative_addons_valid',
        'all_urls_secure_valid',
        'dimension_duration_valid',
        'lda_compliance_valid',
        'rm_creative_coppa_valid'
    ]
    
    # Define column descriptions for header comments with enhanced details
    column_descriptions = {
        'has_issues': 'OVERALL CHECK RESULT - TRUE means there are issues that need attention. FALSE means all checks passed successfully.',
        'naming_valid': 'NAMING CHECK - Verifies if the creative name contains its line item name or campaign name.',
        'ClickUrl_LP_match': 'CLICK URL & LANDING PAGE MATCH - Confirms if the landing page URL in creative attributes matches the click URL domain. EXCEPTION: For template IDs 4 and 6, only checks if landing page URL exists.',
        'technology_vendor_valid': 'TECH VENDOR CHECK - Validates if creative technology vendor ID is exactly 632. This is a mandatory requirement for all creatives.',
        'size_by_naming_valid': 'SIZE IN NAME CHECK - For mobile creatives (MO_, MOA_, MOW_ prefix): must be 320x50, 728x90, or 300x250. For desktop (DE_ prefix): must be 300x600, 160x600, 300x250, or 728x90. NOTE: Video creatives are excluded from this check.',
        'video_attributes_valid': 'VIDEO ATTRIBUTES CHECK - For video creatives: verifies if video duration is specified and skippable flag is set to False.',
        'creative_addons_valid': 'CREATIVE ADDONS CHECK - For banner/display creatives: must have [4]. For video creatives: must have [7].',
        'all_urls_secure_valid': 'SECURE URL CHECK - Verifies all URLs in click_url, pixels, scripts, trackers, content_munge, and landing_page use https instead of http.',
        'dimension_duration_valid': 'DIMENSION/DURATION IN FIELDS - For banners: checks if dimensions (e.g., 300x250) appear in name, content_munge, thumbnail_url. For videos: checks if duration (e.g., _15) appears in creative_name only (content_munge and thumbnail_url are skipped). EXCEPTION: For template ID 4 with celtra.com, skips content_munge size check.',
        'lda_compliance_valid': 'LDA COMPLIANCE CHECK - If campaign is LDA compliant (Yes), verifies that creative_advertiser_category contains IAB8_18 or IAB8_5.',
        'rm_creative_coppa_valid': 'RM CREATIVE COPPA CHECK - For creatives with "_RM_" in name, verifies that content_munge contains the required COPPA tag <!-- coppa = raw %%TFCD%% -->.',
        'creative_advertiser_category': 'Advertiser category from creative attributes (IAB8_18 or IAB8_5 required for LDA compliance)',
        'advertiser_domain': 'Domain of the advertiser extracted from creative attributes',
        'landing_page_url': 'Landing page URL extracted from creative attributes',
        'creative_technology_vendor': 'Technology vendor ID from creative attributes',
        'video_duration': 'Duration in seconds for video creatives',
        'skippable': 'Whether the video creative is skippable',
        'LDA_or_Age_Compliant': 'Whether the campaign is LDA (Legal Drinking Age) or Age Compliant as specified in the brief',
        'LDA_or_Age_Compliant_Notes': 'Additional notes regarding LDA or Age Compliance from the brief'
    }
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Creative QA Results"
        
        # Get all column names for the header
        all_columns = list(qa_df.columns)
        
        # Write header row
        ws.append(all_columns)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid") # Light grey
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply formatting to header (Row 1)
        for col_idx, col_name in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # Set row height for header
            ws.row_dimensions[1].height = 30
            
        # Add description row (Row 2)
        description_row_values = [column_descriptions.get(col, '') for col in all_columns]
        ws.append(description_row_values)
        desc_font = Font(italic=True, size=9)
        desc_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        desc_fill = PatternFill(start_color="FFF0F0F0", end_color="FFF0F0F0", fill_type="solid") # Lighter grey
        
        # Apply formatting to description row (Row 2)
        for col_idx, desc in enumerate(description_row_values, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = desc_font
            cell.fill = desc_fill
            cell.alignment = desc_alignment
        # Set row height for descriptions
        ws.row_dimensions[2].height = 60
        
        # Define fills for data cells (True=Green, False=Red)
        true_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid") # Light green
        false_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid") # Light red
        
        # Write data rows
        col_indices = {name: i + 1 for i, name in enumerate(all_columns)}
        
        # Start data rows from Row 3
        for r_idx, row_data in enumerate(qa_df.to_dict(orient='records'), start=3):
            # Write row data
            excel_row = []
            for col_name in all_columns:
                value = row_data.get(col_name)
                
                # Handle special formatting for boolean values
                if col_name in check_columns:
                    excel_row.append(bool(value))
                else:
                    excel_row.append(value)
            
            ws.append(excel_row)
            
            # Apply color formatting to check result columns
            for check_col in check_columns:
                if check_col in col_indices:
                    col_letter = get_column_letter(col_indices[check_col])
                    cell_coord = f"{col_letter}{r_idx}"
                    cell = ws[cell_coord]
                    
                    # For 'has_issues', invert the color logic (True=Red, False=Green)
                    if check_col == 'has_issues':
                        cell.fill = false_fill if row_data.get(check_col) else true_fill
                    else:
                        # For other check columns (True=Green, False=Red)
                        cell.fill = true_fill if row_data.get(check_col) else false_fill
                    
                    # Add center alignment
                    cell.alignment = Alignment(horizontal='center')
            
            # No comments in cells, details are now in separate columns
            
        # Auto-adjust column widths
        print("Adjusting column widths...")
        for col_idx, column_title in enumerate(all_columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header length (Row 1)
            max_length = max(len(str(ws[f"{column_letter}1"].value)), max_length)
            
            # Check data length (sample rows for speed)
            for r_idx in range(3, min(ws.max_row + 1, 103)): # Check 100 data rows
                cell_value = ws[f"{column_letter}{r_idx}"].value
                if cell_value:
                    # For boolean values, use fixed width
                    if isinstance(cell_value, bool):
                        max_length = max(max_length, 5) # Length of 'FALSE'
                    else:
                        max_length = max(max_length, len(str(cell_value)))
            
            # Add padding, cap width
            adjusted_width = min(max((max_length + 4), 15), 50) # Min width 15, Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print("Output file saved successfully with formatting.")
        
    except Exception as e:
        print(f"Error writing formatted Excel output: {e}")
        print("Attempting to save raw data without formatting...")
        try:
            # Fallback to basic Excel output
            qa_df.to_excel(output_path, index=False)
            print("Raw data saved successfully.")
        except Exception as e2:
            print(f"Failed to save raw data: {e2}")
    
    print("\nCreative QA Analysis finished.")
    print(f"Results saved to: {output_path}")

if __name__ == "__main__":
    main() 