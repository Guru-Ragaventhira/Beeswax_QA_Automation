import pandas as pd
import requests
import json
import os
import re
import argparse
from datetime import datetime
from dotenv import load_dotenv
import openpyxl

class BeeswaxQA:
    def __init__(self, brief_path=None, env_path=None, output_dir=None):
        """
        Initialize the BeeswaxQA class with a brief path and output directory.
        
        Args:
            brief_path (str, optional): Path to the campaign brief Excel file. If None, will use path from .env
            env_path (str, optional): Path to .env file with credentials. If None, will look in current directory
            output_dir (str, optional): Directory to save output files. If None, will use path from .env or script directory
        """
        # Load environment variables first to get paths if not provided
        if env_path:
            if not os.path.exists(env_path):
                print(f"Warning: Environment file {env_path} not found. Using default environment settings.")
            else:
                load_dotenv(env_path)
                print(f"Loaded environment from: {env_path}")
        else:
            # Try to load from default locations
            default_loaded = load_dotenv()
            if default_loaded:
                print("Loaded environment from default .env file")
            else:
                print("No .env file found. Using default settings.")
        
        # Get paths from environment variables if not provided as parameters
        self.brief_path = brief_path or os.getenv("BRIEF_PATH")
        self.output_dir = output_dir or os.getenv("OUTPUT_DIR")
        
        # Use fallback defaults if still not set
        if not self.brief_path:
            self.brief_path = "./Brief/Campaign_Brief.xlsx"
            print(f"No brief path specified. Using default: {self.brief_path}")
        
        if not self.output_dir:
            self.output_dir = "./output_folder"
            print(f"No output directory specified. Using default: {self.output_dir}")
        
        # Resolve relative paths (convert to absolute paths)
        self.brief_path = os.path.abspath(self.brief_path)
        self.output_dir = os.path.abspath(self.output_dir)
        
        print(f"Using brief path: {self.brief_path}")
        print(f"Using output directory: {self.output_dir}")
        
        # Check if brief file exists
        if not os.path.exists(self.brief_path):
            print(f"Warning: Brief file not found at {self.brief_path}")
            print("Please make sure the file exists before proceeding.")
        
        # Create output directory if it doesn't exist
        if not os.path.exists(self.output_dir):
            print(f"Creating output directory: {self.output_dir}")
            os.makedirs(self.output_dir)
        
        # API endpoints
        self.login_url = os.getenv("V2_LOGIN_URL", "https://catalina.api.beeswax.com/rest/v2/authenticate")
        self.campaign_url = os.getenv("CAMPAIGN_URL", "https://catalina.api.beeswax.com/rest/v2/campaigns")
        self.lineitem_url = os.getenv("LINEITEM_URL", "https://catalina.api.beeswax.com/rest/v2/line-items")
        self.creative_url = os.getenv("CREATIVE_URL", "https://catalina.api.beeswax.com/rest/v2/creatives")
        self.creative_lineitem_url = os.getenv("CREATIVE_LINEITEM_URL", "https://catalina.api.beeswax.com/rest/v2/line-items/{line_item_id}/creatives")
        self.advertiser_url = os.getenv("ADVERTISER_URL", "https://catalina.api.beeswax.com/rest/v2/advertisers/{id}")
        self.lineitem_export_url = os.getenv("LINEITEM_EXPORT_URL", "https://catalina.api.beeswax.com/rest/v2/line-items/export")
        self.segment_url = os.getenv("SEGMENT_URL", "https://catalina.api.beeswax.com/rest/v2/ref/segment-tree")
        
        # Login credentials
        self.email = os.getenv("LOGIN_EMAIL")
        self.password = os.getenv("PASSWORD")
        
        if not all([self.email, self.password, self.login_url, self.campaign_url, 
                   self.lineitem_url, self.creative_url, self.creative_lineitem_url]):
            raise ValueError("Missing required environment variables for API access")
        
        # Headers for API requests
        self.headers = {
            "Content-Type": "application/json"
        }
        
        # Will hold the extracted data
        self.campaign_data = None
        self.line_item_data = None
        self.creative_data = None
        self.line_item_creatives = {}
        self.advertiser_data = {}  # Will store advertiser data by ID
        self.line_item_targeting_data = None  # Will store targeting data from export
        
        # Initialize IDs
        self.campaign_ids = set()
        self.line_item_ids = set()
        self.creative_ids = set()
        self.advertiser_ids = set()
        
    def login(self):
        """Authenticate with Beeswax API v2.0."""
        print("Authenticating with Beeswax API v2.0...")
        print(f"Using login URL: {self.login_url}")
        print(f"Using email: {self.email}")
        
        login_payload = {
            "email": self.email,
            "password": self.password
        }
        
        try:
            # Create a session to handle cookies
            session = requests.Session()
            
            # Make the authentication request
            response = session.post(
                self.login_url,
                headers=self.headers,
                json=login_payload
            )
            
            print(f"Response status code: {response.status_code}")
            
            response.raise_for_status()
            
            # Get the session cookie (using the correct cookie name 'sessionid')
            session_cookie = response.cookies.get('sessionid')
            if not session_cookie:
                print("Authentication failed: Session cookie not found in response")
                return False
            
            # Update headers with session cookie for future requests
            self.headers["Cookie"] = f"sessionid={session_cookie}"
            print("Authentication successful")
            return True
            
        except Exception as e:
            print(f"Authentication failed: {e}")
            return False
    
    def load_brief(self):
        """Parse the campaign brief to extract alternative IDs for campaigns, line items, and creatives."""
        print(f"Loading brief from {self.brief_path}")
        
        try:
            # Try to load as Excel first
            if self.brief_path.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(self.brief_path)
            # If it's a CSV file
            elif self.brief_path.endswith('.csv'):
                df = pd.read_csv(self.brief_path)
            else:
                raise ValueError("Unsupported file format. Please provide an Excel or CSV file.")
            
            print(f"Loaded brief with {len(df)} rows and {len(df.columns)} columns")
            print(f"Columns in brief: {df.columns.tolist()}")
                
            # Convert to string and flatten the DataFrame to search for IDs
            flat_data = df.astype(str).values.flatten()
            
            # Extract BVI (campaign), BVT (line item), and BVP (creative) IDs using regex
            for value in flat_data:
                # Campaign IDs
                bvi_matches = re.findall(r'BVI\d{10}', str(value))
                for match in bvi_matches:
                    self.campaign_ids.add(match)
                
                # Line item IDs
                bvt_matches = re.findall(r'BVT\d{9}', str(value))
                for match in bvt_matches:
                    self.line_item_ids.add(match)
                
                # Creative IDs
                bvp_matches = re.findall(r'BVP\d{9}', str(value))
                for match in bvp_matches:
                    self.creative_ids.add(match)
            
            print(f"Found {len(self.campaign_ids)} campaign IDs: {self.campaign_ids}")
            print(f"Found {len(self.line_item_ids)} line item IDs: {self.line_item_ids}")
            print(f"Found {len(self.creative_ids)} creative IDs: {self.creative_ids}")
            
        except Exception as e:
            print(f"Error loading brief: {e}")
            raise
    
    def fetch_campaign_data(self):
        """Fetch campaign data from Beeswax API for the extracted campaign IDs."""
        print("Fetching campaign data...")
        print(f"Campaign IDs to fetch: {self.campaign_ids}")
        all_campaigns = []
        
        for campaign_id in self.campaign_ids:
            try:
                url = f"{self.campaign_url}?alternative_id={campaign_id}"
                print(f"Fetching campaign data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                response.raise_for_status()
                
                data = response.json()
                if 'results' in data and data['results']:
                    all_campaigns.extend(data['results'])
                    print(f"Found {len(data['results'])} campaigns for ID {campaign_id}")
                else:
                    print(f"No campaigns found for ID {campaign_id}")
            except Exception as e:
                print(f"Error fetching campaign {campaign_id}: {e}")
        
        self.campaign_data = pd.DataFrame(all_campaigns) if all_campaigns else pd.DataFrame()
        print(f"Fetched data for {len(self.campaign_data)} campaigns")
    
    def fetch_line_item_data(self):
        """Fetch line item data from Beeswax API for the extracted line item IDs."""
        print("Fetching line item data...")
        print(f"Line item IDs to fetch: {self.line_item_ids}")
        all_line_items = []
        
        for line_item_id in self.line_item_ids:
            try:
                url = f"{self.lineitem_url}?alternative_id={line_item_id}"
                print(f"Fetching line item data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                response.raise_for_status()
                
                data = response.json()
                if 'results' in data and data['results']:
                    all_line_items.extend(data['results'])
                    print(f"Found {len(data['results'])} line items for ID {line_item_id}")
                else:
                    print(f"No line items found for ID {line_item_id}")
            except Exception as e:
                print(f"Error fetching line item {line_item_id}: {e}")
        
        self.line_item_data = pd.DataFrame(all_line_items) if all_line_items else pd.DataFrame()
        print(f"Fetched data for {len(self.line_item_data)} line items")
    
    def fetch_line_item_creatives(self):
        """Fetch creative mappings for each line item."""
        print("Fetching line item-creative mappings...")
        
        if self.line_item_data is None or self.line_item_data.empty:
            print("No line item data available. Run fetch_line_item_data first.")
            return
        
        for _, row in self.line_item_data.iterrows():
            line_item_id = row.get('id')
            if line_item_id:
                try:
                    url = self.creative_lineitem_url.format(line_item_id=line_item_id)
                    print(f"Fetching creative mappings for line item {line_item_id}")
                    
                    response = requests.get(url, headers=self.headers)
                    print(f"Response status code: {response.status_code}")
                    
                    response.raise_for_status()
                    
                    data = response.json()
                    if 'results' in data:
                        self.line_item_creatives[line_item_id] = data['results']
                        print(f"Found {len(data['results'])} creative mappings for line item {line_item_id}")
                    else:
                        print(f"No creative mappings found for line item {line_item_id}")
                except Exception as e:
                    print(f"Error fetching creatives for line item {line_item_id}: {e}")
        
        print(f"Fetched creative mappings for {len(self.line_item_creatives)} line items")
    
    def fetch_creative_data(self):
        """Fetch creative data from Beeswax API for the extracted creative IDs."""
        print("Fetching creative data...")
        print(f"Creative IDs to fetch: {self.creative_ids}")
        all_creatives = []
        
        for creative_id in self.creative_ids:
            try:
                url = f"{self.creative_url}?alternative_id={creative_id}"
                print(f"Fetching creative data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                response.raise_for_status()
                
                data = response.json()
                if 'results' in data and data['results']:
                    all_creatives.extend(data['results'])
                    print(f"Found {len(data['results'])} creatives for ID {creative_id}")
                else:
                    print(f"No creatives found for ID {creative_id}")
            except Exception as e:
                print(f"Error fetching creative {creative_id}: {e}")
        
        self.creative_data = pd.DataFrame(all_creatives) if all_creatives else pd.DataFrame()
        print(f"Fetched data for {len(self.creative_data)} creatives")
    
    def fetch_advertiser_data(self):
        """Fetch advertiser data for all advertisers associated with the campaigns."""
        print("Fetching advertiser data...")
        
        # First, collect all unique advertiser IDs from campaign data
        if self.campaign_data is not None and not self.campaign_data.empty:
            advertiser_ids = self.campaign_data['advertiser_id'].unique()
            self.advertiser_ids.update(advertiser_ids)
        
        print(f"Found {len(self.advertiser_ids)} unique advertiser IDs")
        
        # Fetch data for each advertiser
        for advertiser_id in self.advertiser_ids:
            try:
                url = self.advertiser_url.format(id=advertiser_id)
                print(f"Fetching advertiser data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                response.raise_for_status()
                
                data = response.json()
                # Advertiser endpoint returns data directly, not in 'results' array
                if data and isinstance(data, dict) and 'id' in data:
                    self.advertiser_data[advertiser_id] = data
                    print(f"Successfully fetched data for advertiser {advertiser_id}")
                else:
                    print(f"No valid data found for advertiser {advertiser_id}")
            except Exception as e:
                print(f"Error fetching advertiser {advertiser_id}: {e}")
        
        print(f"Fetched data for {len(self.advertiser_data)} advertisers")
    
    def fetch_line_item_targeting(self):
        """Fetch detailed targeting data for line items using the export endpoint."""
        print("Fetching line item targeting data...")
        
        if self.line_item_data is None or self.line_item_data.empty:
            print("No line item data available. Run fetch_line_item_data first.")
            return
        
        # Extract actual line item IDs before column renaming
        line_item_ids = []
        for _, line_item in self.line_item_data.iterrows():
            # Get the ID directly from the API response data
            line_item_id = line_item.get('id')  # Using 'id' instead of 'line_item_id'
            if line_item_id:
                line_item_ids.append(str(line_item_id))
        
        if not line_item_ids:
            print("No line item IDs found in API response data.")
            return
        
        print(f"Found {len(line_item_ids)} line item IDs from API response")
        
        # Split into chunks of 200 IDs as per API limit
        chunk_size = 200
        id_chunks = [line_item_ids[i:i + chunk_size] for i in range(0, len(line_item_ids), chunk_size)]
        
        all_targeting_data = []
        
        for chunk in id_chunks:
            try:
                # Construct URL with comma-separated IDs
                url = f"{self.lineitem_export_url}?ids={','.join(chunk)}"
                print(f"Fetching targeting data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                if response.status_code == 200:
                    # Response is CSV data
                    if 'text/csv' in response.headers.get('content-type', ''):
                        # Use pandas to read CSV data from the response content
                        chunk_data = pd.read_csv(pd.io.common.StringIO(response.text))
                        all_targeting_data.append(chunk_data)
                        print(f"Successfully fetched targeting data for {len(chunk)} line items")
                    else:
                        print(f"Unexpected content type: {response.headers.get('content-type')}")
                else:
                    print(f"Failed to fetch targeting data. Status code: {response.status_code}")
            
            except Exception as e:
                print(f"Error fetching targeting data for chunk: {e}")
        
        if all_targeting_data:
            # Combine all chunks into one DataFrame
            self.line_item_targeting_data = pd.concat(all_targeting_data, ignore_index=True)
            print(f"Total targeting data rows: {len(self.line_item_targeting_data)}")
        else:
            print("No targeting data was fetched")
    
    def fetch_segment_data(self):
        """Fetch segment data for all segments found in targeting data."""
        print("Fetching segment data...")
        
        if self.line_item_targeting_data is None or self.line_item_targeting_data.empty:
            print("No targeting data available. Run fetch_line_item_targeting first.")
            return None
        
        processed_segments = {} # Cache fetched segment details to avoid redundant parsing
        
        # Get unique segments from only Include Segment column
        segments_to_fetch = set()
        
        for _, row in self.line_item_targeting_data.iterrows():
            include_str = str(row.get('Include Segment', ''))
            
            if pd.notna(include_str) and include_str:
                segments_to_fetch.update(key.strip() for key in include_str.split(',') if key.strip())
        
        print(f"Found {len(segments_to_fetch)} unique include segments to fetch")
        
        # Fetch data for unique segments first
        for segment_key in segments_to_fetch:
            if segment_key in processed_segments:  # Skip if already fetched
                continue
            try:
                url = f"{self.segment_url}?key={segment_key}"
                print(f"Fetching segment data from: {url}")
                
                response = requests.get(url, headers=self.headers)
                print(f"Response status code: {response.status_code}")
                
                if response.status_code == 200:
                    data = response.json()
                    if data and isinstance(data, dict):
                        # Extract relevant fields safely
                        results = data.get('results', [])
                        nam_count = None
                        alternative_id = None
                        name = None
                        
                        if results and isinstance(results, list) and len(results) > 0:
                            segment_details = results[0]
                            if isinstance(segment_details, dict):
                                user_count = segment_details.get('user_count_by_region', {})
                                nam_count = user_count.get('NAM', 0) if isinstance(user_count, dict) else 0
                                alternative_id = segment_details.get('alternative_id')
                                name = segment_details.get('name')
                        
                        # Check if name contains scope/scoping and alternative_id contains r1_test
                        has_scope = False
                        has_r1_test = False
                        
                        if name:
                            name_lower = name.lower()
                            # Check if 'scope' or 'scoping' is NOT in the name (true = good, false = bad)
                            has_scope = not ('scope' in name_lower or 'scoping' in name_lower)
                        
                        if alternative_id:
                            has_r1_test = 'r1_test' in str(alternative_id).lower()
                        
                        processed_segments[segment_key] = {
                            'NAM Count': nam_count,
                            'Segment Alternative ID': alternative_id,
                            'Segment Name': name,
                            'No Scope/Scoping In Name': has_scope,
                            'Contains r1_test': has_r1_test,
                            'Raw Segment Data': json.dumps(data)
                        }
                        print(f"Successfully processed segment data for {segment_key}")
                    else:
                        processed_segments[segment_key] = {
                            'Raw Segment Data': json.dumps(data),
                            'No Scope/Scoping In Name': True,  # Default to True (pass) when no data
                            'Contains r1_test': False
                        }
                        print(f"Empty response for segment {segment_key}")
                else:
                    processed_segments[segment_key] = {
                        'Raw Segment Data': response.text,
                        'No Scope/Scoping In Name': True,  # Default to True (pass) when no data
                        'Contains r1_test': False
                    }
                    print(f"Failed to fetch segment data for {segment_key}. Status code: {response.status_code}")
            
            except Exception as e:
                processed_segments[segment_key] = {
                    'Raw Segment Data': str(e),
                    'No Scope/Scoping In Name': True,  # Default to True (pass) when no data
                    'Contains r1_test': False
                }
                print(f"Error fetching segment data for {segment_key}: {e}")

        # Now, associate fetched data with line items
        final_segment_data = []
        for _, row in self.line_item_targeting_data.iterrows():
            line_item_id = row.get('Line Item ID')
            line_item_name = row.get('Line Item Name')
            include_str = str(row.get('Include Segment', ''))

            if pd.notna(include_str) and include_str:
                segment_keys = [key.strip() for key in include_str.split(',') if key.strip()]
                for segment_key in segment_keys:
                    segment_details = processed_segments.get(segment_key, {})
                    segment_row = {
                        'Line Item ID': line_item_id,
                        'Line Item Name': line_item_name,
                        'Segment Key': segment_key,
                        **segment_details
                    }
                    final_segment_data.append(segment_row)
        
        if final_segment_data:
            # Define expected columns order
            columns_order = [
                'Line Item ID', 'Line Item Name', 'Segment Name', 'Segment Key',
                'Segment Alternative ID', 'NAM Count', 'No Scope/Scoping In Name', 'Contains r1_test', 
                'Raw Segment Data'
            ]
            df = pd.DataFrame(final_segment_data)
            # Ensure all columns exist, fill missing with None, and reorder
            for col in columns_order:
                if col not in df.columns:
                    df[col] = None
            return df[columns_order]
            
        return None
    
    def merge_data(self):
        """Merge all collected data into a comprehensive report."""
        print("Merging data into comprehensive report...")
        
        # First, rename columns in our dataframes to avoid conflicts
        if self.campaign_data is not None and not self.campaign_data.empty:
            # Add prefix to all campaign columns
            campaign_columns = {col: f'campaign_{col}' for col in self.campaign_data.columns}
            self.campaign_data = self.campaign_data.rename(columns=campaign_columns)
            print(f"Campaign columns after renaming: {self.campaign_data.columns.tolist()}")
        
        if self.line_item_data is not None and not self.line_item_data.empty:
            # Add prefix to all line item columns
            line_item_columns = {col: f'line_item_{col}' for col in self.line_item_data.columns}
            self.line_item_data = self.line_item_data.rename(columns=line_item_columns)
            print(f"Line item columns after renaming: {self.line_item_data.columns.tolist()}")
        
        if self.creative_data is not None and not self.creative_data.empty:
            # Add prefix to all creative columns
            creative_columns = {col: f'creative_{col}' for col in self.creative_data.columns}
            self.creative_data = self.creative_data.rename(columns=creative_columns)
            
            # Clean pixels and scripts data
            try:
                # Clean creative_pixels if it exists
                if 'creative_pixels' in self.creative_data.columns:
                    # Convert to string first to handle all data types safely
                    self.creative_data['creative_pixels'] = self.creative_data['creative_pixels'].astype(str)
                    # Clean the strings
                    self.creative_data['creative_pixels'] = self.creative_data['creative_pixels'].str.strip('[]').str.replace("'", "").str.replace('"', '')
                    print("Successfully cleaned creative_pixels data")
                
                # Clean creative_scripts if it exists
                if 'creative_scripts' in self.creative_data.columns:
                    # Convert to string first to handle all data types safely
                    self.creative_data['creative_scripts'] = self.creative_data['creative_scripts'].astype(str)
                    # Clean the strings
                    self.creative_data['creative_scripts'] = self.creative_data['creative_scripts'].str.strip('[]').str.replace("'", "").str.replace('"', '')
                    print("Successfully cleaned creative_scripts data")
            except Exception as e:
                print(f"Error cleaning creative data: {e}")
            
            print(f"Creative columns after renaming: {self.creative_data.columns.tolist()}")
        
        # Create a base report with line item data
        report_data = []
        
        if self.line_item_data is None or self.line_item_data.empty:
            print("No line item data available. Cannot generate report.")
            return None
        
        for _, line_item in self.line_item_data.iterrows():
            line_item_id = line_item.get('line_item_id')
            
            # For each line item, find its associated campaign
            campaign_data = None
            advertiser_data = None
            if self.campaign_data is not None and not self.campaign_data.empty:
                campaign_matches = self.campaign_data[self.campaign_data['campaign_id'] == line_item.get('line_item_campaign_id')]
                if not campaign_matches.empty:
                    campaign_data = campaign_matches.iloc[0]
                    # Get advertiser data if available
                    # First try advertiser_id, then try campaign_advertiser_id after renaming
                    advertiser_id = campaign_data.get('campaign_advertiser_id')
                    if advertiser_id in self.advertiser_data:
                        advertiser_data = self.advertiser_data[advertiser_id]
                        print(f"Found advertiser data for ID {advertiser_id}")
            
            # Get creative mappings for this line item
            creative_mappings = self.line_item_creatives.get(line_item_id, [])
            
            # If no creatives, still add a row with just line item, campaign, and advertiser data
            if not creative_mappings:
                # Start with all line item fields
                row_data = line_item.to_dict()
                
                # Add all campaign fields if available
                if campaign_data is not None:
                    row_data.update(campaign_data.to_dict())
                
                # Add all advertiser fields if available
                if advertiser_data is not None:
                    advertiser_fields = {f'advertiser_{k}': v for k, v in advertiser_data.items()}
                    row_data.update(advertiser_fields)
                    print(f"Added advertiser fields to row: {list(advertiser_fields.keys())}")
                
                # Add empty creative fields
                if self.creative_data is not None and not self.creative_data.empty:
                    for col in self.creative_data.columns:
                        row_data[col] = None
                
                report_data.append(row_data)
            else:
                # For each creative mapping, create a row
                for creative_mapping in creative_mappings:
                    # Start with all line item fields
                    row_data = line_item.to_dict()
                    
                    # Add all campaign fields if available
                    if campaign_data is not None:
                        row_data.update(campaign_data.to_dict())
                    
                    # Add all advertiser fields if available
                    if advertiser_data is not None:
                        advertiser_fields = {f'advertiser_{k}': v for k, v in advertiser_data.items()}
                        row_data.update(advertiser_fields)
                    
                    # Get creative ID from mapping
                    creative_id = creative_mapping.get('creative', {}).get('id')
                    
                    # Find matching creative in our creative data
                    creative_data = None
                    if self.creative_data is not None and not self.creative_data.empty:
                        creative_matches = self.creative_data[self.creative_data['creative_id'] == creative_id]
                        if not creative_matches.empty:
                            creative_data = creative_matches.iloc[0]
                            row_data.update(creative_data.to_dict())
                    
                    # Add creative mapping fields with proper prefixes
                    creative_info = creative_mapping.get('creative', {})
                    for key, value in creative_info.items():
                        row_data[f'creative_{key}'] = value
                    
                    report_data.append(row_data)
        
        # Create DataFrame from collected data
        report = pd.DataFrame(report_data)
        
        # Reorder columns by type
        advertiser_cols = [col for col in report.columns if col.startswith('advertiser_')]
        campaign_cols = [col for col in report.columns if col.startswith('campaign_')]
        line_item_cols = [col for col in report.columns if col.startswith('line_item_')]
        creative_cols = [col for col in report.columns if col.startswith('creative_')]
        other_cols = [col for col in report.columns if not any(col.startswith(prefix) for prefix in ['advertiser_', 'campaign_', 'line_item_', 'creative_'])]
        
        # Combine columns in desired order
        ordered_columns = advertiser_cols + campaign_cols + line_item_cols + creative_cols + other_cols
        report = report[ordered_columns]
        
        print(f"Generated report with {len(report)} rows")
        print(f"Report columns in order:")
        if other_cols:
            print("\nOther columns:", other_cols)
        
        return report
    
    def generate_qa_report(self):
        """Generate a comprehensive QA report and save it to Excel."""
        # Login first
        if not self.login():
            print("Failed to authenticate. Aborting QA report generation.")
            return None
        
        # Run the entire process
        self.load_brief()
        self.fetch_campaign_data()
        self.fetch_line_item_data()
        self.fetch_line_item_creatives()
        self.fetch_creative_data()
        self.fetch_advertiser_data()
        self.fetch_line_item_targeting()
        
        # Fetch segment data
        segment_data_df = self.fetch_segment_data()
        
        # Merge data into comprehensive report
        consolidated_report = self.merge_data()
        
        if consolidated_report is not None:
            # Create output filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.output_dir, f"qa_report_{timestamp}.xlsx")
            
            # Create Excel writer object
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write consolidated report to first sheet
                consolidated_report.to_excel(writer, sheet_name='Consolidated Report', index=False)
                
                # Get the worksheet
                worksheet_consolidated = writer.sheets['Consolidated Report']
                
                # Define NEW colors for different header types in Consolidated Report
                colors_consolidated = {
                    'advertiser': 'FFDAB9',  # PeachPuff (Pastel Orange)
                    'campaign': 'B0C4DE',    # LightSteelBlue
                    'line_item': '98FB98',   # PaleGreen
                    'creative': 'D8BFD8'     # Thistle (Pastel Purple)
                }
                
                # Set column width and apply colors to Consolidated Report
                for col_num, column_name in enumerate(consolidated_report.columns, 1):
                    column_letter = openpyxl.utils.get_column_letter(col_num)
                    worksheet_consolidated.column_dimensions[column_letter].width = 15  # Set width to 15
                    
                    cell = worksheet_consolidated.cell(row=1, column=col_num)
                    fill = None
                    if column_name.startswith('advertiser_'):
                        fill = openpyxl.styles.PatternFill(start_color=colors_consolidated['advertiser'], end_color=colors_consolidated['advertiser'], fill_type='solid')
                    elif column_name.startswith('campaign_'):
                        fill = openpyxl.styles.PatternFill(start_color=colors_consolidated['campaign'], end_color=colors_consolidated['campaign'], fill_type='solid')
                    elif column_name.startswith('line_item_'):
                        fill = openpyxl.styles.PatternFill(start_color=colors_consolidated['line_item'], end_color=colors_consolidated['line_item'], fill_type='solid')
                    elif column_name.startswith('creative_'):
                        fill = openpyxl.styles.PatternFill(start_color=colors_consolidated['creative'], end_color=colors_consolidated['creative'], fill_type='solid')
                    
                    if fill:
                        cell.fill = fill
                
                # Write targeting data to second sheet if available
                if self.line_item_targeting_data is not None and not self.line_item_targeting_data.empty:
                    self.line_item_targeting_data.to_excel(writer, sheet_name='Targeting Data', index=False)
                    
                    # Apply color coding and set column width for Targeting Data sheet
                    worksheet_targeting = writer.sheets['Targeting Data']
                    header_fill_targeting = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    
                    for col_num in range(1, self.line_item_targeting_data.shape[1] + 1):
                        # Set column width
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        worksheet_targeting.column_dimensions[column_letter].width = 15
                        
                        # Apply header color
                        cell = worksheet_targeting.cell(row=1, column=col_num)
                        cell.fill = header_fill_targeting
                
                # Write segment data to third sheet if available
                if segment_data_df is not None and not segment_data_df.empty:
                    segment_data_df.to_excel(writer, sheet_name='Segment Data', index=False)
                    
                    # Apply color coding and set column width for Segment Data sheet
                    worksheet_segment = writer.sheets['Segment Data']
                    header_fill_segment = openpyxl.styles.PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    
                    for col_num in range(1, segment_data_df.shape[1] + 1):
                        # Set column width
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        worksheet_segment.column_dimensions[column_letter].width = 15
                        
                        # Apply header color
                        cell = worksheet_segment.cell(row=1, column=col_num)
                        cell.fill = header_fill_segment
                    
                    # Apply conditional formatting to 'NAM Count' column
                    nam_col_letter = None
                    scope_col_letter = None
                    r1_test_col_letter = None
                    
                    for col_num, col_name in enumerate(segment_data_df.columns, 1):
                        if col_name == 'NAM Count':
                            nam_col_letter = openpyxl.utils.get_column_letter(col_num)
                        elif col_name == 'No Scope/Scoping In Name':
                            scope_col_letter = openpyxl.utils.get_column_letter(col_num)
                        elif col_name == 'Contains r1_test':
                            r1_test_col_letter = openpyxl.utils.get_column_letter(col_num)
                    
                    green_fill = openpyxl.styles.PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    red_fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    
                    if nam_col_letter:
                        worksheet_segment.conditional_formatting.add(
                            f'{nam_col_letter}2:{nam_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
                        )
                        worksheet_segment.conditional_formatting.add(
                            f'{nam_col_letter}2:{nam_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['0'], fill=red_fill)
                        )
                    
                    # Apply conditional formatting to 'No Scope/Scoping In Name' column
                    if scope_col_letter:
                        worksheet_segment.conditional_formatting.add(
                            f'{scope_col_letter}2:{scope_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['TRUE'], fill=green_fill)
                        )
                        worksheet_segment.conditional_formatting.add(
                            f'{scope_col_letter}2:{scope_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill)
                        )
                    
                    # Apply conditional formatting to 'Contains r1_test' column
                    if r1_test_col_letter:
                        worksheet_segment.conditional_formatting.add(
                            f'{r1_test_col_letter}2:{r1_test_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['TRUE'], fill=green_fill)
                        )
                        worksheet_segment.conditional_formatting.add(
                            f'{r1_test_col_letter}2:{r1_test_col_letter}{worksheet_segment.max_row}',
                            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['FALSE'], fill=red_fill)
                        )
            
            print(f"QA report saved to {filename}")
            return filename
        else:
            print("Failed to generate QA report")
            return None

# Example usage
if __name__ == "__main__":
    import argparse
    
    # Look for the default .env file first in current directory and then in input_folder
    default_env_paths = [
        "./beeswax_input_qa.env",  # Check current directory first
        "./input_folder/beeswax_input_qa.env"  # Then check input_folder
    ]
    
    # Find the first .env file that exists
    env_path = None
    for path in default_env_paths:
        if os.path.exists(path):
            env_path = path
            break
    
    # If no .env file found, use the last path as default (will be created later)
    if not env_path:
        env_path = default_env_paths[-1]
        print(f"Warning: No .env file found. Will attempt to use {env_path}")
        
    # Pre-load the .env file to get paths for arguments
    load_dotenv(env_path)
    
    # Set up command line argument parsing with defaults from .env
    parser = argparse.ArgumentParser(description='Generate QA report for Beeswax campaigns')
    parser.add_argument('--env', dest='env_path', help='Path to .env file', default=None)
    parser.add_argument('--brief', dest='brief_path', help='Path to campaign brief file', default=None)
    parser.add_argument('--output', dest='output_dir', help='Directory to save output', default=None)
    args = parser.parse_args()
    
    # Override env_path if specified in command line
    if args.env_path:
        env_path = args.env_path
        # Reload .env if a different file was specified
        load_dotenv(env_path)
        
    print(f"Using environment file: {env_path}")
    
    # Get paths from environment variables first, then override with command line if specified
    brief_path = os.getenv("BRIEF_PATH", "./Brief/Campaign_Brief.xlsx")
    output_dir = os.getenv("OUTPUT_DIR", "./output_folder")
    
    # Only override with command line args if explicitly provided
    if args.brief_path:
        brief_path = args.brief_path
    if args.output_dir:
        output_dir = args.output_dir
    
    print(f"Starting Beeswax QA with:")
    print(f"  Brief path: {brief_path}")
    print(f"  Output directory: {output_dir}")
    
    # Create QA instance and generate report
    qa = BeeswaxQA(brief_path, env_path, output_dir)
    qa.generate_qa_report()