"""
Targeting Data Validation Script

This script performs specific targeting checks on QA reports, including:
- Country validation (USA)
- Line Item Name based validations for:
  - App Bundle Lists
  - Domain Lists
  - Environment Types
  - Operating Systems
  - Device Types
- Segment and Creative validations

It also incorporates general checks from targeting_general.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import glob
import dotenv
from pathlib import Path
from datetime import datetime

# Import general checks and required functions
from targeting_general import (
    main as run_general_checks,
    find_col_name,
    validate_bidding_value,
    qa_checks,  # Import the qa_checks dictionary that contains all general checks
    check_exact_match,  # Import the check functions we need
    check_is_empty,
    check_is_number_one,
    check_is_false,
    check_is_true,
    check_frequency_duration,
    find_latest_qa_report  # Import the function to find the latest QA report
)

# Find the latest QA report and load environment variables
def load_env():
    """Load environment variables from beeswax_input_qa.env file and find the latest targeting output path"""
    # Try to load from the standard location first
    env_path = "./input_folder/beeswax_input_qa.env"
    if os.path.exists(env_path):
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists("./beeswax_input_qa.env"):
        env_path = "./beeswax_input_qa.env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists("config.env"):
        env_path = "config.env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    elif os.path.exists(".env"):
        env_path = ".env"
        print(f"Loading environment from: {env_path}")
        dotenv.load_dotenv(env_path)
    else:
        print("No environment file found, using defaults")
    
    # Get base directory and target path
    base_dir = os.getcwd()
    targeting_output = os.getenv("TARGETING_OUTPUT_PATH", "./targeting_check_output.xlsx")
    
    # Ensure paths use proper separators for the current OS
    targeting_output = targeting_output.replace('/', os.sep)
    
    # Construct the full path
    if not os.path.isabs(targeting_output):
        output_path = os.path.join(base_dir, targeting_output)
    else:
        output_path = targeting_output
    
    print(f"Using output path: {output_path}")
    
    return output_path

# Add these constants at the top of the file after imports
DEFAULT_CONTENT_CATEGORIES = "IAB1_2;IAB8_5;IAB8_18;IAB7_3;IAB7_5;IAB7_28;IAB7_30;IAB7_39;IAB7_42;IAB26_4;IAB26;IAB26_1;IAB26_2;IAB26_3;IAB11_5;IAB25;IAB25_1;IAB25_2;IAB25_3;IAB25_4;IAB25_5;IAB25_6;IAB25_7;IAB23;IAB23_1;IAB23_10;IAB23_2;IAB23_3;IAB23_4;IAB23_5;IAB23_6;IAB23_7;IAB23_8;IAB23_9;IAB15_1;IAB15_5;IAB14_1;IAB14_3;IAB18_2;IAB19_3;IAB19_19;IAB19_20;IAB19_22;IAB19_30;IAB19_33;IAB24;-1;IAB11;IAB11_1;IAB11_2;IAB11_4;IAB11_3"

ADVERTISER_90_CONTENT_CATEGORIES = "IAB1_2;IAB8_5;IAB8_18;IAB7_3;IAB7_5;IAB7_28;IAB7_30;IAB7_39;IAB7_42;IAB26_4;IAB26;IAB26_1;IAB26_3;IAB26_2;IAB11;IAB11_1;IAB11_2;IAB11_3;IAB11_4;IAB11_5;IAB25_2;IAB25_5;IAB25_7;IAB25_3;IAB25_4;IAB25_6;IAB25_1;IAB25;IAB23;IAB23_1;IAB23_2;IAB23_3;IAB23_4;IAB23_5;IAB23_6;IAB23_7;IAB23_8;IAB23_9;IAB23_10;IAB15_1;IAB15_5;IAB14_1;IAB14_3;IAB18_2;IAB19_3;IAB19_19;IAB19_20;IAB19_22;IAB19_30;IAB19_33;IAB12_1;IAB12_2;IAB12_3;IAB12;IAB13_3"

# Add new constant for CTV app names
CTV_EXCLUDED_APPS = "Atmosphere;NRS TV;My NRS Store;VideoElephantTV;Loop for Retail;Loop;Grocery TV;VideoElephant TV;Loop TV;Retail Media TV"

def check_line_item_type(line_item_name):
    """Determine the type of line item based on its prefix and if it contains _RM_."""
    if not isinstance(line_item_name, str):
        return None, False
    
    line_item_name = line_item_name.upper().strip()
    is_rm = '_RM_' in line_item_name
    
    if line_item_name.startswith('MOA_'):
        return 'MOA', is_rm
    elif line_item_name.startswith('CTV_'):
        return 'CTV', is_rm
    elif line_item_name.startswith('MOW_'):
        return 'MOW', is_rm
    elif line_item_name.startswith('DE_'):
        return 'DE', is_rm
    return None, is_rm

def normalize_list_values(value_str):
    """Helper function to normalize list values for comparison.
    Splits string by semicolon, sorts values, and rejoins them.
    Also handles numeric values and empty/NaN cases."""
    if pd.isna(value_str) or value_str == '':
        return ''
    
    # Convert to string, handling both integers and floats
    if isinstance(value_str, (int, float)):
        # Remove .0 from float values
        value_str = str(int(value_str)) if value_str.is_integer() else str(value_str)
    
    # Split, clean, sort, and rejoin
    values = []
    for v in str(value_str).split(';'):
        v = str(v).strip().lower()
        # Try to convert to float and back to string to normalize numeric values
        try:
            float_val = float(v)
            # If it's a whole number, convert to int to remove .0
            if float_val.is_integer():
                v = str(int(float_val))
            else:
                v = str(float_val)
        except ValueError:
            pass
        values.append(v)
    return ';'.join(sorted(values))

def compare_lists(actual, expected):
    """Helper function to compare two lists regardless of order.
    Returns True if both lists contain the same elements."""
    if pd.isna(actual) and pd.isna(expected):
        return True
    
    # Normalize both values using normalize_list_values
    actual_norm = normalize_list_values(actual)
    expected_norm = normalize_list_values(expected)
    
    # If both are empty after normalization, they match
    if not actual_norm and not expected_norm:
        return True
    
    # Split normalized strings into sets
    actual_set = set(actual_norm.split(';')) if actual_norm else set()
    expected_set = set(expected_norm.split(';')) if expected_norm else set()
    
    # Debug output
    print(f"Comparing lists - Actual: {actual_set}, Expected: {expected_set}")
    
    return actual_set == expected_set

def validate_inventory_source(row):
    """Validate Exclude Inventory Source for _RM_ lines."""
    line_type, is_rm = check_line_item_type(row.get('Line Item Name'))
    if not is_rm:
        return True  # Not an _RM_ line, so no check needed
        
    inventory_source = str(row.get('Exclude Inventory Source', '')).strip()
    return compare_lists(inventory_source, 'ap;out')

def validate_app_bundle_list(row):
    """Validate Exclude App Bundle List based on line item type and LDA compliance."""
    line_type, is_rm = check_line_item_type(row.get('Line Item Name'))
    app_bundle = row.get('Exclude App Bundle List')
    advertiser_id = str(row.get('Advertiser ID', '')).strip()
    lda_compliant = str(row.get('brief_lda_compliant', '')).strip().lower()
    line_item_name = str(row.get('Line Item Name', '')).strip()
    
    print(f"\nDebug App Bundle - Line Item: {line_item_name}")
    print(f"Type: {line_type}, RM: {is_rm}, LDA: {lda_compliant}")
    print(f"App Bundle Value: {app_bundle}")
    
    # Helper function to check if a value is effectively empty
    def is_effectively_empty(value):
        is_empty = not value or pd.isna(value) or str(value).strip() == '' or str(value).lower().strip() == 'nan'
        print(f"Empty check for '{value}': {is_empty}")
        return is_empty
    
    if line_type in ['MOA', 'CTV']:
        if lda_compliant == 'yes':
            result = compare_lists(app_bundle, '353')
            print(f"MOA/CTV LDA check - Expected: 353, Result: {result}")
            return result
        else:
            expected_values = ['174']  # Regular case
            if is_rm and line_type == 'MOA':
                expected_values.append('1351')
            if advertiser_id == '90':
                expected_values.append('1358')
            expected = ';'.join(expected_values)
            result = compare_lists(app_bundle, expected)
            print(f"MOA/CTV non-LDA check - Expected: {expected}, Result: {result}")
            return result
    elif line_type in ['MOW', 'DE']:
        result = is_effectively_empty(app_bundle)
        print(f"MOW/DE check - Should be empty, Result: {result}")
        return result
    
    print("No matching line type found")
    return False

def validate_domain_list(row):
    """Validate Exclude Domain List ID based on line item type and LDA compliance."""
    line_type, is_rm = check_line_item_type(row.get('Line Item Name'))
    domain_list = row.get('Exclude Domain List ID')
    advertiser_id = str(row.get('Advertiser ID', '')).strip()
    lda_compliant = str(row.get('brief_lda_compliant', '')).strip().lower()
    line_item_name = str(row.get('Line Item Name', '')).strip()
    
    print(f"\nDebug Domain List - Line Item: {line_item_name}")
    print(f"Type: {line_type}, RM: {is_rm}, LDA: {lda_compliant}")
    print(f"Domain List Value: {domain_list}")
    
    # Helper function to check if a value is effectively empty
    def is_effectively_empty(value):
        is_empty = not value or pd.isna(value) or str(value).strip() == '' or str(value).lower().strip() == 'nan'
        print(f"Empty check for '{value}': {is_empty}")
        return is_empty
    
    if line_type in ['MOW', 'DE']:
        if lda_compliant == 'yes':
            result = compare_lists(domain_list, '352')
            print(f"MOW/DE LDA check - Expected: 352, Result: {result}")
            return result
        else:
            expected_values = ['94']  # Regular case
            if is_rm and line_type == 'MOW':
                expected_values.append('1352')
            if advertiser_id == '90':
                expected_values.append('1357')
            expected = ';'.join(expected_values)
            result = compare_lists(domain_list, expected)
            print(f"MOW/DE non-LDA check - Expected: {expected}, Result: {result}")
            return result
    elif line_type in ['MOA', 'CTV']:
        result = is_effectively_empty(domain_list)
        print(f"MOA/CTV check - Should be empty, Result: {result}")
        return result
    
    print("No matching line type found")
    return False

def validate_environment_type(row):
    """Validate Include Environment Type based on line item type."""
    line_type, _ = check_line_item_type(row.get('Line Item Name'))
    env_type = str(row.get('Include Environment Type', '')).strip()
    
    if line_type in ['MOA', 'CTV']:
        return env_type == '1'
    elif line_type in ['MOW', 'DE']:
        return env_type == '0'
    return False

def validate_operating_system(row):
    """Validate Include Operating System and Device Type based on line item type.
    Each line type should only have one of these fields filled:
    - MOA/MOW: Only Operating System (Android;iOS)
    - DE: Only Operating System (os x;windows;chrome os)
    - CTV: Only Device Type (6;3;8;7)
    Order of values doesn't matter.
    """
    line_type, _ = check_line_item_type(row.get('Line Item Name'))
    os_value = str(row.get('Include Operating System', '')).strip()
    device_type = str(row.get('Include Device Type', '')).strip()
    
    # Helper function to check if a field is effectively empty
    def is_empty(value):
        return not value or pd.isna(value) or value.lower() == 'nan'
    
    # For CTV lines
    if line_type == 'CTV':
        # Should have device type but no OS
        if not is_empty(os_value):
            print(f"Error: CTV line '{row.get('Line Item Name')}' should not have Operating System value")
            return False
        if is_empty(device_type):
            print(f"Error: CTV line '{row.get('Line Item Name')}' missing Device Type")
            return False
        return compare_lists(device_type, '6;3;8;7')
    
    # For MOA/MOW lines
    elif line_type in ['MOA', 'MOW']:
        # Should have OS but no device type
        if not is_empty(device_type):
            print(f"Error: MOA/MOW line '{row.get('Line Item Name')}' should not have Device Type value")
            return False
        if is_empty(os_value):
            print(f"Error: MOA/MOW line '{row.get('Line Item Name')}' missing Operating System")
            return False
        return compare_lists(os_value, 'android;ios')
    
    # For DE lines
    elif line_type == 'DE':
        # Should have OS but no device type
        if not is_empty(device_type):
            print(f"Error: DE line '{row.get('Line Item Name')}' should not have Device Type value")
            return False
        if is_empty(os_value):
            print(f"Error: DE line '{row.get('Line Item Name')}' missing Operating System")
            return False
        return compare_lists(os_value, 'os x;windows;chrome os')
    
    return False

def validate_country(row):
    """Validate that Include Country contains USA."""
    country = str(row.get('Include Country', '')).strip().upper()
    return 'USA' in country.split(';')

def validate_segment(row):
    """Validate Include Segment contains catalina pattern."""
    segment = str(row.get('Include Segment', '')).strip()
    return bool(segment and 'catalina-' in segment.lower())

def validate_creatives(row):
    """Validate Creatives field is not empty and follows pattern."""
    creatives = str(row.get('Creatives', '')).strip()
    if not creatives:
        return False
    # Check for pattern like (ID;"";"";\d+) - at least one creative
    pattern = r'\(\d+;"";"";(?:\d+)\)'
    return bool(re.search(pattern, creatives))

def validate_content_category(row):
    """Validate Exclude Content Category based on Advertiser ID."""
    content_category = row.get('Exclude Content Category', '')
    advertiser_id = str(row.get('Advertiser ID', '')).strip()
    
    # Check based on Advertiser ID
    if advertiser_id == '90':
        expected_categories = ADVERTISER_90_CONTENT_CATEGORIES
    else:
        expected_categories = DEFAULT_CONTENT_CATEGORIES
    
    # Compare lists regardless of order
    return compare_lists(content_category, expected_categories)

def validate_ctv_app_exclusions(row):
    """Validate Exclude App Name for CTV line items and ensure other line items have empty app names."""
    line_item_name = str(row.get('Line Item Name', '')).strip()
    exclude_app_name = str(row.get('Exclude App Name', '')).strip()
    
    # Helper function to check if a value is effectively empty
    def is_effectively_empty(value):
        return not value or pd.isna(value) or str(value).strip() == '' or str(value).lower().strip() == 'nan'
    
    # For non-CTV line items, app names should be empty
    if not line_item_name.upper().startswith('CTV_'):
        result = is_effectively_empty(exclude_app_name)
        print(f"Non-CTV line '{line_item_name}' - App Name should be empty. Value: '{exclude_app_name}', Result: {result}")
        return result
    
    # For CTV line items
    if not exclude_app_name:
        print(f"CTV line '{line_item_name}' missing required app exclusions")
        return False
    
    # Convert both lists to sets for comparison (case-insensitive)
    expected_apps = set(app.lower().strip() for app in CTV_EXCLUDED_APPS.split(';'))
    actual_apps = set(app.lower().strip() for app in exclude_app_name.split(';'))
    
    # Check if all expected apps are present in actual apps
    missing_apps = expected_apps - actual_apps
    if missing_apps:
        print(f"Missing apps for {line_item_name}: {', '.join(missing_apps)}")
        return False
    
    print(f"CTV line '{line_item_name}' - All required apps present")
    return True

def validate_geo_targeting(row):
    """Validate geo targeting columns when brief_geo_required is Yes."""
    geo_required = str(row.get('brief_geo_required', '')).strip().lower()
    if geo_required != 'yes':
        return True  # No geo targeting required, so check passes
    
    # List of columns to check for geo targeting
    geo_columns = [
        'Include Latitude & Longitude List',
        'Exclude Latitude & Longitude List',
        'Include Metro',
        'Exclude Metro',
        'Include Region',
        'Exclude Region',
        'Include Zip Code List',
        'Exclude Zip Code List'
    ]
    
    # Check if any of the geo columns has data
    for col in geo_columns:
        value = row.get(col, '')
        if pd.notna(value) and str(value).strip():
            return True
    
    return False

def validate_deal_id_list(row):
    """Validate Include Deal ID List based on line item type and LDA compliance."""
    line_item_name = str(row.get('Line Item Name', '')).strip()
    deal_id_list = str(row.get('Include Deal ID List', '')).strip()
    lda_compliant = str(row.get('brief_lda_compliant', '')).strip().lower()
    
    # Helper function to check if a value is effectively empty
    def is_effectively_empty(value):
        return not value or pd.isna(value) or str(value).strip() == '' or str(value).lower().strip() == 'nan'
    
    print(f"\nDebug Deal ID List - Line Item: {line_item_name}")
    print(f"LDA Compliant: {lda_compliant}")
    print(f"Deal ID List Value: {deal_id_list}")
    
    # If not LDA compliant, deal ID list should be empty
    if lda_compliant != 'yes':
        result = is_effectively_empty(deal_id_list)
        print(f"Non-LDA check - Should be empty, Result: {result}")
        return result
    
    # For LDA compliant cases
    if line_item_name.upper().startswith('CTV_'):
        result = compare_lists(deal_id_list, '1454')
        print(f"CTV LDA check - Expected: 1454, Result: {result}")
        return result
    else:
        result = compare_lists(deal_id_list, '194')
        print(f"Non-CTV LDA check - Expected: 194, Result: {result}")
        return result

def validate_video_placement_type(row):
    """Validate Include Video Placement Type for video line items.
    
    For line items with brief_platform_media containing "Mobile/Video" or "Desktop/Video",
    Include Video Placement Type should contain exactly "1" with no other values.
    For other line items, Include Video Placement Type should be empty.
    """
    # Get values with careful handling of different types
    platform_media = str(row.get('brief_platform_media', '')).strip()
    video_placement_type_raw = row.get('Include Video Placement Type', '')
    line_item_name = str(row.get('Line Item Name', '')).strip()
    
    # Helper function to check if a value is effectively empty
    def is_effectively_empty(value):
        if pd.isna(value):
            return True
        if isinstance(value, (str)) and (value.strip() == '' or value.lower().strip() == 'nan'):
            return True
        return False
    
    # Check if this is a video line based on platform_media
    is_video_line = 'Video' in platform_media
    
    print(f"\nDebug Video Placement Type - Line Item: {line_item_name}")
    print(f"Platform Media: '{platform_media}'")
    print(f"Is Video Line: {is_video_line}")
    print(f"Video Placement Type (raw): {type(video_placement_type_raw)} - '{video_placement_type_raw}'")
    
    # For video line items, video placement type should be exactly "1"
    if is_video_line:
        # Handle various formats of the value 1
        
        # Direct number check (handles if the value is stored as a number in pandas)
        if isinstance(video_placement_type_raw, (int, float)):
            result = abs(float(video_placement_type_raw) - 1.0) < 0.01
            print(f"Video line check - Numeric value check (must be exactly 1): {result}")
            return result
        
        # Handle empty cases
        if is_effectively_empty(video_placement_type_raw):
            print(f"Video line check - Empty value for video line, Result: False")
            return False
            
        # Handle string cases - must be exactly "1" with no other values
        if isinstance(video_placement_type_raw, str):
            # Remove any whitespace
            clean_value = video_placement_type_raw.strip()
            
            # Check if the value is exactly "1" or "1.0"
            is_exact_one = clean_value == "1" or clean_value == "1.0"
            
            print(f"Video line check - Exact string '1' check (no other values allowed): {is_exact_one}")
            return is_exact_one
        
        # If we get here, we couldn't find a valid match
        print(f"Video line check - No valid match using any method, Result: False")
        return False
    else:
        # For non-video line items, video placement type should be empty
        result = is_effectively_empty(video_placement_type_raw)
        print(f"Non-video line check - Should be empty, Result: {result}")
        return result

def apply_targeting_checks(qa_df):
    """Apply all targeting-specific checks to the DataFrame."""
    # Initialize check result columns
    check_columns = {
        'Country Check': validate_country,
        'App Bundle Check': validate_app_bundle_list,
        'Domain List Check': validate_domain_list,
        'Environment Type Check': validate_environment_type,
        'OS/Device Check': validate_operating_system,
        'Segment Check': validate_segment,
        'Creatives Check': validate_creatives,
        'Inventory Source Check': validate_inventory_source,
        'Content Category Check': validate_content_category,
        'CTV Apps Check': validate_ctv_app_exclusions,
        'Geo Targeting Check': validate_geo_targeting,
        'Deal ID List Check': validate_deal_id_list,
        'Video Placement Type Check': validate_video_placement_type  # New check
    }
    
    # Add check result columns
    for col_name, check_func in check_columns.items():
        qa_df[col_name] = qa_df.apply(check_func, axis=1)
    
    return qa_df

def get_column_notes():
    """Create dictionary of explanatory notes for each column."""
    notes = {
        # Targeting check columns
        'Country Check': 'Validates that Include Country contains USA',
        'App Bundle Check': 'Validates App Bundle List based on line type. Exceptional News and LDA compliance',
        'Domain List Check': 'Validates App Bundle List based on line type. Exceptional News and LDA compliance',
        'Environment Type Check': 'Validates Environment Type (1=App, 0=Web) based on line type',
        'OS/Device Check': 'Validates OS or Device Type settings based on line type',
        'Segment Check': 'Validates that Include Segment contains catalina pattern',
        'Creatives Check': 'Validates Creative field is not empty',
        'Inventory Source Check': 'Validates Exclude Inventory Source for _RM_ lines',
        'Content Category Check': 'Validates Content Category based on Advertiser ID',
        'CTV Apps Check': 'Validates CTV app exclusions for CTV lines',
        'Geo Targeting Check': 'Validates geo targeting when brief_geo_required is Yes. At least one of the Include/Exclude columns for Metro, Region, Zips, LatLong should be filled.',
        'Deal ID List Check': 'Validates Deal ID List based on line type and LDA compliance',
        'Video Placement Type Check': 'Validates that lines with brief_platform_media containing "Video" have Include Video Placement Type = exactly 1 (and no other values)',
        
        # Other important columns
        'Budget Split %': 'Validates sum is 100% and for LDA lines must be 50% or 100%',
        'Campaign Buffer %': 'Should be 6% for Traffic Info=yes, 3% otherwise',
        'Line Buffer %': 'Should be 6% for Traffic Info=yes, 3% otherwise',
        'Bidding Values': 'Validates correct bidding values configuration - eCPM',
        'Bid Shading': 'Should be empty',
        'campaign_budget_type': 'Should be "impressions"',
        'campaign_impressions_budget': 'Should be formatted as {"lifetime":X,"daily":None}',
        'Programmatic Guaranteed': 'Should be FALSE',
        'Budget Type': 'Should be "impressions"',
        'Pacing': 'Should be "lifetime"',
        'Pacing Behavior': 'Should be "even"',
        'Multiplier': 'Should be "1"',
        'Catchup Behavior': 'Should be "even"',
        'Use Custom Bidding Agent': 'Should be FALSE',
        'Frequency Cap ID Type': 'Should be STANDARD',
        'Frequency Duration': 'Should be (1;1;week)',
        'Use Fallback': 'Should be TRUE'
    }
    return notes

def apply_formatting(wb, qa_df):
    """Apply formatting to the Excel output."""
    ws = wb.active
    
    # Define formats
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, vertical='top')
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    notes_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    notes_font = Font(italic=True, size=9)
    notes_alignment = Alignment(wrap_text=True, vertical='top')
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Apply header formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Insert row for notes (after header row)
    ws.insert_rows(2)
    
    # Get column notes
    column_notes = get_column_notes()
    
    # Apply notes to row 2
    for idx, col_name in enumerate(qa_df.columns, 1):
        cell = ws.cell(row=2, column=idx)
        cell.fill = notes_fill
        cell.font = notes_font
        cell.alignment = notes_alignment
        # Add note for this column if available
        cell.value = column_notes.get(col_name, "")
    
    # Get column indices
    col_indices = {name: idx for idx, name in enumerate(qa_df.columns, 1)}
    
    # Get column indices for check result columns
    check_columns = [
        'Country Check', 'App Bundle Check', 'Domain List Check',
        'Environment Type Check', 'OS/Device Check', 'Segment Check',
        'Creatives Check', 'Inventory Source Check', 'Content Category Check',
        'CTV Apps Check', 'Geo Targeting Check', 'Deal ID List Check',
        'Video Placement Type Check'  # Add the new check column
    ]
    
    # Prepare for budget split validation
    budget_split_col_idx = col_indices.get('Budget Split %')
    alt_id_col_name = find_col_name(qa_df.columns, ['Alt ID', 'line_item_alternative_id'])
    alt_id_col_idx = None
    if alt_id_col_name:
        alt_id_col_idx = col_indices.get(alt_id_col_name)
    
    # Group budget splits by Alt ID to check if they sum to 100%
    alt_id_to_rows = {}
    if budget_split_col_idx and alt_id_col_idx:
        for row in range(3, ws.max_row + 1):  # Start at row 3 because row 2 is now notes
            alt_id_cell = ws.cell(row=row, column=alt_id_col_idx)
            alt_id = str(alt_id_cell.value).strip()
            
            if alt_id not in alt_id_to_rows:
                alt_id_to_rows[alt_id] = []
            alt_id_to_rows[alt_id].append(row)
    
    # Apply Budget Split % validation first - this is a critical change
    # Previously this was done at the end, which might have been overridden
    if budget_split_col_idx and alt_id_col_idx:
        print("\n--- Validating Budget Split Percentages ---")
        for alt_id, rows in alt_id_to_rows.items():
            # Extract budget split values and LDA compliance from all rows with this Alt ID
            split_values = []
            split_rows = []  # Store row indices along with their split values
            
            # Check LDA compliance - we'll need this for additional validation
            lda_compliant = "unknown"  # Default value to track if we found it
            lda_col_name = find_col_name(qa_df.columns, ['brief_lda_compliant'])
            
            if lda_col_name:
                brief_lda_compliant_col = col_indices.get(lda_col_name)
                if brief_lda_compliant_col:
                    # Get LDA compliance from the first row (should be the same for all rows with same Alt ID)
                    first_row = rows[0]
                    lda_cell = ws.cell(row=first_row, column=brief_lda_compliant_col)
                    lda_value = str(lda_cell.value).strip()
                    lda_compliant = lda_value.lower()
                    print(f"  Found LDA value for {alt_id}: '{lda_value}' (normalized to '{lda_compliant}')")
            
            for row in rows:
                split_cell = ws.cell(row=row, column=budget_split_col_idx)
                split_value = str(split_cell.value).strip()
                
                try:
                    # Extract numeric value from percentage
                    if split_value.endswith('%'):
                        numeric_value = float(split_value.rstrip('%'))
                        split_values.append(numeric_value)
                        split_rows.append((row, numeric_value))
                except ValueError:
                    # Skip non-numeric values
                    continue
            
            # Check if sum rounds to 100% (use tolerance for floating-point comparison)
            valid_sum = False
            if len(split_values) == 1:
                # If only one line item, it should be 100%
                valid_sum = abs(split_values[0] - 100.0) < 0.2  # Small tolerance
                print(f"Alt ID {alt_id}: Single line item with split {split_values[0]}% - {'Valid (100%)' if valid_sum else 'Invalid (should be 100%)'}")
            else:
                # Multiple line items - check if they sum to approximately 100%
                sum_value = sum(split_values)
                valid_sum = abs(sum_value - 100.0) < 0.2  # Small tolerance
                print(f"Alt ID {alt_id}: {len(split_values)} line items with splits {split_values} - Sum: {sum_value:.1f}% - {'Valid (≈100%)' if valid_sum else 'Invalid (not ≈100%)'}")
            
            # Apply LDA compliance check if applicable
            if lda_compliant == "yes":
                print(f"  LDA Compliant: Yes - Checking if splits are either 50% or 100%")
                lda_valid = True
                
                # For LDA-compliant items, splits should only be 50% or 100%
                for _, split_value in split_rows:
                    if not (abs(split_value - 50.0) < 0.2 or abs(split_value - 100.0) < 0.2):
                        lda_valid = False
                        print(f"  Invalid split value for LDA: {split_value}% (should be 50% or 100%)")
                
                if not lda_valid:
                    print("  LDA compliance check failed - marking as invalid")
                    valid_sum = False  # Override validation if LDA requirement fails
                else:
                    print("  All split values are valid for LDA compliance")
            
            # Apply formatting based on validation
            for row in rows:
                split_cell = ws.cell(row=row, column=budget_split_col_idx)
                if valid_sum:
                    split_cell.fill = green_fill
                    print(f"  Row {row}: VALID (green)")
                else:
                    split_cell.fill = red_fill
                    print(f"  Row {row}: INVALID (red)")
    
    # Apply formatting for each row - adjust for the new row 2
    for row in range(3, ws.max_row + 1):  # Start at row 3 because row 2 is now notes
        current_row = row
        row_data = qa_df.iloc[row-3]  # Adjust index for data rows
        
        # Apply formatting for targeting check columns
        for check_col in check_columns:
            if check_col in col_indices:
                cell = ws.cell(row=current_row, column=col_indices[check_col])
                if str(cell.value).lower() == 'true':
                    cell.fill = green_fill
                    cell.value = "No Issue"  # Replace True with "No Issue"
                else:
                    cell.fill = red_fill
                    cell.value = "Check Issue"  # Replace False with "Check Issue"
                    # Add debug info for failed checks
                    line_item_name = row_data.get('Line Item Name', 'N/A')
                    print(f"Check failed: {check_col} for Line Item: {line_item_name}")
        
        # Apply formatting for Bid Shading
        if 'Bid Shading' in col_indices:
            cell = ws.cell(row=current_row, column=col_indices['Bid Shading'])
            value = row_data.get('Bid Shading')
            if check_is_empty(value, None):
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # Apply formatting for campaign_impressions_budget
        if 'campaign_impressions_budget' in col_indices:
            cell = ws.cell(row=current_row, column=col_indices['campaign_impressions_budget'])
            value = row_data.get('campaign_impressions_budget')
            try:
                # Remove spaces and single quotes
                value_str = str(value).replace(" ", "").replace("'", '"')
                # Check if it matches the pattern
                if re.match(r'^\{"lifetime":\d+,"daily":None\}$', value_str):
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            except:
                cell.fill = red_fill

        # Apply formatting for general QA checks from targeting_general.py
        for check_name, (check_func, expected) in qa_checks.items():
            col_name = check_name  # The actual column name in the data
            if col_name in col_indices and col_name not in ['Bid Shading', 'campaign_impressions_budget']:
                cell = ws.cell(row=current_row, column=col_indices[col_name])
                if check_func and expected is not None:
                    value = row_data.get(col_name)
                    if check_func(value, expected):
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill
                        print(f"Check failed: {col_name} - Expected: {expected}, Got: {value}")

        # Campaign Budget Type Check
        campaign_budget_type_col = 'campaign_budget_type'
        if campaign_budget_type_col in col_indices:
            cell = ws.cell(row=current_row, column=col_indices[campaign_budget_type_col])
            if str(cell.value).lower().strip() == 'impressions':
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # Bidding Values Check
        bidding_col = 'Bidding Values'
        if bidding_col in col_indices:
            cell = ws.cell(row=current_row, column=col_indices[bidding_col])
            is_valid, _, _, _, _ = validate_bidding_value(row_data)
            if is_valid:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

        # Buffer Checks
        for buffer_col in ['Campaign Buffer %', 'Line Buffer %']:
            if buffer_col in col_indices:
                cell = ws.cell(row=current_row, column=col_indices[buffer_col])
                buffer_value = str(cell.value).strip()
                traffic_info = row_data.get('brief_traffic_info', '')
                
                if buffer_value != '' and buffer_value != 'Error':
                    try:
                        diff_value = float(buffer_value.strip('%').replace('+', ''))
                        expected_diff = 6.0 if str(traffic_info).lower() == 'yes' else 3.0
                        
                        if abs(diff_value - expected_diff) < 0.1:
                            cell.fill = green_fill
                        else:
                            cell.fill = red_fill
                    except ValueError as e:
                        print(f"Warning: Could not parse buffer value '{buffer_value}' in row {current_row}: {e}")
                        cell.fill = red_fill
                else:
                    cell.fill = red_fill
    
    # Auto-adjust column widths with improved readability
    for col_idx, column_title in enumerate(qa_df.columns, 1):
        column_letter = get_column_letter(col_idx)
        # Calculate max length based on header, notes and data
        max_length = len(str(column_title))
        
        # Check notes length
        notes_value = ws[f"{column_letter}2"].value
        if notes_value:
            max_length = max(max_length, min(len(str(notes_value)), 50))  # Cap at 50 chars for width calc
        
        # Check data rows
        for r_idx in range(3, ws.max_row + 1):  # Start from row 3 now
            cell_value = ws[f"{column_letter}{r_idx}"].value
            if cell_value is not None:
                max_length = max(max_length, min(len(str(cell_value)), 50))  # Cap at 50 chars
        
        # Set width with better padding and limits
        if column_title in ['Budget', 'Budget Split %', 'Target Required', 'Campaign_Imps_Required']:
            # Give more width to budget-related columns
            adjusted_width = min(max(max_length + 6, 20), 30)
        elif column_title in check_columns:
            # Give more width to check columns for notes
            adjusted_width = min(max(max_length + 4, 25), 40)
        else:
            # Standard width for other columns
            adjusted_width = min(max(max_length + 4, 12), 40)
        
        ws.column_dimensions[column_letter].width = adjusted_width

    # Center align check columns and buffer columns
    columns_to_center = check_columns + ['Campaign Buffer %', 'Line Buffer %', 'Budget Split %']
    for col_name in columns_to_center:
        if col_name in col_indices:
            col_letter = get_column_letter(col_indices[col_name])
            for r_idx in range(3, ws.max_row + 1):  # Start at row 3 (skip header and notes)
                cell = ws[f"{col_letter}{r_idx}"]
                cell.alignment = Alignment(horizontal='center')
    
    # FINAL HACK: Force correct coloring for Budget Split % based on LDA validation
    # This is added as the very last step to ensure nothing overrides it
    if budget_split_col_idx and alt_id_col_idx:
        print("\n--- FINAL CHECK: Forcing Budget Split % validation with LDA rule ---")
        for alt_id, rows in alt_id_to_rows.items():
            # Extract values again
            split_values = []
            split_rows = []
            
            # Get LDA value
            first_row = rows[0]
            lda_compliant = "unknown"
            lda_col_name = find_col_name(qa_df.columns, ['brief_lda_compliant'])
            if lda_col_name:
                brief_lda_compliant_col = col_indices.get(lda_col_name)
                if brief_lda_compliant_col:
                    lda_cell = ws.cell(row=first_row, column=brief_lda_compliant_col)
                    lda_value = str(lda_cell.value).strip()
                    lda_compliant = lda_value.lower()
                    print(f"  FINAL CHECK: LDA value for {alt_id} is '{lda_compliant}'")
            
            # Collect split values
            for row in rows:
                cell = ws.cell(row=row, column=budget_split_col_idx)
                split_value = str(cell.value).strip()
                try:
                    if split_value.endswith('%'):
                        numeric_value = float(split_value.rstrip('%'))
                        split_values.append(numeric_value)
                        split_rows.append((row, numeric_value))
                except ValueError:
                    continue
            
            # Check sum to 100%
            valid_sum = False
            if len(split_values) == 1:
                valid_sum = abs(split_values[0] - 100.0) < 0.2
            else:
                sum_value = sum(split_values)
                valid_sum = abs(sum_value - 100.0) < 0.2
            
            # LDA validation (stricter enforcement)
            if lda_compliant == "yes":
                print(f"  FINAL CHECK: LDA validation for Alt ID {alt_id} - Splits: {split_values}")
                
                # Check each split value strictly
                lda_valid = True
                for row, value in split_rows:
                    # Must be exactly 50% or 100% (within tolerance)
                    is_valid_split = abs(value - 50.0) < 0.2 or abs(value - 100.0) < 0.2
                    
                    if not is_valid_split:
                        lda_valid = False
                        print(f"    Invalid LDA split at row {row}: {value}% (must be 50% or 100%)")
                
                # If any value fails LDA check, mark ALL rows in the group as invalid
                if not lda_valid:
                    valid_sum = False
                    print(f"    Group {alt_id} FAILS LDA check - All rows will be marked RED")
                else:
                    print(f"    Group {alt_id} PASSES LDA check - All rows will be marked GREEN")
            
            # Force apply formatting - use direct RGB hex codes to ensure consistency
            for row in rows:
                cell = ws.cell(row=row, column=budget_split_col_idx)
                if valid_sum:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    print(f"    Row {row}: FORCE GREEN")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                    print(f"    Row {row}: FORCE RED")

def apply_standard_checks(row, ws, current_row, col_indices, green_fill, red_fill):
    """Apply standard checks for specific columns with expected values."""
    standard_checks = {
        'Programmatic Guaranteed': ('FALSE', True),  # (expected_value, case_sensitive)
        'Budget Type': ('impressions', False),
        'Pacing': ('lifetime', False),
        'Pacing Behavior': ('even', False),
        'Multiplier': ('1', True),
        'Catchup Behavior': ('even', False),
        'Bid Shading': ('', True),  # Should be empty
        'Use Custom Bidding Agent': ('FALSE', True),
        'Frequency Cap ID Type': ('STANDARD', True),
        'Frequency Duration': ('(1;1;week)', True),
        'Use Fallback': ('TRUE', True)
    }
    
    for col_name, (expected_value, case_sensitive) in standard_checks.items():
        if col_name in col_indices:
            cell = ws.cell(row=current_row, column=col_indices[col_name])
            cell_value = str(row.get(col_name, '')).strip()
            
            if not case_sensitive:
                cell_value = cell_value.lower()
                expected_value = expected_value.lower()
            
            # Special handling for empty check
            if expected_value == '':
                is_valid = cell_value == ''
            else:
                is_valid = cell_value == expected_value
            
            cell.fill = green_fill if is_valid else red_fill

def main():
    print("Starting Targeting Check...")
    
    # Run general checks first
    merged_df = run_general_checks()
    
    if merged_df is None:
        print("Error: General targeting checks failed. Cannot proceed with specific targeting checks.")
        return
        
    print("\nStarting specific targeting checks...")
    
    # --- Configuration ---
    output_path = load_env()
    print(f"Output will be saved to: {output_path}")
    
    # Create output directory if needed
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")
    
    try:
        # Apply targeting-specific checks
        print("\nApplying targeting-specific checks...")
        merged_df = apply_targeting_checks(merged_df)
        
        # Define the exact column order with additional columns in their specific positions
        final_columns = [
            # Initial columns
            'Line Item ID', 'Advertiser ID', 'Campaign ID',
            # Campaign budget related columns (additional)
            'campaign_budget_type', 'campaign_impressions_budget', 'Campaign_Imps_Required', 'Campaign Buffer %',
            # Brief data columns (additional)
            'brief_product_type', 'brief_dairy_milk_restrictions', 'brief_lda_compliant',
            'brief_viewability_contracted', 'brief_measurement_type', 'brief_viewability_goal',
            # Line item details
            'Line Item Name', 'Line Item Status', 'Line Item Type', 'Programmatic Guaranteed',
            'Start Date', 'End Date',
            # Budget related columns
            'Budget Type', 'Daily Budget', 'Budget', 'Budget Split %', 'Target Required', 'Line Buffer %',
            # Bidding and pacing
            'Bidding Strategy', 'Pacing', 'Pacing Behavior', 'Multiplier', 'Catchup Behavior',
            'Bidding Values', 'Bid Shading', 'Use Custom Bidding Agent',
            # Alt ID and brief data
            'Alt ID', 'brief_bvp_id', 'brief_platform_media', 'brief_impressions',
            'brief_geo_required', 'brief_traffic_info',
            # Original remaining columns
            'Notes', 'Flights', 'Flight Carry Over', 'Frequency Cap ID Type', 'Frequency Duration',
            'Use Fallback', 'Frequency Cap Vendor', 'Vendor Fees', 'Custom Event IDs',
            'Revenue Type', 'Revenue Amount', 'Enable SKAd Tracking', 'Target SKAd Enabled Supply',
            'SKAd Assignment Level', 'Ghost Bidding User ID Type', 'Test Group ID',
            'Include App Bundle List', 'Exclude App Bundle List', 'Include App ID List',
            'Exclude App ID List', 'Include App Name', 'Exclude App Name', 'Include Deal ID',
            'Exclude Deal ID', 'Include Deal ID List', 'Exclude Deal ID List',
            'Include Domain List ID', 'Exclude Domain List ID', 'Include Placement ID',
            'Exclude Placement ID', 'Include Placement ID List', 'Exclude Placement ID List',
            'Include Publisher ID', 'Exclude Publisher ID', 'Include Publisher ID List',
            'Exclude Publisher ID List', 'Include Site ID', 'Exclude Site ID',
            'Include Site ID List', 'Exclude Site ID List', 'Include Content Category',
            'Exclude Content Category', 'Include Content Genre', 'Exclude Content Genre',
            'Include Content Rating', 'Exclude Content Rating', 'Include Language',
            'Exclude Language', 'Include Ad Position', 'Exclude Ad Position',
            'Include Ads Txt', 'Exclude Ads Txt', 'Include Environment Type',
            'Exclude Environment Type', 'Include Interstitial', 'Exclude Interstitial',
            'Include Interstitial Type', 'Exclude Interstitial Type',
            'Include Native Layout', 'Exclude Native Layout', 'Include Rewarded',
            'Exclude Rewarded', 'Include Topframe', 'Exclude Topframe',
            'Include Video API', 'Exclude Video API', 'Include Auction Type',
            'Exclude Auction Type', 'Include Inventory Source', 'Exclude Inventory Source',
            'Include City', 'Exclude City', 'Include Country', 'Exclude Country',
            'Include Latitude & Longitude List', 'Exclude Latitude & Longitude List',
            'Include Latitude & Longitude Present', 'Exclude Latitude & Longitude Present',
            'Include Location Type', 'Exclude Location Type', 'Include Metro',
            'Exclude Metro', 'Include Region', 'Exclude Region',
            'Include Zip Code List', 'Exclude Zip Code List', 'Include Zip',
            'Exclude Zip', 'Include Bandwidth', 'Exclude Bandwidth',
            'Include Browser', 'Exclude Browser', 'Include Browser Version',
            'Exclude Browser Version', 'Include Carrier', 'Exclude Carrier',
            'Include Device Make', 'Exclude Device Make', 'Include Device Model',
            'Exclude Device Model', 'Include Device Screen Size',
            'Exclude Device Screen Size', 'Include Device Type', 'Exclude Device Type',
            'Include Operating System', 'Exclude Operating System',
            'Include Operating System Version', 'Exclude Operating System Version',
            'Include Time Of Week', 'Exclude Time Of Week',
            'Include User Time Of Week', 'Exclude User Time Of Week',
            'Include IP Address', 'Exclude IP Address', 'Include Segment',
            'Exclude Segment', 'Require Segment', 'Boolean Expression Segment',
            'Include User ID', 'Exclude User ID', 'Include Companion Required',
            'Exclude Companion Required', 'Include Playback Method',
            'Exclude Playback Method', 'Include Player Size', 'Exclude Player Size',
            'Include Video Start Delay', 'Exclude Video Start Delay',
            'Include Video Placement Type', 'Exclude Video Placement Type',
            'Include Audio Companion Required', 'Exclude Audio Companion Required',
            'Include Audio Start Delay', 'Exclude Audio Start Delay',
            'Bid Modifier ID', 'Bid Modifier Max Bid', 'Bid Modifier Min Bid',
            'Delivery Modifier ID', 'Delivery Modifier Model ID', 'Fallback Weight',
            'Creatives'
        ]
        
        # Add targeting check columns at the end
        targeting_check_cols = [
            'Country Check', 'App Bundle Check', 'Domain List Check',
            'Environment Type Check', 'OS/Device Check', 'Segment Check',
            'Creatives Check', 'Inventory Source Check', 'Content Category Check',
            'CTV Apps Check', 'Geo Targeting Check', 'Deal ID List Check',
            'Video Placement Type Check'  # Add the new check column
        ]
        final_columns.extend(targeting_check_cols)
        
        # Filter out any columns that don't exist in the DataFrame
        final_columns = [col for col in final_columns if col in merged_df.columns]
        
        # Reorder the DataFrame columns
        merged_df = merged_df[final_columns]
        
        # --- Write to Excel ---
        print(f"\nWriting {len(merged_df)} rows to Excel...")
        
        # Check for Budget Split % column
        if 'Budget Split %' in merged_df.columns:
            print(f"Budget Split % values: {merged_df['Budget Split %'].tolist()}")
            # Check Alt ID column
            alt_id_col = find_col_name(merged_df.columns, ['Alt ID', 'line_item_alternative_id'])
            if alt_id_col:
                print(f"Alt ID values: {merged_df[alt_id_col].tolist()}")
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Targeting Check Results"
        
        # Write header row
        ws.append(list(merged_df.columns))
        
        # Apply header formatting with different colors
        header_font = Font(bold=True)
        original_header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray
        additional_header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue
        
        # Define additional columns for different header color
        additional_cols = [
            'campaign_budget_type', 'campaign_impressions_budget', 'Campaign_Imps_Required', 'Campaign Buffer %',
            'brief_product_type', 'brief_dairy_milk_restrictions', 'brief_lda_compliant',
            'brief_viewability_contracted', 'brief_measurement_type', 'brief_viewability_goal',
            'Budget Split %', 'Target Required', 'Line Buffer %',
            'brief_bvp_id', 'brief_platform_media', 'brief_impressions',
            'brief_geo_required', 'brief_traffic_info'
        ]
        
        # Apply header formatting
        for idx, col in enumerate(merged_df.columns):
            cell = ws[1][idx]
            cell.font = header_font
            if col in additional_cols:
                cell.fill = additional_header_fill
            else:
                cell.fill = original_header_fill
                
        # Add data rows
        for r_idx, row in merged_df.iterrows():
            excel_row = []
            for col_name in merged_df.columns:
                value = row.get(col_name, '')
                if pd.isna(value):
                    excel_row.append('') 
                else:
                    excel_row.append(value)
            
            ws.append(excel_row)
        
        # Apply formatting
        print("Applying formatting and validations...")
        apply_formatting(wb, merged_df)
        
        # Save the file
        print(f"Saving workbook to {output_path}...")
        wb.save(output_path)
        if os.path.exists(output_path):
            print(f"Output file saved successfully to {output_path}")
            print(f"File size: {os.path.getsize(output_path)} bytes")
        else:
            print(f"Error: Failed to save file to {output_path}")
        
        print("\nTargeting Check finished.")
    except Exception as e:
        import traceback
        print(f"Error during targeting check: {str(e)}")
        print(traceback.format_exc())
        return None

if __name__ == "__main__":
    main() 